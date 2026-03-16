#!/usr/bin/env python3
"""
mobile_api.py - REST API for Symphony AI iOS App

Provides endpoints for:
- System status and health
- Quick actions (bids, proposals, invoices)
- Knowledge search
- Service status dashboard

Run: python3 mobile_api.py
Access: http://localhost:8420

For remote access via Tailscale:
  http://bob-mac-mini.tail1234.ts.net:8420
"""

import asyncio
import csv
import json
import os
import re
import shutil
import signal
import sqlite3
import subprocess
import sys
import tempfile
import time
import urllib.request
import zipfile
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional
from uuid import uuid4

from dotenv import load_dotenv

BASE_DIR = Path(__file__).resolve().parent.parent
load_dotenv(BASE_DIR / ".env")
sys.path.insert(0, str(BASE_DIR / "api"))
try:
    from api.common_api_utils import run_command as shared_run_command, run_tool_script
except ModuleNotFoundError:
    from common_api_utils import run_command as shared_run_command, run_tool_script

# FastAPI with fallback to simple HTTP server
try:
    from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
    from fastapi.middleware.cors import CORSMiddleware
    from fastapi.responses import FileResponse, JSONResponse
    from pydantic import BaseModel
    import uvicorn
    HAS_FASTAPI = True
except ImportError:
    HAS_FASTAPI = False
    print("FastAPI not installed. Run: pip install fastapi uvicorn")

try:
    import openpyxl  # type: ignore
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False

sys.path.insert(0, str(BASE_DIR))

API_PORT = int(os.environ.get("MOBILE_API_PORT", "8420"))
API_BIND_HOST = os.environ.get("MOBILE_API_BIND_HOST", "127.0.0.1")
def _normalize_auth_token(value: str) -> str:
    token = (value or "").strip()
    if len(token) >= 2 and token[0] == token[-1] and token[0] in {"'", '"'}:
        token = token[1:-1].strip()
    return token


def normalize_contact_identifier(value: str) -> str:
    raw = (value or "").strip()
    if not raw:
        return ""
    if "@" in raw:
        return raw.lower()
    digits = re.sub(r"\D", "", raw)
    if digits.startswith("1") and len(digits) == 11:
        digits = digits[1:]
    return digits or raw.lower()


API_AUTH_TOKEN = _normalize_auth_token(os.environ.get("SYMPHONY_API_TOKEN", ""))
DTOOLS_PRODUCT_AGENT_DIR = BASE_DIR / "data" / "dtools_product_agent"
DTOOLS_PRODUCT_AGENT_DIR.mkdir(parents=True, exist_ok=True)
DTOOLS_PRODUCT_DB = DTOOLS_PRODUCT_AGENT_DIR / "products.sqlite3"
MANUAL_DIGEST_DIR = BASE_DIR / "data" / "manual_digest"
MANUAL_DIGEST_DIR.mkdir(parents=True, exist_ok=True)
ASK_BOB_MEMORY_DIR = BASE_DIR / "data" / "ask_bob_memory"
ASK_BOB_MEMORY_DIR.mkdir(parents=True, exist_ok=True)
IMESSAGE_WATCHER_STATE_FILE = BASE_DIR / "data" / "imessage_watcher_state.json"
IMESSAGE_WORK_LOG_FILE = BASE_DIR / "knowledge" / "imessages" / "work_talk.jsonl"
CONTACTS_INDEX_FILE = BASE_DIR / "data" / "contacts" / "contacts_index.json"
CLIENTS_REGISTRY_FILE = BASE_DIR / "data" / "contacts" / "clients_registry.json"
NOTES_PROJECT_LINKS_FILE = BASE_DIR / "data" / "notes_project_links.json"
TASK_UPLOAD_DIR = BASE_DIR / "data" / "intake_uploads"
TASK_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
TASK_UPLOAD_QUEUE_FILE = TASK_UPLOAD_DIR / "uploads_queue.jsonl"
TASK_UPLOAD_FINDINGS_DIR = TASK_UPLOAD_DIR / "findings"
TASK_UPLOAD_FINDINGS_DIR.mkdir(parents=True, exist_ok=True)
TASK_UPLOAD_EXTRACTION_QUEUE_DIR = TASK_UPLOAD_DIR / "extraction_queue"
TASK_UPLOAD_EXTRACTION_QUEUE_DIR.mkdir(parents=True, exist_ok=True)
PROJECT_WATCHES_FILE = TASK_UPLOAD_DIR / "project_watches.json"
PROJECT_WATCH_STATE_FILE = TASK_UPLOAD_DIR / "project_watch_state.json"
PROJECT_WATCH_LOG_DIR = TASK_UPLOAD_DIR / "project_watch_logs"
PROJECT_WATCH_LOG_DIR.mkdir(parents=True, exist_ok=True)
EMPLOYEE_BOTS_FILE = BASE_DIR / "data" / "employee_bots.json"
EMPLOYEE_BOTS_RUNTIME_DIR = BASE_DIR / "data" / "employee_bots_runtime"
EMPLOYEE_BOTS_RUNTIME_DIR.mkdir(parents=True, exist_ok=True)
TASK_BOARD_DB = BASE_DIR / "orchestrator" / "task_board.db"
KNOWLEDGE_PROJECTS_DIR = BASE_DIR / "knowledge" / "projects"
KNOWLEDGE_PROJECTS_DIR.mkdir(parents=True, exist_ok=True)
PROJECT_WATCH_INTERVAL_SEC = int(os.environ.get("PROJECT_WATCH_INTERVAL_SEC", "300"))
PROJECT_WATCH_AUTORUN = os.environ.get("PROJECT_WATCH_AUTORUN", "1").strip().lower() not in {"0", "false", "no"}
PROJECT_WATCH_AUTO_DISCOVER = os.environ.get("PROJECT_WATCH_AUTO_DISCOVER", "1").strip().lower() not in {"0", "false", "no"}
PROJECT_WATCH_PROJECTS_ROOT = os.environ.get(
    "PROJECT_WATCH_PROJECTS_ROOT",
    str(Path.home() / "Library/Mobile Documents/com~apple~CloudDocs/Symphony SH/Projects"),
).strip()
PROJECT_WATCH_MAX_FILE_BYTES = int(os.environ.get("PROJECT_WATCH_MAX_FILE_BYTES", str(200 * 1024 * 1024)))
PROJECT_WATCH_STALE_SCAN_MIN = int(os.environ.get("PROJECT_WATCH_STALE_SCAN_MIN", "30"))
PROJECT_WATCH_IGNORE_DIR_PATTERNS = tuple(
    p.strip()
    for p in os.environ.get(
        "PROJECT_WATCH_IGNORE_DIR_PATTERNS",
        "__MACOSX,.git,.cursor,node_modules,.venv,venv,.idea,DerivedData",
    ).split(",")
    if p.strip()
)
PROJECT_WATCH_IGNORE_FILE_PATTERNS = tuple(
    p.strip().lower()
    for p in os.environ.get(
        "PROJECT_WATCH_IGNORE_FILE_PATTERNS",
        ".ds_store,thumbs.db,desktop.ini,.tmp,.swp,.crdownload,.part,.icloud",
    ).split(",")
    if p.strip()
)
PDF_EXTRACT_MAX_PAGES = int(os.environ.get("PDF_EXTRACT_MAX_PAGES", "80"))
PDF_EXTRACT_MAX_CHARS = int(os.environ.get("PDF_EXTRACT_MAX_CHARS", "300000"))
MANUALS_AUTO_DIGEST_ENABLED = os.environ.get("MANUALS_AUTO_DIGEST_ENABLED", "1").strip().lower() not in {"0", "false", "no"}
MANUALS_LIBRARY_ROOT = Path(
    os.environ.get(
        "MANUALS_LIBRARY_ROOT",
        str(Path.home() / "Library/Mobile Documents/com~apple~CloudDocs/Symphony SH/Manuals"),
    )
).expanduser()
MANUALS_AUTO_DIGEST_STATE_FILE = MANUAL_DIGEST_DIR / "manuals_auto_digest_state.json"
MANUALS_AUTO_DIGEST_KNOWLEDGE_FILE = BASE_DIR / "knowledge" / "cortex" / "company" / "manuals_auto_digest.jsonl"
MANUALS_AUTO_DIGEST_KNOWLEDGE_FILE.parent.mkdir(parents=True, exist_ok=True)
MANUALS_AUTO_DIGEST_MAX_FILE_BYTES = int(os.environ.get("MANUALS_AUTO_DIGEST_MAX_FILE_BYTES", str(200 * 1024 * 1024)))
PROJECT_WATCH_TASK: Optional[asyncio.Task] = None
NETWORK_WATCH_DIR = BASE_DIR / "data" / "network_watch"
NETWORK_WATCH_DIR.mkdir(parents=True, exist_ok=True)
NETWORK_WATCH_PID_FILE = NETWORK_WATCH_DIR / "dropout_watch.pid"
NETWORK_WATCH_STATUS_FILE = NETWORK_WATCH_DIR / "dropout_watch_status.json"
NETWORK_WATCH_EVENTS_FILE = NETWORK_WATCH_DIR / "dropout_watch_events.jsonl"
NETWORK_WATCH_STDOUT = BASE_DIR / "logs" / "dropout_watch.log"
NETWORK_WATCH_STDERR = BASE_DIR / "logs" / "dropout_watch_error.log"
INVENTORY_REPORTS_DIR = BASE_DIR / "knowledge" / "reports"
INVENTORY_REPORTS_DIR.mkdir(parents=True, exist_ok=True)
INVENTORY_STOCK_DIR = BASE_DIR / "data" / "inventory"
INVENTORY_STOCK_DIR.mkdir(parents=True, exist_ok=True)
INVENTORY_STOCK_FILE = INVENTORY_STOCK_DIR / "stock_levels.json"
IMESSAGE_AUTOMATION_DIR = BASE_DIR / "data" / "imessage_automation"
IMESSAGE_AUTOMATION_DIR.mkdir(parents=True, exist_ok=True)
IMESSAGE_INVOICE_DRAFTS_FILE = IMESSAGE_AUTOMATION_DIR / "service_invoice_drafts.jsonl"
IMESSAGE_APPOINTMENT_DRAFTS_FILE = IMESSAGE_AUTOMATION_DIR / "appointment_schedule_drafts.jsonl"
IMESSAGE_INTAKE_AUDIT_FILE = IMESSAGE_AUTOMATION_DIR / "intake_actions_audit.jsonl"
IMESSAGE_INTAKE_FAILURES_FILE = IMESSAGE_AUTOMATION_DIR / "intake_failures.json"
MESSAGES_DB_PATH = Path.home() / "Library" / "Messages" / "chat.db"

PARSE_PROFILES: Dict[str, Dict[str, Any]] = {
    "auto": {
        "label": "Auto detect",
        "expected_columns": [],
    },
    "msrp_three_tiers": {
        "label": "MSRP + Standard/Silver/Gold",
        "expected_columns": [
            "MODEL NAME",
            "PART NUMBER",
            "SKU",
            "DESCRIPTION",
            "MSRP",
            "STANDARD DEALER",
            "SILVER DEALER",
            "GOLD DEALER",
        ],
    },
    "msrp_standard_only": {
        "label": "MSRP + Standard dealer",
        "expected_columns": [
            "MODEL NAME",
            "PART NUMBER",
            "SKU",
            "DESCRIPTION",
            "MSRP",
            "STANDARD DEALER",
        ],
    },
    "minimal": {
        "label": "Minimal (model, part, description, price)",
        "expected_columns": [
            "MODEL",
            "PART NUMBER",
            "DESCRIPTION",
            "PRICE",
        ],
    },
    "invoice_rexel": {
        "label": "Invoice (Rexel / electrical order)",
        "expected_columns": [
            "ITEM NUMBER",
            "MANUFACTURER",
            "SKU",
            "DESCRIPTION",
            "QUANTITY",
            "SHIP QTY",
            "UNIT PRICE",
            "EXTENDED PRICE",
            "UPC",
        ],
    },
}

KNOWN_DEVICE_BRANDS = [
    "Control4",
    "Lutron",
    "Sonos",
    "Araknis",
    "Episode",
    "WattBox",
    "Luma",
    "Pakedge",
    "Snap One",
    "Modern Atomics",
    "Josh.ai",
    "Apple TV",
    "Samsung",
    "Sony",
    "Yamaha",
]


def _slugify(value: str, fallback: str = "na", max_len: int = 48) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9]+", "-", (value or "").strip().lower()).strip("-")
    if not cleaned:
        cleaned = fallback
    return cleaned[:max_len]


def _safe_filename_segment(value: str, fallback: str, max_len: int = 64) -> str:
    # Keep human-readable capitalization/spaces; strip filename-unsafe chars.
    cleaned = re.sub(r'[\\/:*?"<>|]+', "", (value or "").strip())
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" .-_")
    if not cleaned:
        cleaned = fallback
    return cleaned[:max_len]


def _categorize_bundle_file(filename: str) -> str:
    ext = Path(filename).suffix.lower()
    if ext in {".dwg"}:
        return "drawing"
    if ext in {".png", ".jpg", ".jpeg", ".webp", ".gif", ".heic", ".tif", ".tiff"}:
        return "image"
    if ext in {".xlsx", ".xls", ".csv"}:
        return "proposal"
    if ext in {".pdf"}:
        return "drawing"
    return "document"


def _safe_extract_zip(zip_path: Path, dest_dir: Path, max_files: int = 250) -> List[Path]:
    extracted: List[Path] = []
    root = dest_dir.resolve()
    with zipfile.ZipFile(zip_path, "r") as zf:
        for info in zf.infolist():
            if len(extracted) >= max_files:
                break
            if info.is_dir():
                continue
            member = (info.filename or "").replace("\\", "/").strip("/")
            if not member or member.startswith("__MACOSX/") or member.endswith(".DS_Store"):
                continue
            target = (dest_dir / member).resolve()
            if not str(target).startswith(str(root) + os.sep):
                continue
            target.parent.mkdir(parents=True, exist_ok=True)
            with zf.open(info, "r") as src, target.open("wb") as out:
                shutil.copyfileobj(src, out)
            extracted.append(target)
    return extracted


def _build_intake_filename(
    category: str,
    project_name: str,
    client_name: str,
    original_filename: str,
    address_line: str = "",
    location_name: str = "",
    discipline: str = "",
    sheet_number: str = "",
    revision: str = "",
    issue_date: str = "",
) -> str:
    def _client_last_name(value: str) -> str:
        tokens = re.findall(r"[A-Za-z0-9]+", value or "")
        return tokens[-1] if tokens else ""

    def _address_number_street(value: str, fallback_project: str) -> str:
        source = value or fallback_project or ""
        match = re.search(
            r"(\d{1,6}\s+[A-Za-z0-9.\-']+(?:\s+[A-Za-z0-9.\-']+){0,4})",
            source,
            flags=re.IGNORECASE,
        )
        return match.group(1) if match else ""

    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    client_last = _safe_filename_segment(_client_last_name(client_name), fallback="Client", max_len=28)
    addr = _safe_filename_segment(_address_number_street(address_line, project_name), fallback="Address", max_len=48)
    location = _safe_filename_segment(location_name, fallback="Location", max_len=32)
    cat = _slugify(category, fallback="document", max_len=16)
    orig_stem = _slugify(Path(original_filename or "upload").stem, fallback="file", max_len=28)
    disc = _slugify(discipline, fallback="na", max_len=12)
    sheet = _slugify(sheet_number, fallback="na", max_len=12)
    rev = _slugify(revision, fallback="na", max_len=6)
    issue = _slugify(issue_date, fallback=ts.split("-")[0], max_len=12)
    ext = (Path(original_filename or "").suffix or ".bin").lower()
    prefix = f"{client_last} - {addr} - {location}"
    return f"{prefix} -- {cat} -- {disc} -- {sheet} -- r-{rev} -- {issue} -- {orig_stem}{ext}"


def _load_markup_symbol_catalog() -> List[Dict[str, str]]:
    """
    Build a lightweight symbol catalog from SYMBOL_SPEC and symbols.svg.
    Returns rows like: {"id": "spk-z1", "label": "S 1", "sku": "TS-IC62"}.
    """
    catalog: List[Dict[str, str]] = []
    spec_path = BASE_DIR / "knowledge" / "symbols" / "SYMBOL_SPEC.md"
    svg_path = BASE_DIR / "tools" / "markup_app" / "symbols" / "symbols.svg"

    if spec_path.exists():
        for line in spec_path.read_text(encoding="utf-8", errors="ignore").splitlines():
            stripped = line.strip()
            if not stripped.startswith("| `"):
                continue
            parts = [p.strip() for p in stripped.strip("|").split("|")]
            if len(parts) < 6:
                continue
            sym_id = parts[0].strip("`")
            label = parts[1]
            sku = parts[5]
            if sym_id and sym_id.lower() != "id":
                catalog.append({"id": sym_id, "label": label, "sku": sku})

    if svg_path.exists():
        svg_text = svg_path.read_text(encoding="utf-8", errors="ignore")
        for match in re.findall(r'id="sym-([a-zA-Z0-9\-]+)"', svg_text):
            if not any(c.get("id") == match for c in catalog):
                catalog.append({"id": match, "label": match.replace("-", " ").upper(), "sku": ""})

    return catalog


def _normalize_room_name(value: str) -> str:
    name = re.sub(r"\s+", " ", (value or "").strip())
    if not name:
        return ""
    if len(name) < 2 or len(name) > 80:
        return ""
    lowered = name.lower()
    blocked = {
        "name",
        "room",
        "rooms",
        "page",
        "pages",
        "project",
        "markup",
        "symbol",
        "symbols",
        "placed",
        "currentpage",
    }
    if lowered in blocked:
        return ""
    return name


def _room_model_profile(system_profile: str) -> Dict[str, List[str]]:
    profile = (system_profile or "control4").strip().lower()
    if profile == "lutron":
        return {
            "scope": [
                "Lighting load map and keypad engravings",
                "Shade groups and scene assignments",
                "Processor/network/IP planning",
            ],
            "checks": [
                "Confirm fixture counts and dimming compatibility",
                "Confirm keypad locations and gang depth",
            ],
        }
    if profile == "sonos":
        return {
            "scope": [
                "Audio zone map and amp/speaker assignment",
                "Network drops and PoE budget review",
                "Rack power and thermal planning",
            ],
            "checks": [
                "Validate speaker wire home-runs per zone",
                "Confirm LAN wiring for Sonos core devices",
            ],
        }
    if profile == "hybrid":
        return {
            "scope": [
                "Control4 automation + Lutron lighting integration plan",
                "Per-room AV/control points and keypad strategy",
                "Rack/core network and power design",
            ],
            "checks": [
                "Confirm protocol boundaries (LEAP, mDNS, IGMP)",
                "Validate room-by-room control surface consistency",
            ],
        }
    return {
        "scope": [
            "Control4 room endpoint and control-surface plan",
            "AV/network prewire checklist by room",
            "Rack/core dependency and driver notes",
        ],
        "checks": [
            "Confirm hardwired TV/audio endpoints",
            "Confirm control keypad/touchpanel locations",
        ],
    }


def _extract_rooms_from_symphony_markup(payload: Dict[str, Any]) -> Dict[str, Any]:
    rooms_map: Dict[str, Dict[str, Any]] = {}

    pages = payload.get("pages")
    if isinstance(pages, dict):
        for page_key, page_data in pages.items():
            if not isinstance(page_data, dict):
                continue
            room_entries = page_data.get("rooms")
            if isinstance(room_entries, list):
                for entry in room_entries:
                    if isinstance(entry, dict):
                        raw_name = str(entry.get("name") or entry.get("room") or entry.get("label") or "").strip()
                    else:
                        raw_name = str(entry or "").strip()
                    room_name = _normalize_room_name(raw_name)
                    if not room_name:
                        continue
                    info = rooms_map.setdefault(
                        room_name,
                        {"room": room_name, "pages": set(), "symbol_count": 0},
                    )
                    info["pages"].add(str(page_key))

            placed = page_data.get("placed")
            if isinstance(placed, list):
                for symbol in placed:
                    if not isinstance(symbol, dict):
                        continue
                    raw_name = str(
                        symbol.get("room")
                        or symbol.get("roomName")
                        or symbol.get("area")
                        or symbol.get("areaName")
                        or symbol.get("label")
                        or ""
                    ).strip()
                    room_name = _normalize_room_name(raw_name)
                    if not room_name:
                        continue
                    info = rooms_map.setdefault(
                        room_name,
                        {"room": room_name, "pages": set(), "symbol_count": 0},
                    )
                    info["pages"].add(str(page_key))
                    info["symbol_count"] += 1

    if not rooms_map:
        text = json.dumps(payload, ensure_ascii=False)
        candidates = re.findall(
            r'"(?:room|roomName|area|areaName|name)"\s*:\s*"([^"]{2,80})"',
            text,
            flags=re.IGNORECASE,
        )
        for cand in candidates:
            room_name = _normalize_room_name(cand)
            if not room_name:
                continue
            rooms_map.setdefault(room_name, {"room": room_name, "pages": set(), "symbol_count": 0})

    rooms: List[Dict[str, Any]] = []
    for name in sorted(rooms_map.keys()):
        info = rooms_map[name]
        pages_list = sorted(list(info.get("pages", set())))
        rooms.append(
            {
                "room": name,
                "pages": pages_list,
                "symbol_count": int(info.get("symbol_count", 0)),
            }
        )
    return {"rooms": rooms, "room_count": len(rooms)}


def _extract_text_from_dwg(path: Path) -> str:
    """
    Best-effort DWG text extraction without heavyweight dependencies.
    Uses `strings` to surface embedded note text and block names.
    """
    try:
        result = subprocess.run(
            ["strings", str(path)],
            capture_output=True,
            text=True,
            timeout=25,
        )
        if result.returncode == 0:
            return result.stdout
    except Exception:
        pass
    return ""


def _extract_legend_terms(text: str) -> List[str]:
    """
    Heuristic legend extraction:
    - pulls terms from lines containing LEGEND/SYMBOL
    - includes known shorthand tokens common in AV drawings
    """
    terms: List[str] = []
    for line in text.splitlines():
        up = line.upper().strip()
        if not up:
            continue
        if "LEGEND" in up or "SYMBOL" in up:
            terms.extend(re.findall(r"[A-Z0-9\-]{2,}", up))
            continue
        if any(tok in up for tok in [" AP ", " TV ", " KP ", " DIM ", " CAM ", " DEMARC ", " DATA "]):
            terms.extend(re.findall(r"[A-Z0-9\-]{2,}", up))

    cleaned = []
    for t in terms:
        if t in {"LEGEND", "SYMBOL", "SCHEDULE", "NOTES"}:
            continue
        if len(t) < 2:
            continue
        cleaned.append(t)
    return sorted(set(cleaned))[:200]


def _extract_dtools_qv(text: str, fallback_text: str = "") -> Dict[str, str]:
    """
    Parse D-Tools quote/version markers such as:
    - Q-1234|V2
    - Q-1234 V2
    - Quote Q-1234 Version 2
    """
    combined = f"{text or ''}\n{fallback_text or ''}".upper()
    normalized = (
        combined.replace("I", "1")
        .replace("L", "1")
        .replace("O", "0")
        .replace("S", "5")
    )
    quote_id = ""
    version = ""

    # Prefer explicit quote/version pair candidates and choose best by frequency + quote length.
    pair_candidates = re.findall(r"\bQ[-\s]?(\d{1,8})\s*(?:\||\s)\s*V?(\d{1,2})\b", normalized)
    if pair_candidates:
        max_len = max(len(q_digits) for q_digits, _ in pair_candidates)
        score: Dict[str, int] = {}
        q_to_versions: Dict[str, List[str]] = {}
        for q_digits, v_digits in pair_candidates:
            if len(q_digits) < max_len:
                continue
            q = f"Q-{q_digits}"
            v = f"V{v_digits}"
            score[q] = score.get(q, 0) + 50 + len(q_digits) * 10
            q_to_versions.setdefault(q, []).append(v)
        # Pick highest score; tie-break by longer numeric part.
        quote_id = sorted(score.keys(), key=lambda q: (score[q], len(q.split("-")[-1])), reverse=True)[0]
        versions = q_to_versions.get(quote_id, [])
        if versions:
            version = sorted(set(versions), key=lambda v: versions.count(v), reverse=True)[0]
    else:
        q_matches = re.findall(r"\bQ[-\s]?(\d{1,8})\b", normalized)
        v_matches = re.findall(r"\bV(\d{1,2})\b", normalized)
        if q_matches:
            # Prefer longer quote IDs; then most frequent.
            q_freq: Dict[str, int] = {}
            for qd in q_matches:
                q_freq[qd] = q_freq.get(qd, 0) + 1
            best_q = sorted(q_freq.keys(), key=lambda qd: (len(qd), q_freq[qd]), reverse=True)[0]
            quote_id = f"Q-{best_q}"
        if v_matches:
            v_freq: Dict[str, int] = {}
            for vd in v_matches:
                v_freq[vd] = v_freq.get(vd, 0) + 1
            best_v = sorted(v_freq.keys(), key=lambda vd: (v_freq[vd], len(vd)), reverse=True)[0]
            version = f"V{best_v}"

    # Accept compact/boxed style: "Q-195 4" or "Q-195|4"
    if quote_id and not version:
        compact = re.search(r"\bQ[-\s]?\d{1,8}\s*(?:\||\s)\s*(?:V)?(\d{1,2})\b", normalized)
        if compact:
            version = f"V{compact.group(1)}"

    # Fallback for "Version 2" formatting
    if not version:
        v_word = re.search(r"\bVER(?:SION)?\s*#?\s*(\d{1,2})\b", normalized)
        if v_word:
            version = f"V{v_word.group(1)}"

    return {
        "quote_id": quote_id,
        "version": version,
        "quote_version": f"{quote_id}|{version}" if quote_id and version else "",
    }


def _ocr_titleblock_qv_from_image(path: Path) -> str:
    """
    Best-effort OCR focused on title-block top-right region where Q/V appears.
    Requires Pillow + tesseract binary. Returns extracted text (possibly empty).
    """
    if shutil.which("tesseract") is None:
        return ""
    try:
        from PIL import Image, ImageOps  # type: ignore
    except Exception:
        return ""

    try:
        with Image.open(path) as img:
            img = img.convert("RGB")
            w, h = img.size

            # Region where user said Q/V always lives: left of logo, top-right banner.
            regions = [
                (int(w * 0.60), int(h * 0.02), int(w * 0.84), int(h * 0.22)),
                (int(w * 0.55), int(h * 0.00), int(w * 0.90), int(h * 0.30)),
            ]

            text_chunks: List[str] = []
            for idx, box in enumerate(regions):
                crop = img.crop(box)
                gray = ImageOps.grayscale(crop)
                boosted = ImageOps.autocontrast(gray)
                upscaled = boosted.resize((boosted.width * 3, boosted.height * 3))
                bw = upscaled.point(lambda p: 255 if p > 135 else 0)
                with tempfile.NamedTemporaryFile(suffix=f"_qv_{idx}.png", delete=False) as tf:
                    tmp_path = Path(tf.name)
                try:
                    bw.save(tmp_path, format="PNG")
                    for psm in ("7", "6", "11"):
                        result = subprocess.run(
                            [
                                "tesseract",
                                str(tmp_path),
                                "stdout",
                                "--psm",
                                psm,
                                "-l",
                                "eng",
                                "-c",
                                "tessedit_char_whitelist=QV0123456789|- ",
                            ],
                            capture_output=True,
                            text=True,
                            timeout=20,
                        )
                        if result.returncode == 0 and result.stdout:
                            text_chunks.append(result.stdout)
                finally:
                    try:
                        tmp_path.unlink(missing_ok=True)
                    except Exception:
                        pass
            return "\n".join(text_chunks)
    except Exception:
        return ""


def _match_legend_to_symbols(legend_terms: List[str], symbol_catalog: List[Dict[str, str]]) -> List[Dict[str, str]]:
    matches: List[Dict[str, str]] = []
    for term in legend_terms:
        term_n = term.lower().replace("_", " ").replace("-", " ")
        for sym in symbol_catalog:
            sid = (sym.get("id") or "").lower().replace("-", " ")
            label = (sym.get("label") or "").lower()
            if not sid and not label:
                continue
            if term_n == sid or term_n == label or term_n in sid or term_n in label:
                matches.append(
                    {
                        "legend_term": term,
                        "symbol_id": sym.get("id", ""),
                        "symbol_label": sym.get("label", ""),
                        "dtools_sku": sym.get("sku", ""),
                    }
                )
                break
    # De-dupe by legend_term + symbol_id
    seen = set()
    dedup = []
    for m in matches:
        key = (m["legend_term"], m["symbol_id"])
        if key in seen:
            continue
        seen.add(key)
        dedup.append(m)
    return dedup[:200]


def _first_pass_drawing_findings(stored_path: Path, project_name: str, client_name: str) -> Dict[str, Any]:
    text = ""
    suffix = stored_path.suffix.lower()
    if suffix == ".pdf":
        text = _extract_text_from_pdf(stored_path)
    elif suffix == ".dwg":
        text = _extract_text_from_dwg(stored_path)
    elif suffix in {".png", ".jpg", ".jpeg", ".webp"}:
        text = _ocr_titleblock_qv_from_image(stored_path)
    else:
        text = _extract_text_from_path(stored_path)

    symbol_catalog = _load_markup_symbol_catalog()
    legend_terms = _extract_legend_terms(text)
    symbol_matches = _match_legend_to_symbols(legend_terms, symbol_catalog)
    sheet_refs = sorted(set(re.findall(r"\b[A-Z]{1,3}\d{1,2}(?:\.\d{1,2})?\b", text.upper())))
    dtools_qv = _extract_dtools_qv(text=text, fallback_text=stored_path.name)

    findings = {
        "generated_at": datetime.now().isoformat(),
        "project_name": project_name,
        "client_name": client_name,
        "file_name": stored_path.name,
        "file_path": str(stored_path),
        "file_type": suffix.lstrip("."),
        "text_char_count": len(text),
        "sheet_references": sheet_refs[:200],
        "legend_terms": legend_terms[:200],
        "symbol_matches": symbol_matches,
        "dtools_quote_id": dtools_qv.get("quote_id", ""),
        "dtools_version": dtools_qv.get("version", ""),
        "dtools_quote_version": dtools_qv.get("quote_version", ""),
        "markup_symbol_catalog_size": len(symbol_catalog),
    }
    return findings


def _load_project_watches() -> List[Dict[str, Any]]:
    try:
        if not PROJECT_WATCHES_FILE.exists():
            return []
        raw = json.loads(PROJECT_WATCHES_FILE.read_text(encoding="utf-8"))
        if isinstance(raw, list):
            return [x for x in raw if isinstance(x, dict)]
    except Exception:
        pass
    return []


def _save_project_watches(watches: List[Dict[str, Any]]) -> None:
    PROJECT_WATCHES_FILE.write_text(json.dumps(watches, indent=2), encoding="utf-8")


def _load_project_watch_state() -> Dict[str, Any]:
    try:
        if not PROJECT_WATCH_STATE_FILE.exists():
            return {}
        raw = json.loads(PROJECT_WATCH_STATE_FILE.read_text(encoding="utf-8"))
        if isinstance(raw, dict):
            return raw
    except Exception:
        pass
    return {}


def _save_project_watch_state(state: Dict[str, Any]) -> None:
    PROJECT_WATCH_STATE_FILE.write_text(json.dumps(state, indent=2), encoding="utf-8")


def _project_slug(project_name: str, client_name: str, location_name: str) -> str:
    if project_name.strip():
        return _slugify(project_name, fallback="project", max_len=64)
    combined = " ".join([client_name.strip(), location_name.strip()]).strip()
    return _slugify(combined, fallback="project", max_len=64)


def _metadata_from_project_folder_name(folder_name: str) -> Dict[str, str]:
    """
    Parse expected naming style:
      <Client Last Name>-<Address(#+street)>-<Location>
    Accepts minor spacing inconsistencies around hyphens.
    """
    cleaned = re.sub(r"\s+", " ", (folder_name or "").strip())
    parts = [p.strip() for p in re.split(r"\s*-\s*", cleaned) if p.strip()]
    if len(parts) >= 3:
        client_last = parts[0]
        address = parts[1]
        location = " - ".join(parts[2:])
        project_name = f"{client_last} - {address}"
        return {
            "project_name": project_name,
            "client_name": client_last,
            "address_line": address,
            "location_name": location,
        }
    return {
        "project_name": cleaned,
        "client_name": "",
        "address_line": "",
        "location_name": "",
    }


def _should_ignore_watch_path(rel_path: str, filename: str) -> bool:
    rel_lower = rel_path.lower()
    file_lower = filename.lower()
    for pat in PROJECT_WATCH_IGNORE_DIR_PATTERNS:
        token = pat.lower().strip()
        if not token:
            continue
        if rel_lower.startswith(token + "/") or f"/{token}/" in rel_lower:
            return True
    for pat in PROJECT_WATCH_IGNORE_FILE_PATTERNS:
        token = pat.lower().strip()
        if not token:
            continue
        if token.startswith(".") and file_lower.endswith(token):
            return True
        if file_lower == token:
            return True
    return False


def _extract_project_signals_from_text(text: str) -> Dict[str, List[str]]:
    upper = (text or "").upper()
    rfi_matches = re.findall(r"\bRFI[-\s:]?\d{1,4}\b", upper)
    wants_matches = re.findall(r"(?:CLIENT\s+WANTS?|WANTS/NEEDS|WANTS|NEEDS)", upper)
    scope_matches = re.findall(r"(?:CHANGE\s+ORDER|ALLOWANCE|EXCLUSION|ASSUMPTION)", upper)
    return {
        "rfi_tags": sorted(set(rfi_matches))[:50],
        "wants_needs_tags": sorted(set(wants_matches))[:50],
        "scope_risk_tags": sorted(set(scope_matches))[:50],
    }


def _load_manuals_auto_digest_state() -> Dict[str, Any]:
    try:
        if not MANUALS_AUTO_DIGEST_STATE_FILE.exists():
            return {}
        raw = json.loads(MANUALS_AUTO_DIGEST_STATE_FILE.read_text(encoding="utf-8"))
        if isinstance(raw, dict):
            return raw
    except Exception:
        pass
    return {}


def _save_manuals_auto_digest_state(state: Dict[str, Any]) -> None:
    MANUALS_AUTO_DIGEST_STATE_FILE.write_text(json.dumps(state, indent=2), encoding="utf-8")


def _manuals_auto_digest_once(max_new_files: int = 25) -> Dict[str, Any]:
    root = MANUALS_LIBRARY_ROOT.resolve()
    if not root.exists() or not root.is_dir():
        return {"success": False, "error": f"Manuals root missing: {root}"}

    state = _load_manuals_auto_digest_state()
    seen: Dict[str, str] = state.get("files", {})
    supported = {".pdf", ".txt", ".md", ".doc", ".docx", ".xlsx", ".xls", ".csv", ".png", ".jpg", ".jpeg", ".webp"}

    candidates: List[Path] = []
    for p in root.rglob("*"):
        if not p.is_file():
            continue
        rel = str(p.relative_to(root))
        if rel.startswith(".") or "/." in rel or p.name == ".DS_Store":
            continue
        if p.suffix.lower() not in supported:
            continue
        try:
            st = p.stat()
            if st.st_size > MANUALS_AUTO_DIGEST_MAX_FILE_BYTES:
                continue
            sig = f"{st.st_size}:{int(st.st_mtime)}"
        except Exception:
            continue
        if seen.get(rel) != sig:
            seen[rel] = sig
            candidates.append(p)

    candidates = sorted(candidates, key=lambda x: x.stat().st_mtime)[: max(1, min(max_new_files, 300))]
    processed = 0
    errors = 0
    for p in candidates:
        try:
            text = _extract_text_from_path(p)
            signals = _extract_manual_digest_signals(text) if text else {}
            row = {
                "timestamp": datetime.now().isoformat(),
                "source": "manuals_auto_digest",
                "manual_root": str(root),
                "file_path": str(p),
                "file_name": p.name,
                "text_chars": len(text),
                "signals": signals,
            }
            with MANUALS_AUTO_DIGEST_KNOWLEDGE_FILE.open("a", encoding="utf-8") as fh:
                fh.write(json.dumps(row) + "\n")
            processed += 1
        except Exception:
            errors += 1

    state["files"] = seen
    state["last_run_at"] = datetime.now().isoformat()
    state["last_processed"] = processed
    state["last_errors"] = errors
    state["root"] = str(root)
    _save_manuals_auto_digest_state(state)
    return {
        "success": True,
        "root": str(root),
        "processed": processed,
        "errors": errors,
        "candidate_count": len(candidates),
        "knowledge_file": str(MANUALS_AUTO_DIGEST_KNOWLEDGE_FILE),
        "last_run_at": state["last_run_at"],
    }


def _update_project_intelligence_summary(project_slug: str) -> Optional[Path]:
    ingest_dir = KNOWLEDGE_PROJECTS_DIR / project_slug / "ingest"
    if not ingest_dir.exists():
        return None
    records: List[Dict[str, Any]] = []
    for p in sorted(ingest_dir.glob("*.json"), key=lambda x: x.stat().st_mtime)[-300:]:
        try:
            records.append(json.loads(p.read_text(encoding="utf-8")))
        except Exception:
            continue
    if not records:
        return None

    category_counts: Dict[str, int] = {}
    rfi_tags: set[str] = set()
    wants_tags: set[str] = set()
    scope_risk_tags: set[str] = set()
    recent_files: List[Dict[str, Any]] = []

    for rec in records:
        cat = (rec.get("category", "unknown") or "unknown").lower()
        category_counts[cat] = category_counts.get(cat, 0) + 1
        stored_path = rec.get("stored_path", "")
        stored_name = Path(stored_path).name if stored_path else ""
        if stored_name:
            recent_files.append(
                {
                    "file_name": stored_name,
                    "category": cat,
                    "ingested_at": rec.get("ingested_at", ""),
                }
            )

        # Signal extraction for text-like documents and PDFs.
        try:
            suffix = Path(stored_path).suffix.lower()
            if Path(stored_path).exists():
                text = ""
                if suffix in {".txt", ".md", ".csv", ".json"}:
                    text = Path(stored_path).read_text(encoding="utf-8", errors="ignore")[:120000]
                elif suffix == ".pdf":
                    text = _extract_text_from_pdf(Path(stored_path))[:120000]
                if not text:
                    continue
                signals = _extract_project_signals_from_text(text)
                rfi_tags.update(signals["rfi_tags"])
                wants_tags.update(signals["wants_needs_tags"])
                scope_risk_tags.update(signals["scope_risk_tags"])
        except Exception:
            pass

    summary = {
        "generated_at": datetime.now().isoformat(),
        "project_slug": project_slug,
        "total_ingested_files": len(records),
        "category_counts": category_counts,
        "signals": {
            "rfi_tags": sorted(rfi_tags)[:100],
            "wants_needs_tags": sorted(wants_tags)[:100],
            "scope_risk_tags": sorted(scope_risk_tags)[:100],
        },
        "recent_files": recent_files[-25:],
        "recommended_next_actions": [
            "Review files with RFI tags and create/merge formal RFI tasks.",
            "Confirm client wants/needs mentions against proposal scope.",
            "Flag scope risk tags (allowance/exclusion/assumption/change order) for PM review.",
        ],
    }
    out_path = KNOWLEDGE_PROJECTS_DIR / project_slug / "project_intelligence_summary.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")
    return out_path


def _backfill_recent_watch_files(hours: int = 24, max_files: int = 250) -> Dict[str, Any]:
    cutoff = datetime.now().timestamp() - max(1, hours) * 3600
    watches = [w for w in _load_project_watches() if w.get("enabled", True)]
    if not watches:
        return {"watches": 0, "candidates": 0, "processed": 0}

    existing_source_paths: set[str] = set()
    if TASK_UPLOAD_QUEUE_FILE.exists():
        for line in TASK_UPLOAD_QUEUE_FILE.read_text(encoding="utf-8", errors="ignore").splitlines():
            line = line.strip()
            if not line:
                continue
            try:
                row = json.loads(line)
            except Exception:
                continue
            src = (row.get("source_path") or "").strip()
            if src:
                existing_source_paths.add(src)

    candidates: List[tuple[float, Dict[str, Any], Path]] = []
    for watch in watches:
        folder = Path(watch.get("folder_path", "")).expanduser().resolve()
        if not folder.exists() or not folder.is_dir():
            continue
        for p in folder.rglob("*"):
            if not p.is_file():
                continue
            rel = str(p.relative_to(folder))
            if rel.startswith(".") or "/." in rel or _should_ignore_watch_path(rel, p.name):
                continue
            try:
                st = p.stat()
                if st.st_mtime < cutoff:
                    continue
                if st.st_size > PROJECT_WATCH_MAX_FILE_BYTES:
                    continue
            except Exception:
                continue
            source_path = str(p.resolve())
            if source_path in existing_source_paths:
                continue
            candidates.append((st.st_mtime, watch, p))

    candidates.sort(key=lambda x: x[0], reverse=True)
    candidates = candidates[: max(1, min(max_files, 1000))]

    processed = 0
    errors = 0
    error_samples: List[str] = []
    touched_slugs: set[str] = set()
    for _, watch, path in candidates:
        try:
            _ingest_watch_file(watch, path)
            processed += 1
            if watch.get("project_slug"):
                touched_slugs.add(str(watch["project_slug"]))
        except Exception as e:
            errors += 1
            if len(error_samples) < 5:
                error_samples.append(f"{path}: {e}")

    for slug in touched_slugs:
        _update_project_intelligence_summary(slug)

    return {
        "watches": len(watches),
        "candidates": len(candidates),
        "processed": processed,
        "errors": errors,
        "error_samples": error_samples,
        "hours": hours,
        "max_files": max_files,
    }


def _write_project_ingest_record(project_slug: str, record: Dict[str, Any]) -> None:
    project_dir = KNOWLEDGE_PROJECTS_DIR / project_slug / "ingest"
    project_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    rec_path = project_dir / f"{ts}--{record.get('upload_id', uuid4().hex[:8])}.json"
    rec_path.write_text(json.dumps(record, indent=2), encoding="utf-8")
    _update_project_intelligence_summary(project_slug)


def _register_project_watch(
    project_name: str,
    client_name: str,
    address_line: str,
    location_name: str,
    folder_path: str,
) -> Dict[str, Any]:
    folder = Path(folder_path).expanduser().resolve()
    if not folder.exists() or not folder.is_dir():
        raise ValueError("Watch folder does not exist or is not a directory")

    watches = _load_project_watches()
    for w in watches:
        if Path(w.get("folder_path", "")).expanduser().resolve() == folder:
            return w

    task_id: Optional[int] = None
    try:
        sys.path.insert(0, str(BASE_DIR))
        from orchestrator.task_board import add_task as tb_add_task

        task_title = f"Watch ingest synthesis: {project_name or client_name or folder.name}"
        task_desc = "\n".join(
            [
                "Auto-created project watch synthesis task.",
                f"Project: {project_name or 'n/a'}",
                f"Client: {client_name or 'n/a'}",
                f"Address: {address_line or 'n/a'}",
                f"Location: {location_name or 'n/a'}",
                f"Watch folder: {folder}",
                "Purpose: aggregate new watched files into one running synthesis task.",
            ]
        )
        task_id = tb_add_task(
            title=task_title,
            description=task_desc,
            task_type="documentation",
            priority="high",
        )
    except Exception:
        task_id = None

    watch = {
        "watch_id": uuid4().hex[:12],
        "created_at": datetime.now().isoformat(),
        "enabled": True,
        "project_name": project_name or "",
        "client_name": client_name or "",
        "address_line": address_line or "",
        "location_name": location_name or "",
        "folder_path": str(folder),
        "project_slug": _project_slug(project_name, client_name, location_name),
        "task_id": task_id,
    }
    watches.append(watch)
    _save_project_watches(watches)
    return watch


def _seed_watch_state_from_existing_files(watch: Dict[str, Any]) -> Dict[str, Any]:
    folder = Path(watch.get("folder_path", "")).expanduser().resolve()
    if not folder.exists() or not folder.is_dir():
        return {"watch_id": watch.get("watch_id", ""), "seeded": 0, "error": "Watch folder missing"}
    state = _load_project_watch_state()
    watch_id = watch.get("watch_id", "")
    watch_state = state.get(watch_id, {"files": {}})
    files_map: Dict[str, str] = watch_state.get("files", {})
    seeded = 0
    for p in folder.rglob("*"):
        if not p.is_file():
            continue
        rel = str(p.relative_to(folder))
        if rel.startswith(".") or "/." in rel:
            continue
        if _should_ignore_watch_path(rel, p.name):
            continue
        try:
            stat = p.stat()
            sig = f"{stat.st_size}:{int(stat.st_mtime)}"
            if rel not in files_map:
                files_map[rel] = sig
                seeded += 1
        except Exception:
            continue
    watch_state["files"] = files_map
    watch_state["seeded_at"] = datetime.now().isoformat()
    watch_state["last_scan_at"] = datetime.now().isoformat()
    watch_state["last_processed_count"] = 0
    watch_state["last_error"] = ""
    state[watch_id] = watch_state
    _save_project_watch_state(state)
    return {"watch_id": watch_id, "seeded": seeded}


def _auto_discover_project_watches(max_new: int = 25) -> Dict[str, Any]:
    root = Path(PROJECT_WATCH_PROJECTS_ROOT).expanduser().resolve()
    if not root.exists() or not root.is_dir():
        return {"root": str(root), "discovered": 0, "registered": 0, "error": "Projects root missing"}

    existing = _load_project_watches()
    existing_paths = {
        str(Path(w.get("folder_path", "")).expanduser().resolve())
        for w in existing
        if w.get("folder_path")
    }

    discovered = 0
    registered = 0
    for child in sorted(root.iterdir(), key=lambda p: p.name.lower()):
        if not child.is_dir():
            continue
        discovered += 1
        child_resolved = str(child.resolve())
        if child_resolved in existing_paths:
            continue
        if registered >= max_new:
            break
        meta = _metadata_from_project_folder_name(child.name)
        try:
            watch = _register_project_watch(
                project_name=meta["project_name"],
                client_name=meta["client_name"],
                address_line=meta["address_line"],
                location_name=meta["location_name"],
                folder_path=child_resolved,
            )
            _seed_watch_state_from_existing_files(watch)
            registered += 1
            existing_paths.add(child_resolved)
        except Exception:
            continue

    return {
        "root": str(root),
        "discovered": discovered,
        "registered": registered,
    }


def _ingest_watch_file(watch: Dict[str, Any], source_path: Path) -> Dict[str, Any]:
    category_norm = _categorize_bundle_file(source_path.name)
    day_dir = datetime.now().strftime("%Y-%m-%d")
    upload_id = uuid4().hex[:12]
    stored_filename = _build_intake_filename(
        category=category_norm,
        project_name=watch.get("project_name", ""),
        client_name=watch.get("client_name", ""),
        original_filename=source_path.name,
        address_line=watch.get("address_line", ""),
        location_name=watch.get("location_name", ""),
    )
    target_dir = TASK_UPLOAD_DIR / category_norm / day_dir
    target_dir.mkdir(parents=True, exist_ok=True)
    stored_path = target_dir / stored_filename
    last_error: Optional[Exception] = None
    for _ in range(4):
        try:
            shutil.copy2(source_path, stored_path)
            last_error = None
            break
        except OSError as e:
            last_error = e
            time.sleep(0.6)
    if last_error:
        # iCloud-backed files can fail metadata-preserving copy with deadlock errors;
        # retry with plain bytes copy and then shell cp as a final fallback.
        for _ in range(3):
            try:
                data = source_path.read_bytes()
                stored_path.write_bytes(data)
                last_error = None
                break
            except OSError as e:
                last_error = e
                time.sleep(0.6)
    if last_error:
        try:
            subprocess.run(["/bin/cp", "-f", str(source_path), str(stored_path)], check=True, timeout=20)
            last_error = None
        except Exception as e:
            last_error = e
    if last_error:
        raise last_error

    queue_record = {
        "upload_id": upload_id,
        "watch_id": watch.get("watch_id", ""),
        "created_at": datetime.now().isoformat(),
        "category": category_norm,
        "project_name": watch.get("project_name", ""),
        "client_name": watch.get("client_name", ""),
        "address_line": watch.get("address_line", ""),
        "location_name": watch.get("location_name", ""),
        "source_filename": source_path.name,
        "source_path": str(source_path),
        "stored_filename": stored_filename,
        "stored_path": str(stored_path),
        "task_id": watch.get("task_id"),
    }

    findings: Dict[str, Any] = {}
    if category_norm == "drawing":
        try:
            findings = _first_pass_drawing_findings(
                stored_path=stored_path,
                project_name=watch.get("project_name", ""),
                client_name=watch.get("client_name", ""),
            )
            findings["upload_id"] = upload_id
            findings["watch_id"] = watch.get("watch_id", "")
            findings_path = TASK_UPLOAD_FINDINGS_DIR / f"{upload_id}__findings.json"
            findings_path.write_text(json.dumps(findings, indent=2), encoding="utf-8")
            queue_record["findings_path"] = str(findings_path)
        except Exception as extract_error:
            queue_record["findings_error"] = str(extract_error)

    _append_upload_queue_record(queue_record)
    _write_project_ingest_record(
        project_slug=watch.get("project_slug", "project"),
        record={
            "ingested_at": datetime.now().isoformat(),
            "watch_id": watch.get("watch_id", ""),
            "upload_id": upload_id,
            "category": category_norm,
            "source_path": str(source_path),
            "stored_path": str(stored_path),
            "project_name": watch.get("project_name", ""),
            "client_name": watch.get("client_name", ""),
            "address_line": watch.get("address_line", ""),
            "location_name": watch.get("location_name", ""),
            "findings_summary": {
                "legend_terms_count": len(findings.get("legend_terms", [])) if findings else 0,
                "symbol_matches_count": len(findings.get("symbol_matches", [])) if findings else 0,
                "sheet_references_count": len(findings.get("sheet_references", [])) if findings else 0,
                "dtools_quote_version": findings.get("dtools_quote_version", "") if findings else "",
            },
        },
    )
    return queue_record


def _ensure_watch_task_id(watch: Dict[str, Any]) -> Dict[str, Any]:
    if watch.get("task_id"):
        return watch
    try:
        sys.path.insert(0, str(BASE_DIR))
        from orchestrator.task_board import add_task as tb_add_task

        title = f"Watch ingest synthesis: {watch.get('project_name') or watch.get('client_name') or watch.get('watch_id')}"
        desc = "\n".join(
            [
                "Backfilled watch synthesis task.",
                f"Watch ID: {watch.get('watch_id')}",
                f"Project: {watch.get('project_name') or 'n/a'}",
                f"Folder: {watch.get('folder_path') or 'n/a'}",
            ]
        )
        watch["task_id"] = tb_add_task(
            title=title,
            description=desc,
            task_type="documentation",
            priority="high",
        )
    except Exception:
        return watch

    watches = _load_project_watches()
    changed = False
    for idx, item in enumerate(watches):
        if item.get("watch_id") == watch.get("watch_id"):
            watches[idx] = watch
            changed = True
            break
    if changed:
        _save_project_watches(watches)
    return watch


def _scan_project_watch(watch: Dict[str, Any], max_new_files: int = 40) -> Dict[str, Any]:
    watch = _ensure_watch_task_id(watch)
    folder = Path(watch.get("folder_path", "")).expanduser().resolve()
    if not folder.exists() or not folder.is_dir():
        return {"watch_id": watch.get("watch_id", ""), "processed": 0, "error": "Watch folder missing"}

    started_at = datetime.now()
    state = _load_project_watch_state()
    watch_id = watch.get("watch_id", "")
    watch_state = state.get(watch_id, {"files": {}})
    stale_age_min = 0
    previous_start = watch_state.get("scan_started_at", "")
    if watch_state.get("scan_in_progress") and previous_start:
        try:
            prev_dt = datetime.fromisoformat(previous_start)
            stale_age_min = int((started_at - prev_dt).total_seconds() // 60)
        except Exception:
            stale_age_min = 0
    watch_state["scan_started_at"] = started_at.isoformat()
    watch_state["scan_in_progress"] = True
    known: Dict[str, str] = watch_state.get("files", {})

    candidates: List[Path] = []
    skipped_too_large = 0
    skipped_ignored = 0
    for p in folder.rglob("*"):
        if not p.is_file():
            continue
        rel = str(p.relative_to(folder))
        if rel.startswith(".") or "/." in rel:
            continue
        if _should_ignore_watch_path(rel, p.name):
            skipped_ignored += 1
            continue
        try:
            stat = p.stat()
            if stat.st_size > PROJECT_WATCH_MAX_FILE_BYTES:
                skipped_too_large += 1
                continue
            sig = f"{stat.st_size}:{int(stat.st_mtime)}"
            if known.get(rel) != sig:
                known[rel] = sig
                candidates.append(p)
        except Exception:
            continue

    candidates = sorted(candidates, key=lambda x: x.stat().st_mtime)[:max_new_files]
    processed_records: List[Dict[str, Any]] = []
    for path in candidates:
        try:
            processed_records.append(_ingest_watch_file(watch, path))
        except Exception:
            continue

    watch_state["files"] = known
    watch_state["last_scan_at"] = datetime.now().isoformat()
    watch_state["last_processed_count"] = len(processed_records)
    watch_state["last_candidate_count"] = len(candidates)
    watch_state["last_skipped_too_large"] = skipped_too_large
    watch_state["last_skipped_ignored"] = skipped_ignored
    watch_state["scan_duration_sec"] = round((datetime.now() - started_at).total_seconds(), 2)
    watch_state["scan_in_progress"] = False
    watch_state["last_error"] = ""
    if stale_age_min >= PROJECT_WATCH_STALE_SCAN_MIN:
        watch_state["stale_scan_detected"] = True
        watch_state["stale_scan_age_min"] = stale_age_min
    state[watch_id] = watch_state
    _save_project_watch_state(state)

    log_file = PROJECT_WATCH_LOG_DIR / f"{watch_id}.jsonl"
    with log_file.open("a", encoding="utf-8") as fh:
        fh.write(
            json.dumps(
                {
                    "scanned_at": datetime.now().isoformat(),
                    "processed_count": len(processed_records),
                    "candidate_count": len(candidates),
                    "skipped_too_large": skipped_too_large,
                    "skipped_ignored": skipped_ignored,
                    "processed_upload_ids": [r.get("upload_id") for r in processed_records],
                }
            )
            + "\n"
        )

    return {
        "watch_id": watch_id,
        "folder_path": str(folder),
        "processed": len(processed_records),
        "candidate_count": len(candidates),
        "skipped_too_large": skipped_too_large,
        "skipped_ignored": skipped_ignored,
        "task_id": watch.get("task_id"),
        "last_scan_at": watch_state.get("last_scan_at"),
    }


async def _project_watch_loop() -> None:
    while True:
        try:
            if PROJECT_WATCH_AUTO_DISCOVER:
                _auto_discover_project_watches(max_new=10)
            watches = [w for w in _load_project_watches() if w.get("enabled", True)]
            for watch in watches:
                _scan_project_watch(watch, max_new_files=40)
            if MANUALS_AUTO_DIGEST_ENABLED:
                _manuals_auto_digest_once(max_new_files=15)
        except Exception:
            pass
        await asyncio.sleep(max(30, PROJECT_WATCH_INTERVAL_SEC))


def _append_upload_queue_record(record: Dict[str, Any]) -> None:
    TASK_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    with TASK_UPLOAD_QUEUE_FILE.open("a", encoding="utf-8") as fh:
        fh.write(json.dumps(record) + "\n")


def _task_status(task_id: Optional[int]) -> str:
    if not task_id:
        return "unknown"
    try:
        if not TASK_BOARD_DB.exists():
            return "unknown"
        con = sqlite3.connect(TASK_BOARD_DB)
        cur = con.cursor()
        row = cur.execute("SELECT status FROM tasks WHERE id = ?", (task_id,)).fetchone()
        con.close()
        return (row[0] if row else "unknown") or "unknown"
    except Exception:
        return "unknown"


def _parse_allowed_origins(env_key: str, default_csv: str) -> list[str]:
    raw = (os.environ.get(env_key, default_csv) or "").strip()
    origins = [x.strip() for x in raw.split(",") if x.strip()]
    return origins or ["http://localhost", "http://127.0.0.1"]


def _auth_exempt_path(path: str) -> bool:
    if path in {"/", "/health", "/openapi.json", "/docs", "/redoc"}:
        return True
    return path.startswith("/docs") or path.startswith("/redoc")

if HAS_FASTAPI:
    app = FastAPI(
        title="Symphony AI Mobile API",
        description="REST API for Symphony Smart Homes AI Operations",
        version="1.0.0"
    )
    
    # CORS allowlist (no wildcard in production hardening mode)
    allowed_origins = _parse_allowed_origins(
        "MOBILE_APP_ALLOWED_ORIGINS",
        "http://localhost,http://127.0.0.1,http://bobs-mac-mini:8420",
    )
    app.add_middleware(
        CORSMiddleware,
        allow_origins=allowed_origins,
        allow_credentials=False,
        allow_methods=["*"],
        allow_headers=["Content-Type", "Authorization", "X-Symphony-Token"],
    )

    @app.on_event("startup")
    async def project_watch_startup():
        global PROJECT_WATCH_TASK
        if PROJECT_WATCH_AUTORUN and PROJECT_WATCH_INTERVAL_SEC >= 30:
            PROJECT_WATCH_TASK = asyncio.create_task(_project_watch_loop())

    @app.on_event("shutdown")
    async def project_watch_shutdown():
        global PROJECT_WATCH_TASK
        if PROJECT_WATCH_TASK and not PROJECT_WATCH_TASK.done():
            PROJECT_WATCH_TASK.cancel()
        PROJECT_WATCH_TASK = None

    @app.middleware("http")
    async def symphony_api_auth(request: Request, call_next):
        """
        Require API token on all non-health routes when SYMPHONY_API_TOKEN is set.
        Accepts:
          - X-Symphony-Token: <token>
          - Authorization: Bearer <token>
        """
        if not API_AUTH_TOKEN or _auth_exempt_path(request.url.path):
            return await call_next(request)

        header_token = _normalize_auth_token(request.headers.get("x-symphony-token", ""))
        auth = request.headers.get("authorization", "").strip()
        bearer = _normalize_auth_token(auth[7:].strip()) if auth.lower().startswith("bearer ") else ""
        token = header_token or bearer

        if token != API_AUTH_TOKEN:
            return JSONResponse(status_code=401, content={"detail": "Unauthorized"})
        return await call_next(request)


# --- Helper Functions ---

def _extract_json_from_output(result: Dict, fallback_key: str = "output") -> Dict:
    """Parse JSON from run_command output; return clean result for API consumers."""
    output = result.get("output", "")
    if not output:
        return {"success": False, "error": result.get("error", "No output")}
    # Handle output that may have progress text before JSON (e.g. "Checking...\n{...}")
    try:
        data = json.loads(output)
        return data
    except json.JSONDecodeError:
        pass
    # Try to extract JSON object from output
    import re
    match = re.search(r'\{[\s\S]*\}', output)
    if match:
        try:
            return json.loads(match.group(0))
        except json.JSONDecodeError:
            pass
    return {"success": False, "error": "Could not parse output", fallback_key: output[:500]}


def run_command(cmd: List[str], timeout: int = 30) -> Dict:
    """Run a command and return result."""
    return shared_run_command(cmd, timeout=timeout, cwd=BASE_DIR)


def run_tool_endpoint(script: str, args: List[str], timeout: int = 60) -> Dict:
    """Run a tools/ script and return {success, output, error}."""
    return run_tool_script(BASE_DIR, script, args=args, timeout=timeout)


def _network_watch_pid() -> Optional[int]:
    try:
        if not NETWORK_WATCH_PID_FILE.exists():
            return None
        pid = int(NETWORK_WATCH_PID_FILE.read_text(encoding="utf-8").strip())
        return pid if pid > 0 else None
    except Exception:
        return None


def _pid_is_alive(pid: Optional[int]) -> bool:
    if not pid:
        return False
    try:
        os.kill(pid, 0)
        return True
    except OSError:
        return False


def _read_json_file(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _read_recent_jsonl(path: Path, limit: int = 20) -> List[Dict[str, Any]]:
    """Read the last N JSONL records from a file."""
    if not path.exists():
        return []
    capped_limit = max(1, min(int(limit), 200))
    try:
        with path.open("r", encoding="utf-8") as f:
            lines = f.readlines()
        recent = lines[-capped_limit:]
        items: List[Dict[str, Any]] = []
        for raw in recent:
            raw = raw.strip()
            if not raw:
                continue
            try:
                payload = json.loads(raw)
                if isinstance(payload, dict):
                    items.append(payload)
            except json.JSONDecodeError:
                continue
        return list(reversed(items))
    except Exception:
        return []


def _read_all_jsonl(path: Path) -> List[Dict[str, Any]]:
    if not path.exists():
        return []
    items: List[Dict[str, Any]] = []
    try:
        with path.open("r", encoding="utf-8") as f:
            for raw in f:
                raw = raw.strip()
                if not raw:
                    continue
                try:
                    payload = json.loads(raw)
                    if isinstance(payload, dict):
                        items.append(payload)
                except json.JSONDecodeError:
                    continue
    except Exception:
        return []
    return items


def _write_all_jsonl(path: Path, items: List[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for item in items:
            f.write(json.dumps(item) + "\n")


def _append_jsonl(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(payload) + "\n")


def _update_draft(path: Path, draft_id: str, updates: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    rows = _read_all_jsonl(path)
    changed = False
    target: Optional[Dict[str, Any]] = None
    for row in rows:
        if str(row.get("draft_id", "")) != draft_id:
            continue
        row.update(updates)
        changed = True
        target = row
    if changed:
        _write_all_jsonl(path, rows)
    return target


def _parse_iso_or_default(value: Optional[str], fallback: datetime) -> datetime:
    if not value:
        return fallback
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except Exception:
        return fallback


def _send_imessage_text(handle: str, message: str) -> Dict[str, Any]:
    target = (handle or "").strip()
    body = (message or "").strip()
    if not target:
        return {"success": False, "error": "Missing destination handle"}
    if not body:
        return {"success": False, "error": "Missing message body"}

    script = (
        'on run argv\n'
        'set targetHandle to item 1 of argv\n'
        'set bodyText to item 2 of argv\n'
        'tell application "Messages"\n'
        'set targetService to first service whose service type = iMessage\n'
        'set targetBuddy to buddy targetHandle of targetService\n'
        'send bodyText to targetBuddy\n'
        'end tell\n'
        'return "ok"\n'
        'end run\n'
    )
    try:
        result = subprocess.run(
            ["osascript", "-e", script, target, body],
            capture_output=True,
            text=True,
            timeout=20,
        )
        if result.returncode != 0:
            return {
                "success": False,
                "error": (result.stderr or result.stdout or "osascript failed").strip(),
            }
        return {"success": True}
    except Exception as exc:
        return {"success": False, "error": str(exc)}


def _create_calendar_event(title: str, notes: str, start_at: datetime, duration_min: int = 60) -> Dict[str, Any]:
    end_at = start_at.timestamp() + max(15, int(duration_min)) * 60
    end_dt = datetime.fromtimestamp(end_at)
    # RFC3339-like strings accepted by AppleScript date parser.
    start_str = start_at.strftime("%Y-%m-%d %H:%M:%S")
    end_str = end_dt.strftime("%Y-%m-%d %H:%M:%S")
    script = (
        'on run argv\n'
        'set titleText to item 1 of argv\n'
        'set notesText to item 2 of argv\n'
        'set startText to item 3 of argv\n'
        'set endText to item 4 of argv\n'
        'tell application "Calendar"\n'
        'if not (exists calendar "Symphony Ops") then\n'
        'make new calendar with properties {name:"Symphony Ops"}\n'
        'end if\n'
        'set startDate to date startText\n'
        'set endDate to date endText\n'
        'tell calendar "Symphony Ops"\n'
        'set newEvent to make new event with properties {summary:titleText, description:notesText, start date:startDate, end date:endDate}\n'
        'end tell\n'
        'end tell\n'
        'return "ok"\n'
        'end run\n'
    )
    try:
        result = subprocess.run(
            ["osascript", "-e", script, title, notes, start_str, end_str],
            capture_output=True,
            text=True,
            timeout=20,
        )
        if result.returncode != 0:
            return {
                "success": False,
                "error": (result.stderr or result.stdout or "calendar osascript failed").strip(),
            }
        return {"success": True, "start": start_str, "end": end_str}
    except Exception as exc:
        return {"success": False, "error": str(exc)}


def _lookup_handle_for_rowid(rowid: Any) -> str:
    try:
        rowid_int = int(rowid)
    except Exception:
        return ""


def _load_intake_failures() -> List[Dict[str, Any]]:
    payload = _read_json_file(IMESSAGE_INTAKE_FAILURES_FILE)
    rows = payload.get("items", []) if isinstance(payload, dict) else []
    return [r for r in rows if isinstance(r, dict)]


def _save_intake_failures(rows: List[Dict[str, Any]]) -> None:
    IMESSAGE_INTAKE_FAILURES_FILE.parent.mkdir(parents=True, exist_ok=True)
    IMESSAGE_INTAKE_FAILURES_FILE.write_text(json.dumps({"items": rows}, indent=2), encoding="utf-8")


def _append_intake_audit(
    action: str,
    kind: str,
    draft_id: str,
    success: bool,
    payload: Optional[Dict[str, Any]] = None,
    result: Optional[Dict[str, Any]] = None,
    error: str = "",
) -> None:
    _append_jsonl(
        IMESSAGE_INTAKE_AUDIT_FILE,
        {
            "id": f"audit_{uuid4().hex[:12]}",
            "timestamp": datetime.now().isoformat(),
            "action": action,
            "kind": kind,
            "draft_id": draft_id,
            "success": bool(success),
            "error": error,
            "payload": payload or {},
            "result": result or {},
        },
    )


def _enqueue_intake_failure(
    action: str,
    kind: str,
    draft_id: str,
    payload: Dict[str, Any],
    error: str,
) -> str:
    rows = _load_intake_failures()
    failure_id = f"ifail_{uuid4().hex[:10]}"
    rows.append(
        {
            "id": failure_id,
            "created_at": datetime.now().isoformat(),
            "status": "pending",
            "attempt_count": 0,
            "last_attempt_at": None,
            "resolved_at": None,
            "action": action,
            "kind": kind,
            "draft_id": draft_id,
            "payload": payload,
            "last_error": error,
        }
    )
    _save_intake_failures(rows)
    return failure_id


def _execute_invoice_approve(payload: Dict[str, Any]) -> Dict[str, Any]:
    draft_id = str(payload.get("draft_id") or "").strip()
    if not draft_id:
        return {"success": False, "error": "draft_id is required"}
    now = datetime.now().isoformat()
    updated = _update_draft(
        IMESSAGE_INVOICE_DRAFTS_FILE,
        draft_id,
        {
            "status": "approved",
            "approved_at": now,
            "approval_note": str(payload.get("note") or "").strip(),
            "last_action_at": now,
        },
    )
    if not updated:
        return {"success": False, "error": "Invoice draft not found", "draft_id": draft_id}

    sent = {"success": False, "skipped": True}
    if bool(payload.get("send_confirmation", True)):
        handle = str(updated.get("handle_raw") or "") or _lookup_handle_for_rowid(updated.get("rowid"))
        msg = str(payload.get("confirmation_message") or "").strip() or "Your service invoice draft has been approved. We will send final details shortly."
        if handle:
            sent = _send_imessage_text(handle, msg)
        else:
            sent = {"success": False, "error": "Draft has no raw handle for outbound text"}
    ok = bool(sent.get("success", True) if not sent.get("skipped") else True)
    return {"success": ok, "draft_id": draft_id, "status": "approved", "confirmation": sent}


def _execute_appointment_schedule(payload: Dict[str, Any]) -> Dict[str, Any]:
    draft_id = str(payload.get("draft_id") or "").strip()
    if not draft_id:
        return {"success": False, "error": "draft_id is required"}
    rows = _read_all_jsonl(IMESSAGE_APPOINTMENT_DRAFTS_FILE)
    target = next((r for r in rows if str(r.get("draft_id", "")) == draft_id), None)
    if not target:
        return {"success": False, "error": "Appointment draft not found", "draft_id": draft_id}

    fallback_start = datetime.now().replace(hour=9, minute=0, second=0, microsecond=0)
    if fallback_start <= datetime.now():
        fallback_start = fallback_start + timedelta(days=1)
    start_at = _parse_iso_or_default(payload.get("proposed_start"), fallback_start)
    title = f"Service Appointment - {target.get('contact_name_masked') or target.get('handle_masked') or 'Client'}"
    notes = f"{target.get('request_text_redacted', '')}\n\nProject: {target.get('project_hint', '')}\nDraft ID: {draft_id}"
    cal = _create_calendar_event(title, notes, start_at, duration_min=max(15, int(payload.get("duration_min") or 60)))

    now = datetime.now().isoformat()
    status_value = "scheduled" if cal.get("success") else "schedule_failed"
    updated = _update_draft(
        IMESSAGE_APPOINTMENT_DRAFTS_FILE,
        draft_id,
        {
            "status": status_value,
            "proposed_start": start_at.isoformat(),
            "calendar_event_created_at": now if cal.get("success") else None,
            "calendar_result": cal,
            "last_action_at": now,
        },
    )
    if not updated:
        return {"success": False, "error": "Appointment draft not found", "draft_id": draft_id}

    sent = {"success": False, "skipped": True}
    if bool(payload.get("send_confirmation", True)):
        handle = str(updated.get("handle_raw") or "") or _lookup_handle_for_rowid(updated.get("rowid"))
        when_text = start_at.strftime("%a %b %d at %I:%M %p")
        msg = str(payload.get("confirmation_message") or "").strip() or f"Your appointment is scheduled for {when_text}. Reply here if you need to adjust."
        if handle:
            sent = _send_imessage_text(handle, msg)
        else:
            sent = {"success": False, "error": "Draft has no raw handle for outbound text"}

    ok = bool(cal.get("success")) and bool(sent.get("success", True) if not sent.get("skipped") else True)
    return {"success": ok, "draft_id": draft_id, "status": status_value, "calendar": cal, "confirmation": sent}


def _execute_confirmation_send(payload: Dict[str, Any]) -> Dict[str, Any]:
    kind = str(payload.get("kind") or "").strip().lower()
    draft_id = str(payload.get("draft_id") or "").strip()
    if kind not in {"invoice", "appointment"}:
        return {"success": False, "error": "kind must be invoice or appointment"}
    if not draft_id:
        return {"success": False, "error": "draft_id is required"}

    draft_file = IMESSAGE_INVOICE_DRAFTS_FILE if kind == "invoice" else IMESSAGE_APPOINTMENT_DRAFTS_FILE
    rows = _read_all_jsonl(draft_file)
    target = next((r for r in rows if str(r.get("draft_id", "")) == draft_id), None)
    if not target:
        return {"success": False, "error": "Draft not found", "draft_id": draft_id, "kind": kind}

    handle = str(target.get("handle_raw") or "") or _lookup_handle_for_rowid(target.get("rowid"))
    if not handle:
        return {"success": False, "error": "Draft has no raw handle for outbound text", "draft_id": draft_id, "kind": kind}

    default_message = (
        "Your service invoice draft has been approved. We will send final details shortly."
        if kind == "invoice"
        else "Your appointment request is in progress. We will confirm your time shortly."
    )
    message = str(payload.get("message") or "").strip() or default_message
    sent = _send_imessage_text(handle, message)
    now = datetime.now().isoformat()
    _update_draft(
        draft_file,
        draft_id,
        {
            "last_confirmation_at": now if sent.get("success") else None,
            "last_confirmation_message": message,
            "last_action_at": now,
        },
    )
    return {"success": bool(sent.get("success")), "draft_id": draft_id, "kind": kind, "result": sent}
    if not MESSAGES_DB_PATH.exists():
        return ""
    try:
        conn = sqlite3.connect(f"file:{MESSAGES_DB_PATH}?mode=ro", uri=True)
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            """
            SELECT COALESCE(h.id, '') AS handle
            FROM message m
            LEFT JOIN handle h ON h.ROWID = m.handle_id
            WHERE m.ROWID = ?
            LIMIT 1
            """,
            (rowid_int,),
        ).fetchone()
        conn.close()
        return str(row["handle"]) if row else ""
    except Exception:
        return ""


def _http_get_json(url: str, timeout: int = 6) -> Dict[str, Any]:
    req = urllib.request.Request(url, headers={"Accept": "application/json"})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return json.loads(resp.read().decode("utf-8"))


def _read_csv_rows(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    try:
        with path.open("r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            rows: List[Dict[str, str]] = []
            for row in reader:
                rows.append({str(k): str(v or "").strip() for k, v in row.items() if k is not None})
            return rows
    except Exception:
        return []


def _normalize_sku_key(value: str) -> str:
    return re.sub(r"[^A-Z0-9\-]+", "-", (value or "").strip().upper()).strip("-")


def _load_inventory_stock_levels() -> Dict[str, Dict[str, Any]]:
    if not INVENTORY_STOCK_FILE.exists():
        try:
            INVENTORY_STOCK_FILE.write_text(json.dumps({"items": []}, indent=2), encoding="utf-8")
        except Exception:
            pass
    payload = _read_json_file(INVENTORY_STOCK_FILE)
    raw_items = payload.get("items", []) if isinstance(payload, dict) else []
    stock: Dict[str, Dict[str, Any]] = {}
    if isinstance(raw_items, list):
        for item in raw_items:
            if not isinstance(item, dict):
                continue
            sku = _normalize_sku_key(str(item.get("sku") or item.get("model_sku") or ""))
            if not sku:
                continue
            try:
                on_hand = int(float(item.get("on_hand", 0) or 0))
            except Exception:
                on_hand = 0
            try:
                reorder = int(float(item.get("reorder_point", 0) or 0))
            except Exception:
                reorder = 0
            stock[sku] = {
                "sku": sku,
                "on_hand": on_hand,
                "reorder_point": reorder,
                "supplier": str(item.get("supplier", "") or "").strip(),
                "notes": str(item.get("notes", "") or "").strip(),
            }
    return stock


def _build_inventory_summary(low_stock_limit: int = 25, top_limit: int = 100) -> Dict[str, Any]:
    inventory_rows = _read_csv_rows(INVENTORY_REPORTS_DIR / "SKU_Inventory.csv")
    manual_queue_rows = _read_csv_rows(INVENTORY_REPORTS_DIR / "manual_fetch_queue.csv")
    stock_map = _load_inventory_stock_levels()

    items: List[Dict[str, Any]] = []
    for row in inventory_rows:
        sku = _normalize_sku_key(row.get("model_sku", ""))
        if not sku:
            continue
        try:
            observed = int(float(row.get("count", 0) or 0))
        except Exception:
            observed = 0
        stock = stock_map.get(sku, {})
        on_hand = int(stock.get("on_hand", 0) or 0)
        reorder = int(stock.get("reorder_point", 0) or 0)
        low = bool(reorder > 0 and on_hand <= reorder)
        items.append({
            "sku": sku,
            "manufacturer": row.get("manufacturer", ""),
            "category": row.get("category", ""),
            "observed_count": observed,
            "on_hand": on_hand,
            "reorder_point": reorder,
            "low_stock": low,
            "supplier": stock.get("supplier", ""),
            "notes": stock.get("notes", ""),
        })

    items.sort(key=lambda r: (not r.get("low_stock", False), -(r.get("observed_count", 0) or 0), r.get("sku", "")))
    low_stock_items = [r for r in items if r.get("low_stock")][: max(1, min(low_stock_limit, 200))]
    top_items = items[: max(1, min(top_limit, 500))]
    missing_stock_setup = [
        r for r in items
        if r.get("observed_count", 0) >= 2 and int(r.get("reorder_point", 0) or 0) <= 0
    ][:50]

    todo_manuals = [
        r for r in manual_queue_rows
        if str(r.get("status", "todo") or "todo").strip().lower() in {"", "todo", "pending"}
    ]

    return {
        "success": True,
        "generated_at": datetime.now().isoformat(),
        "paths": {
            "inventory_csv": str(INVENTORY_REPORTS_DIR / "SKU_Inventory.csv"),
            "manual_queue_csv": str(INVENTORY_REPORTS_DIR / "manual_fetch_queue.csv"),
            "stock_file": str(INVENTORY_STOCK_FILE),
        },
        "counts": {
            "inventory_rows": len(inventory_rows),
            "tracked_stock_items": len(stock_map),
            "low_stock_count": len([r for r in items if r.get("low_stock")]),
            "manual_queue_todo_count": len(todo_manuals),
        },
        "low_stock_items": low_stock_items,
        "top_items": top_items,
        "missing_stock_setup": missing_stock_setup,
        "manual_queue_preview": todo_manuals[:25],
    }


def _has_env_key(name: str) -> bool:
    value = (os.environ.get(name, "") or "").strip()
    return bool(value)


def _valid_task_types() -> List[str]:
    return [
        "research",
        "documentation",
        "troubleshooting",
        "proposal",
        "commissioning",
        "learning",
        "idea",
        "maintenance",
        "integration",
        "claude",
    ]


def _default_employee_bots() -> Dict[str, Any]:
    return {
        "workers": {
            "betty": {
                "name": "Betty",
                "role": "Research & Documentation Specialist",
                "emoji": "📚",
                "skills": ["research", "documentation", "learning", "troubleshooting"],
                "intro": "Hi! I'm Betty, the research specialist.",
                "token_env": "BETTY_BOT_TOKEN",
            },
            "beatrice": {
                "name": "Beatrice",
                "role": "Proposals & Commissioning Specialist",
                "emoji": "📋",
                "skills": ["proposal", "commissioning", "integration"],
                "intro": "Hi! I'm Beatrice, the proposals specialist.",
                "token_env": "BEATRICE_BOT_TOKEN",
            },
            "bill": {
                "name": "Bill",
                "role": "Maintenance & Integration Specialist",
                "emoji": "🔧",
                "skills": ["maintenance", "troubleshooting", "integration"],
                "intro": "Hi! I'm Bill, the maintenance specialist.",
                "token_env": "BILL_BOT_TOKEN",
            },
        }
    }


def _load_employee_bots() -> Dict[str, Any]:
    if not EMPLOYEE_BOTS_FILE.exists():
        EMPLOYEE_BOTS_FILE.parent.mkdir(parents=True, exist_ok=True)
        payload = _default_employee_bots()
        try:
            EMPLOYEE_BOTS_FILE.write_text(json.dumps(payload, indent=2), encoding="utf-8")
        except Exception:
            pass
        return payload
    payload = _read_json_file(EMPLOYEE_BOTS_FILE)
    if not isinstance(payload, dict) or "workers" not in payload:
        return _default_employee_bots()
    return payload


def _save_employee_bots(payload: Dict[str, Any]) -> None:
    EMPLOYEE_BOTS_FILE.parent.mkdir(parents=True, exist_ok=True)
    EMPLOYEE_BOTS_FILE.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def _employee_bot_paths(worker_id: str) -> Dict[str, Path]:
    safe_id = _slugify(worker_id, fallback="worker")
    return {
        "pid": EMPLOYEE_BOTS_RUNTIME_DIR / f"{safe_id}.pid",
        "log": BASE_DIR / "logs" / f"employee-bot-{safe_id}.log",
        "err": BASE_DIR / "logs" / f"employee-bot-{safe_id}.err.log",
    }


def _employee_bot_pid(worker_id: str) -> Optional[int]:
    paths = _employee_bot_paths(worker_id)
    pid_file = paths["pid"]
    if not pid_file.exists():
        return None
    try:
        pid = int(pid_file.read_text(encoding="utf-8").strip())
        return pid if pid > 0 else None
    except Exception:
        return None


def _pick_python_with_module(module_name: str) -> str:
    candidates = [
        str(BASE_DIR / ".venv" / "bin" / "python3"),
        "/opt/homebrew/bin/python3",
        "/usr/local/bin/python3",
        sys.executable,
        "/usr/bin/python3",
    ]
    for candidate in candidates:
        if not candidate:
            continue
        path = Path(candidate)
        if not path.exists():
            continue
        try:
            result = subprocess.run(
                [str(path), "-c", f"import {module_name}"],
                capture_output=True,
                text=True,
                timeout=5,
            )
            if result.returncode == 0:
                return str(path)
        except Exception:
            continue
    return ""


def _employee_bot_runtime_status(worker_id: str) -> Dict[str, Any]:
    pid = _employee_bot_pid(worker_id)
    running = _pid_is_alive(pid)
    paths = _employee_bot_paths(worker_id)
    return {
        "worker_id": worker_id,
        "running": running,
        "pid": pid if running else None,
        "pid_file": str(paths["pid"]),
        "log_file": str(paths["log"]),
        "error_log_file": str(paths["err"]),
    }


def _to_money(value: Any) -> Optional[float]:
    """Coerce numeric-ish value to float."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("$", "").replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _pick_unit_cost(prices: List[float], dealer_tier: str) -> Optional[float]:
    """
    Price columns expected in many manufacturer sheets:
    MSRP, Standard, Silver, Gold, Dist/Fab
    """
    if len(prices) < 2:
        return None
    tier = (dealer_tier or "standard").strip().lower()
    if tier == "fabricator":
        idx = 4
    elif tier == "gold":
        idx = 3
    elif tier == "silver":
        idx = 2
    else:
        idx = 1
    if idx < len(prices):
        return prices[idx]
    return prices[-1]


def _looks_like_price_line(line: str) -> bool:
    nums = re.findall(r"\d{1,3}(?:,\d{3})*(?:\.\d{2})", line)
    return len(nums) >= 2


def _extract_part_number_from_line(line: str) -> Optional[str]:
    upper = line.upper().strip()
    # Prefer SKU-like tokens whose first segment is alpha-heavy.
    tokens = re.findall(r"[A-Z0-9#]+(?:-[A-Z0-9#]+)+", upper)
    if not tokens:
        return None
    candidates: List[str] = []
    for token in tokens:
        parts = token.split("-")
        first = parts[0]
        if first.isdigit():
            continue
        if len(first) < 3:
            continue
        if first.endswith("YR"):
            continue
        if re.match(r"^\d{4}-\d{2}-\d{2}$", token):
            continue
        if token.startswith("Q") and token[1:].isdigit():
            continue
        if len(parts) == 2 and parts[1].isdigit() and first in {"POE", "DC"}:
            continue
        if "WARRANTY" in token:
            continue
        if not any(ch.isalpha() for ch in token):
            continue
        candidates.append(token)
    if not candidates:
        return None
    # Usually the SKU is the longest code in the row.
    return sorted(candidates, key=len, reverse=True)[0]


def _extract_text_from_pdf(path: Path) -> str:
    """Extract PDF text using pypdf, with pdftotext fallback."""
    try:
        from pypdf import PdfReader  # type: ignore
        reader = PdfReader(str(path))
        chunks: List[str] = []
        total = 0
        for idx, page in enumerate(reader.pages):
            if idx >= max(1, PDF_EXTRACT_MAX_PAGES):
                break
            chunks.append(page.extract_text() or "")
            total += len(chunks[-1])
            if total >= max(1000, PDF_EXTRACT_MAX_CHARS):
                break
        return "\n".join(chunks)[: max(1000, PDF_EXTRACT_MAX_CHARS)]
    except Exception:
        try:
            # macOS/homebrew fallback.
            result = subprocess.run(
                ["pdftotext", str(path), "-"],
                capture_output=True,
                text=True,
                timeout=45,
            )
            if result.returncode == 0 and result.stdout.strip():
                return result.stdout[: max(1000, PDF_EXTRACT_MAX_CHARS)]
        except Exception:
            pass
        raise RuntimeError(
            "Could not parse PDF text. Install pypdf (`pip3 install pypdf`) or ensure pdftotext is installed."
        )


def _parse_price_sheet_text(text: str, dealer_tier: str) -> List[Dict[str, Any]]:
    """
    Parse semi-structured price-book text into D-Tools product drafts.
    Works best for sheets with rows that include a SKU and nearby pricing line.
    """
    lines = [ln.strip() for ln in text.splitlines()]
    lines = [ln for ln in lines if ln]
    products: List[Dict[str, Any]] = []

    def _is_noise_line(value: str) -> bool:
        up = value.upper()
        return (
            "PRICE BOOK" in up
            or "MODEL NAME" in up
            or "PART NUMBER" in up
            or "ALL BRAND NAMES" in up
            or "ARE THE PROPERTY" in up
            or "DEALER LEVEL" in up
            or "TERMS:" in up
            or up.startswith("-- ")
            or up.startswith("#")
        )

    # Pass 1: detect SKU anchors.
    sku_rows: List[tuple[int, str]] = []
    for idx, line in enumerate(lines):
        if _is_noise_line(line):
            continue
        part = _extract_part_number_from_line(line)
        if not part:
            continue
        words = line.split()
        if line.strip().upper() == part:
            sku_rows.append((idx, part))
            continue
        if line.strip().upper().endswith(part) and len(words) <= 6:
            sku_rows.append((idx, part))

    for row_idx, (start, part) in enumerate(sku_rows):
        end = sku_rows[row_idx + 1][0] if row_idx + 1 < len(sku_rows) else min(len(lines), start + 35)
        block = lines[start + 1:end]

        # Model: nearest previous human-readable line.
        model = ""
        j = start - 1
        while j >= 0:
            prev = lines[j].strip()
            if _is_noise_line(prev):
                j -= 1
                continue
            if "MADE IN USA" in prev.upper():
                j -= 1
                continue
            if _extract_part_number_from_line(prev):
                j -= 1
                continue
            prev_upper = prev.upper()
            if any(
                word in prev_upper
                for word in ["DEALER", "MSRP", "SKU", "DESCRIPTION", "WARRANTY", "NOW", "Q226", "Q326", "Q426"]
            ):
                j -= 1
                continue
            if re.search(r"[A-Za-z]", prev):
                model = prev
                break
            j -= 1
        if not model:
            model = part

        # Description: text lines before dense pricing starts.
        desc_parts: List[str] = []
        prices: List[float] = []
        for ln in block:
            nums = re.findall(r"\d{1,3}(?:,\d{3})*(?:\.\d{2})", ln)
            if nums:
                prices.extend([float(n.replace(",", "")) for n in nums])
                continue
            up = ln.upper()
            if (
                "$" in ln
                or up in {"NOW", "Q226", "Q326", "Q426"}
                or _extract_part_number_from_line(ln)
                or _is_noise_line(ln)
            ):
                continue
            if len(ln) <= 2:
                continue
            desc_parts.append(ln)

        prices = prices[:5]
        if not prices:
            # Some PDFs print price columns before SKU in reading order.
            nearby = lines[max(0, start - 20):end]
            backfill: List[float] = []
            for ln in nearby:
                nums = re.findall(r"\d{1,3}(?:,\d{3})*(?:\.\d{2})", ln)
                backfill.extend([float(n.replace(",", "")) for n in nums])
                if len(backfill) >= 5:
                    break
            prices = backfill[:5]
        msrp = prices[0] if prices else None
        unit_cost = _pick_unit_cost(prices, dealer_tier)
        unit_price = msrp
        short_desc = " ".join(desc_parts).strip()
        if len(short_desc) > 300:
            short_desc = short_desc[:297] + "..."
        if not short_desc:
            short_desc = f"{model} {part}".strip()

        up_model = model.upper()
        if "TERMS" in up_model or "DEALER LEVEL" in up_model:
            continue

        bad_model_tokens = ["DEALER", "WARRANTY", "AVAIL", "Q226", "Q326", "Q426"]
        if any(tok in up_model for tok in bad_model_tokens):
            model = part
        if model.startswith("10YR") or model.startswith("Q") or "/" in model:
            model = part

        keywords = ", ".join([kw for kw in ["Modern Atomics", model, part, "power", "pdu"] if kw])
        products.append(
            {
                "brand": "Modern Atomics",
                "model": model.replace("™", "").replace("®", "").strip(),
                "part_number": part.strip(),
                "category": "Power Management",
                "short_description": short_desc,
                "keywords": keywords[:280],
                "unit_price": unit_price,
                "unit_cost": unit_cost,
                "msrp": msrp,
                "supplier": "Modern Atomics",
            }
        )

    # Deduplicate by part number.
    dedup: Dict[str, Dict[str, Any]] = {}
    for p in products:
        key = (p.get("part_number") or "").upper()
        if key and key not in dedup:
            dedup[key] = p
    return list(dedup.values())


def _parse_sheet_file(path: Path, dealer_tier: str) -> List[Dict[str, Any]]:
    return _parse_sheet_file_with_profile(path, dealer_tier, parse_profile="auto", custom_columns=None)


def _normalize_header_name(value: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", value.upper())


def _parse_sheet_file_with_profile(
    path: Path,
    dealer_tier: str,
    parse_profile: str,
    custom_columns: Optional[List[str]],
) -> List[Dict[str, Any]]:
    suffix = path.suffix.lower()
    profile = PARSE_PROFILES.get(parse_profile, PARSE_PROFILES["auto"])
    expected_columns = custom_columns or profile.get("expected_columns", [])
    expected_normalized = [_normalize_header_name(col) for col in expected_columns if col]

    def as_text(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def load_excel_rows(excel_path: Path) -> List[Dict[str, Any]]:
        if not HAS_OPENPYXL:
            raise RuntimeError("Excel parsing not available (openpyxl not installed).")
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)  # type: ignore[name-defined]
        ws = wb["Products"] if "Products" in wb.sheetnames else wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return []
        headers = [as_text(h) or f"col_{idx}" for idx, h in enumerate(header_row)]
        parsed_rows: List[Dict[str, Any]] = []
        for row in rows_iter:
            if row is None:
                continue
            row_dict: Dict[str, Any] = {}
            has_value = False
            for idx, value in enumerate(row):
                if idx >= len(headers):
                    continue
                row_dict[headers[idx]] = value
                if value not in (None, ""):
                    has_value = True
            if has_value:
                parsed_rows.append(row_dict)
        return parsed_rows

    def pick_value_by_patterns(row: Dict[str, Any], patterns: List[str]) -> str:
        if not row:
            return ""
        normalized = { _normalize_header_name(k): v for k, v in row.items() if k is not None }
        for pat in patterns:
            pat_n = _normalize_header_name(pat)
            for key_n, value in normalized.items():
                if pat_n and pat_n in key_n:
                    return str(value or "").strip()
        return ""

    rows: List[Dict[str, Any]] = []
    if suffix == ".csv":
        with path.open("r", encoding="utf-8", errors="ignore", newline="") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
    elif suffix in {".xlsx", ".xlsm", ".xltx"}:
        rows = load_excel_rows(path)
    elif suffix == ".xls":
        raise RuntimeError("Legacy .xls not supported. Please re-save as .xlsx and upload again.")

    if rows:
        first_keys = {_normalize_header_name(k) for k in rows[0].keys() if k is not None}
        # Auto-detect Rexel-style invoice headers.
        if parse_profile == "auto":
            if {"ITEMNUMBER", "MANUFACTURER", "SKU", "DESCRIPTION", "UNITPRICE"} <= first_keys:
                parse_profile = "invoice_rexel"

        parsed: List[Dict[str, Any]] = []
        for row in rows:
            if parse_profile == "invoice_rexel":
                part = (
                    pick_value_by_patterns(row, ["SKU", "PART NUMBER", "PART"])
                    or as_text(row.get("SKU"))
                    or as_text(row.get("Part Number"))
                )
                model = (
                    pick_value_by_patterns(row, ["MODEL NAME", "MODEL"])
                    or part
                )
                if not model and not part:
                    continue
                brand = (
                    pick_value_by_patterns(row, ["MANUFACTURER", "BRAND"])
                    or as_text(row.get("Manufacturer"))
                    or "Unknown"
                )
                description = (
                    pick_value_by_patterns(row, ["DESCRIPTION", "ITEM DESCRIPTION"])
                    or as_text(row.get("Description"))
                    or model
                )
                unit_price = _to_money(
                    pick_value_by_patterns(row, ["UNIT PRICE", "PRICE", "COST"])
                    or row.get("Unit Price")
                )
                item_number = as_text(
                    pick_value_by_patterns(row, ["ITEM NUMBER"]) or row.get("Item Number")
                )
                upc = as_text(pick_value_by_patterns(row, ["UPC"]) or row.get("UPC"))
                qty = as_text(
                    pick_value_by_patterns(row, ["SHIP QTY", "QUANTITY"]) or row.get("Ship Qty") or row.get("Quantity")
                )
                keywords = ", ".join([x for x in [brand, model, part, f"qty:{qty}" if qty else "", f"item:{item_number}" if item_number else ""] if x])
                if upc:
                    keywords = f"{keywords}, upc:{upc}" if keywords else f"upc:{upc}"

                parsed.append(
                    {
                        "brand": brand,
                        "model": model,
                        "part_number": part,
                        "category": "Electrical",
                        "short_description": description[:300],
                        "keywords": keywords[:280],
                        "unit_price": unit_price,
                        "unit_cost": unit_price,
                        "msrp": unit_price,
                        "supplier": "Rexel",
                    }
                )
                continue

            model = (
                pick_value_by_patterns(row, ["MODEL NAME", "MODEL", "NAME"])
                or as_text(row.get("Model"))
                or as_text(row.get("MODEL"))
                or as_text(row.get("Name"))
            )
            part = (
                pick_value_by_patterns(row, ["PART NUMBER", "SKU", "PART"])
                or as_text(row.get("Part Number"))
                or as_text(row.get("SKU"))
                or as_text(row.get("Part"))
            )
            if not model and not part:
                continue
            msrp = _to_money(
                pick_value_by_patterns(row, ["MSRP", "LIST PRICE", "PRICE"])
                or row.get("MSRP")
            )

            if expected_normalized:
                # If profile declares dealer columns, respect that order.
                normalized_row = {
                    _normalize_header_name(k): v for k, v in row.items() if k is not None
                }
                profile_prices: List[float] = []
                for col_n in expected_normalized:
                    raw = normalized_row.get(col_n, "")
                    val = _to_money(raw)
                    if val is not None and any(token in col_n for token in ["MSRP", "DEALER", "PRICE", "COST"]):
                        profile_prices.append(val)
                if profile_prices:
                    unit_cost = _pick_unit_cost(profile_prices, dealer_tier)
                    if msrp is None:
                        msrp = profile_prices[0]
                else:
                    unit_cost = _to_money(
                        pick_value_by_patterns(row, ["STANDARD DEALER", "DEALER", "COST"])
                        or row.get("Dealer")
                        or row.get("Cost")
                    )
            else:
                unit_cost = _to_money(
                    pick_value_by_patterns(row, ["STANDARD DEALER", "DEALER", "COST"])
                    or row.get("Dealer")
                    or row.get("Cost")
                )

            parsed.append(
                {
                    "brand": (
                        pick_value_by_patterns(row, ["BRAND", "MANUFACTURER"])
                        or as_text(row.get("Brand"))
                        or as_text(row.get("Manufacturer"))
                        or "Unknown"
                    ).strip(),
                    "model": model or part,
                    "part_number": part,
                    "category": (
                        pick_value_by_patterns(row, ["CATEGORY"])
                        or as_text(row.get("Category"))
                        or "General"
                    ).strip(),
                    "short_description": (
                        pick_value_by_patterns(row, ["DESCRIPTION", "SHORT DESCRIPTION"])
                        or as_text(row.get("Description"))
                        or ""
                    ).strip()[:300],
                    "keywords": ", ".join(
                        [
                            k
                            for k in [
                                pick_value_by_patterns(row, ["BRAND", "MANUFACTURER"]) or as_text(row.get("Brand")),
                                model,
                                part,
                            ]
                            if k
                        ]
                    )[:280],
                    "unit_price": msrp,
                    "unit_cost": unit_cost,
                    "msrp": msrp,
                    "supplier": (as_text(row.get("Supplier")) or as_text(row.get("Brand"))).strip(),
                }
            )
        return parsed

    if suffix == ".pdf":
        text = _extract_text_from_pdf(path)
        return _parse_price_sheet_text(text, dealer_tier)

    raise RuntimeError(f"Unsupported file type: {suffix}. Use PDF, CSV, or XLSX.")


def _product_unique_key(brand: str, part_number: str, model: str) -> str:
    b = (brand or "unknown").strip().lower()
    p = (part_number or "").strip().lower()
    m = (model or "").strip().lower()
    return f"{b}::{p or m}"


def _extract_text_from_path(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return _extract_text_from_pdf(path)
    if suffix in {".txt", ".md", ".csv", ".json", ".yaml", ".yml", ".log"}:
        return path.read_text(encoding="utf-8", errors="ignore")
    if suffix in {".rtf"}:
        return path.read_text(encoding="utf-8", errors="ignore")
    return ""


def _extract_manual_digest_signals(text: str) -> Dict[str, Any]:
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    recommended: List[str] = []
    scope_lines: List[str] = []
    risk_lines: List[str] = []
    questions: List[str] = []

    for line in lines:
        low = line.lower()
        if any(tok in low for tok in ["recommend", "recommended", "should use", "suggested", "suggestion"]):
            recommended.append(line)
        if any(tok in low for tok in ["scope", "requirements", "deliverable", "rooms", "system includes"]):
            scope_lines.append(line)
        if any(tok in low for tok in ["risk", "constraint", "budget", "timeline", "deadline", "lead time"]):
            risk_lines.append(line)
        if "?" in line or any(tok in low for tok in ["confirm", "verify", "need to know", "clarify"]):
            questions.append(line)

    sku_matches = sorted(
        set(re.findall(r"\b[A-Z0-9#]+(?:-[A-Z0-9#]+){1,}\b", text.upper()))
    )
    sku_matches = [m for m in sku_matches if len(m) >= 5][:200]

    brands_found = []
    text_low = text.lower()
    for brand in KNOWN_DEVICE_BRANDS:
        if brand.lower() in text_low:
            brands_found.append(brand)

    emails = sorted(set(re.findall(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", text)))
    phones = sorted(set(re.findall(r"(?:\+?1[-.\s]?)?\(?\d{3}\)?[-.\s]\d{3}[-.\s]\d{4}", text)))

    return {
        "recommended_notes": recommended[:40],
        "scope_notes": scope_lines[:40],
        "risk_notes": risk_lines[:30],
        "open_questions": questions[:30],
        "detected_skus": sku_matches,
        "detected_brands": brands_found,
        "detected_contacts": {
            "emails": emails[:20],
            "phones": phones[:20],
        },
    }


def _extract_proposal_scope_signals(text: str) -> Dict[str, Any]:
    """
    Proposal-focused scope extraction for finished proposal uploads.
    Produces PM-ready buckets: scope, inclusions, exclusions, assumptions, risks.
    """
    if not text:
        return {
            "scope_of_work": [],
            "included_items": [],
            "excluded_items": [],
            "allowances": [],
            "assumptions": [],
            "schedule_notes": [],
            "risk_tags": [],
            "open_questions": [],
            "detected_skus": [],
            "detected_brands": [],
        }

    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    scope_of_work: List[str] = []
    included_items: List[str] = []
    excluded_items: List[str] = []
    allowances: List[str] = []
    assumptions: List[str] = []
    schedule_notes: List[str] = []
    risk_tags: set[str] = set()
    open_questions: List[str] = []

    for line in lines:
        low = line.lower()
        if len(line) > 260:
            continue

        if any(tok in low for tok in ["scope of work", "statement of work", "project scope", "work includes"]):
            scope_of_work.append(line)
        if any(tok in low for tok in ["included", "includes", "provide", "install", "furnish"]):
            included_items.append(line)
        if any(tok in low for tok in ["exclude", "not included", "excluded", "owner provided", "by others"]):
            excluded_items.append(line)
        if any(tok in low for tok in ["allowance", "allowances", "allow.", "t&m", "time and material"]):
            allowances.append(line)
            risk_tags.add("ALLOWANCE")
        if any(tok in low for tok in ["assumption", "assumes", "assuming", "subject to"]):
            assumptions.append(line)
            risk_tags.add("ASSUMPTION")
        if any(tok in low for tok in ["schedule", "timeline", "lead time", "duration", "completion", "milestone"]):
            schedule_notes.append(line)
        if any(tok in low for tok in ["change order", "contingency", "unforeseen", "permit", "inspection"]):
            risk_tags.add("SCOPE_RISK")
        if "?" in line or any(tok in low for tok in ["to be confirmed", "tbd", "verify", "confirm"]):
            open_questions.append(line)

    base = _extract_manual_digest_signals(text)
    proj_signals = _extract_project_signals_from_text(text)
    for tag in proj_signals.get("scope_risk_tags", []):
        risk_tags.add(tag)

    def _uniq(items: List[str], limit: int) -> List[str]:
        seen: set[str] = set()
        out: List[str] = []
        for item in items:
            key = item.lower()
            if key in seen:
                continue
            seen.add(key)
            out.append(item)
            if len(out) >= limit:
                break
        return out

    return {
        "scope_of_work": _uniq(scope_of_work or base.get("scope_notes", []), 40),
        "included_items": _uniq(included_items or base.get("recommended_notes", []), 60),
        "excluded_items": _uniq(excluded_items, 40),
        "allowances": _uniq(allowances, 30),
        "assumptions": _uniq(assumptions, 30),
        "schedule_notes": _uniq(schedule_notes, 30),
        "risk_tags": sorted(risk_tags)[:60],
        "open_questions": _uniq(open_questions or base.get("open_questions", []), 40),
        "detected_skus": base.get("detected_skus", [])[:100],
        "detected_brands": base.get("detected_brands", [])[:60],
    }


def _merge_digest_signals(items: List[Dict[str, Any]]) -> Dict[str, Any]:
    merged = {
        "recommended_notes": [],
        "scope_notes": [],
        "risk_notes": [],
        "open_questions": [],
        "detected_skus": [],
        "detected_brands": [],
        "detected_contacts": {"emails": [], "phones": []},
    }
    for item in items:
        for key in ["recommended_notes", "scope_notes", "risk_notes", "open_questions", "detected_skus", "detected_brands"]:
            merged[key].extend(item.get(key, []))
        contacts = item.get("detected_contacts", {})
        merged["detected_contacts"]["emails"].extend(contacts.get("emails", []))
        merged["detected_contacts"]["phones"].extend(contacts.get("phones", []))

    for key in ["recommended_notes", "scope_notes", "risk_notes", "open_questions", "detected_skus", "detected_brands"]:
        merged[key] = sorted(set(merged[key]))
    merged["detected_contacts"]["emails"] = sorted(set(merged["detected_contacts"]["emails"]))
    merged["detected_contacts"]["phones"] = sorted(set(merged["detected_contacts"]["phones"]))
    return merged


def _ai_project_digest_summary(project_name: str, merged: Dict[str, Any]) -> Dict[str, Any]:
    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not api_key:
        return {}
    model = os.environ.get("OPENAI_BACKUP_MODEL", "gpt-4.1-mini")
    payload = {
        "model": model,
        "temperature": 0.2,
        "messages": [
            {
                "role": "system",
                "content": (
                    "You are an AV integration project analyst. Output strict JSON with keys: "
                    "recommended_devices (array), key_findings (array), risks (array), "
                    "clarifying_questions (array), next_steps (array)."
                ),
            },
            {
                "role": "user",
                "content": json.dumps(
                    {
                        "project_name": project_name,
                        "digest": merged,
                    }
                ),
            },
        ],
        "response_format": {"type": "json_object"},
    }
    req = urllib.request.Request(
        "https://api.openai.com/v1/chat/completions",
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=45) as resp:
            raw = json.loads(resp.read().decode("utf-8"))
        content = (
            raw.get("choices", [{}])[0]
            .get("message", {})
            .get("content", "{}")
        )
        return json.loads(content)
    except Exception:
        return {}


def _session_memory_file(session_id: str) -> Path:
    safe = re.sub(r"[^a-zA-Z0-9_-]", "_", (session_id or "default"))[:80]
    return ASK_BOB_MEMORY_DIR / f"{safe}.jsonl"


def _append_ask_bob_memory(session_id: str, question: str, answer: str, source: str) -> None:
    path = _session_memory_file(session_id)
    with path.open("a", encoding="utf-8") as f:
        f.write(
            json.dumps(
                {
                    "timestamp": datetime.now().isoformat(),
                    "question": question,
                    "answer": answer,
                    "source": source,
                }
            )
            + "\n"
        )


def _load_recent_ask_bob_memory(session_id: str, limit: int = 8) -> List[Dict[str, Any]]:
    path = _session_memory_file(session_id)
    if not path.exists():
        return []
    lines = path.read_text(encoding="utf-8", errors="ignore").splitlines()
    out: List[Dict[str, Any]] = []
    for line in lines[-limit:]:
        try:
            out.append(json.loads(line))
        except Exception:
            continue
    return out


def _extract_project_hint(question: str) -> Optional[str]:
    q = question.strip()
    match = re.search(r"\b([A-Za-z][A-Za-z0-9_-]{2,})\s+project\b", q, flags=re.IGNORECASE)
    if match:
        return match.group(1)
    match = re.search(r"\bfor\s+([A-Za-z][A-Za-z0-9_-]{2,})\b", q, flags=re.IGNORECASE)
    if match:
        return match.group(1)
    return None


def _collect_project_file_context(project_hint: str, question: str, max_files: int = 16) -> Dict[str, Any]:
    """
    Scan likely project-related directories for files matching project hint
    and collect concise snippets for Ask Bob context.
    """
    hint = (project_hint or "").strip()
    if not hint:
        return {"project_hint": "", "files": [], "context": ""}

    candidates: List[Path] = []
    roots = [
        BASE_DIR / "knowledge" / "projects",
        BASE_DIR / "knowledge" / "generated_proposals",
        BASE_DIR / "knowledge" / "proposals",
        BASE_DIR / "data" / "manual_digest",
        BASE_DIR / "knowledge" / "cortex",
    ]
    hint_low = hint.lower()
    for root in roots:
        if not root.exists():
            continue
        for path in root.rglob("*"):
            if not path.is_file():
                continue
            rel = str(path.relative_to(BASE_DIR)).lower()
            if hint_low in rel:
                candidates.append(path)

    # Prefer digest JSON + text docs.
    candidates = sorted(
        candidates,
        key=lambda p: (
            0 if p.name == "digest.json" else 1,
            0 if p.suffix.lower() in {".md", ".txt", ".json", ".csv"} else 1,
            str(p),
        ),
    )[: max_files]

    snippets: List[str] = []
    files_used: List[str] = []
    q_terms = [t for t in re.findall(r"[A-Za-z0-9]{3,}", question.lower()) if t not in {"what", "which", "with", "from", "that", "this"}]

    for path in candidates:
        rel = str(path.relative_to(BASE_DIR))
        try:
            text = _extract_text_from_path(path)
            if not text:
                continue
            lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
            if not lines:
                continue
            picked: List[str] = []
            for ln in lines:
                low = ln.lower()
                if any(term in low for term in q_terms):
                    picked.append(ln)
                if len(picked) >= 4:
                    break
            if not picked:
                picked = lines[:3]
            snippet = "\n".join(picked)
            snippets.append(f"[{rel}]\n{snippet}")
            files_used.append(rel)
        except Exception:
            continue

    context = "\n\n".join(snippets)[:12000]
    return {
        "project_hint": hint,
        "files": files_used,
        "context": context,
    }


def _init_product_db(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS approved_batches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at TEXT NOT NULL,
            source_file TEXT,
            parse_profile TEXT,
            dealer_tier TEXT,
            notes TEXT,
            total_products INTEGER NOT NULL DEFAULT 0
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS approved_products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_id INTEGER,
            unique_key TEXT NOT NULL UNIQUE,
            brand TEXT,
            model TEXT NOT NULL,
            part_number TEXT,
            category TEXT,
            short_description TEXT,
            keywords TEXT,
            unit_price REAL,
            unit_cost REAL,
            msrp REAL,
            supplier TEXT,
            approved_at TEXT NOT NULL,
            FOREIGN KEY(batch_id) REFERENCES approved_batches(id)
        )
        """
    )
    conn.commit()


def _safe_export_filename(value: str) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9._-]+", "_", (value or "").strip())
    return cleaned.strip("._") or "export"


def get_launchd_job_status(label: str) -> Dict:
    """Return launchd status details for a given label."""
    status = {
        "label": label,
        "loaded": False,
        "running": False,
        "pid": None,
        "last_exit_code": None,
        "plist_exists": False,
        "plist_path": None,
    }
    plist_candidates = [
        Path.home() / "Library" / "LaunchAgents" / f"{label}.plist",
        Path("/Library/LaunchAgents") / f"{label}.plist",
        Path("/Library/LaunchDaemons") / f"{label}.plist",
    ]
    for candidate in plist_candidates:
        if candidate.exists():
            status["plist_exists"] = True
            status["plist_path"] = str(candidate)
            break

    try:
        result = subprocess.run(
            ["launchctl", "list"],
            capture_output=True, text=True, timeout=5
        )
        if result.returncode != 0:
            status["error"] = result.stderr.strip() or "launchctl list failed"
            return status

        for line in result.stdout.splitlines():
            if label not in line:
                continue
            parts = line.split()
            if len(parts) < 3:
                continue
            # launchctl list format: PID  LastExitStatus  Label
            pid_part = parts[0]
            exit_part = parts[1]
            job_label = parts[-1]
            if job_label != label:
                continue

            status["loaded"] = True
            if pid_part != "-":
                try:
                    status["pid"] = int(pid_part)
                    status["running"] = True
                except ValueError:
                    status["pid"] = None
            if exit_part != "-":
                try:
                    status["last_exit_code"] = int(exit_part)
                except ValueError:
                    status["last_exit_code"] = None
            return status
    except Exception as e:
        status["error"] = str(e)
    return status


def run_trading_api_watchdog() -> Dict:
    """Kill stale :8421 listeners and kickstart trading API launchd service."""
    label = "com.symphony.trading-api"
    user_id = str(os.getuid())
    steps: List[Dict] = []

    # Kill any direct python process for trading_api.py.
    pkill_result = run_command(["pkill", "-f", "api/trading_api.py"], timeout=10)
    steps.append({"step": "pkill trading_api.py", "result": pkill_result})

    # Kill stale listener on port 8421 if present.
    listener_pid: Optional[int] = None
    lsof_cmd = ["/usr/sbin/lsof", "-t", "-iTCP:8421", "-sTCP:LISTEN"]
    lsof_result = run_command(lsof_cmd, timeout=10)
    if (not lsof_result.get("success")) and "No such file or directory" in str(lsof_result.get("error", "")):
        lsof_result = run_command(["lsof", "-t", "-iTCP:8421", "-sTCP:LISTEN"], timeout=10)
    if lsof_result.get("success") and lsof_result.get("output"):
        raw_pid = (lsof_result.get("output", "").splitlines() or [""])[0].strip()
        try:
            listener_pid = int(raw_pid)
            kill_result = run_command(["kill", "-9", str(listener_pid)], timeout=10)
            steps.append({
                "step": "kill stale port 8421 listener",
                "pid": listener_pid,
                "result": kill_result,
            })
        except ValueError:
            steps.append({
                "step": "kill stale port 8421 listener",
                "result": {"success": False, "error": f"Unexpected PID output: {raw_pid}"},
            })
    else:
        steps.append({"step": "check stale port 8421 listener", "result": lsof_result})

    # Restart launchd service for trading API.
    kickstart_result = run_command(
        ["launchctl", "kickstart", "-k", f"gui/{user_id}/{label}"],
        timeout=15,
    )
    steps.append({"step": "launchctl kickstart trading-api", "result": kickstart_result})

    # Verify health endpoint.
    health_result = run_command(
        [
            "python3",
            "-c",
            (
                "import time, urllib.request\n"
                "ok=False\n"
                "last=''\n"
                "for _ in range(8):\n"
                "    try:\n"
                "        urllib.request.urlopen('http://127.0.0.1:8421/health', timeout=2)\n"
                "        ok=True\n"
                "        break\n"
                "    except Exception as e:\n"
                "        last=str(e)\n"
                "        time.sleep(1)\n"
                "print('healthy' if ok else last)\n"
                "raise SystemExit(0 if ok else 1)\n"
            ),
        ],
        timeout=20,
    )
    steps.append({"step": "verify trading api health", "result": health_result})

    job = get_launchd_job_status(label)
    return {
        "success": bool(health_result.get("success")),
        "timestamp": datetime.now().isoformat(),
        "label": label,
        "job": job,
        "stale_listener_pid": listener_pid,
        "steps": steps,
    }


def get_service_status() -> List[Dict]:
    """Get status of all Symphony services."""
    services = []
    
    # Check common ports
    port_services = [
        (3000, "Voice Receptionist", "voice"),
        (5678, "Bob Orchestrator", "orchestrator"),
        (8080, "Mission Control", "dashboard"),
        (8091, "Symphony Markup", "markup"),
        (11434, "Ollama", "ollama"),
        (8420, "Mobile API", "api"),
    ]
    
    import socket
    for port, name, key in port_services:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(1)
        result = sock.connect_ex(('127.0.0.1', port))
        sock.close()
        
        services.append({
            "name": name,
            "key": key,
            "port": port,
            "status": "running" if result == 0 else "stopped",
            "url": f"http://localhost:{port}" if result == 0 else None
        })
    
    # Check launchd jobs
    try:
        result = subprocess.run(
            ["launchctl", "list"],
            capture_output=True, text=True, timeout=5
        )
        running_jobs = result.stdout if result.returncode == 0 else ""
        
        symphony_jobs = [
            "com.symphony.morning-checklist",
            "com.symphony.daily-digest",
            "com.symphony.subscription-audit",
            "com.symphony.watcher",
            "com.symphony.memory-guard",
        ]
        
        for job in symphony_jobs:
            is_running = job in running_jobs
            services.append({
                "name": job.replace("com.symphony.", "").replace("-", " ").title(),
                "key": job,
                "type": "scheduled",
                "status": "loaded" if is_running else "unloaded"
            })
    except:
        pass
    
    return services


def get_quick_stats() -> Dict:
    """Get quick stats for dashboard."""
    stats = {
        "timestamp": datetime.now().isoformat(),
        "bids": {"new": 0, "pending": 0},
        "proposals": {"draft": 0, "sent": 0, "accepted": 0},
        "invoices": {"pending": 0, "overdue": 0, "paid_this_month": 0},
        "cortex": {"articles": 0, "size_kb": 0},
        "subscriptions": {"monthly_total": 0, "count": 0}
    }
    
    # Count proposals
    proposals_dir = BASE_DIR / "knowledge" / "projects"
    if proposals_dir.exists():
        for proj in proposals_dir.iterdir():
            if proj.is_dir():
                pf = proj / "proposal.json"
                if pf.exists():
                    try:
                        data = json.loads(pf.read_text())
                        status = data.get("status", "draft")
                        if status == "draft":
                            stats["proposals"]["draft"] += 1
                        elif status == "sent":
                            stats["proposals"]["sent"] += 1
                        elif status == "accepted":
                            stats["proposals"]["accepted"] += 1
                    except:
                        pass
    
    # Count cortex articles
    cortex_dir = BASE_DIR / "knowledge" / "cortex"
    if cortex_dir.exists():
        total_size = 0
        for f in cortex_dir.rglob("*.json"):
            stats["cortex"]["articles"] += 1
            total_size += f.stat().st_size
        stats["cortex"]["size_kb"] = round(total_size / 1024, 1)
    
    # Subscriptions
    subs_file = BASE_DIR / "knowledge" / "subscriptions.json"
    if subs_file.exists():
        try:
            data = json.loads(subs_file.read_text())
            subs = data.get("subscriptions", [])
            stats["subscriptions"]["count"] = len(subs)
            monthly = 0
            for s in subs:
                cost = s.get("cost", 0)
                if s.get("billing_cycle") == "yearly":
                    cost = cost / 12
                monthly += cost
            stats["subscriptions"]["monthly_total"] = round(monthly, 2)
        except:
            pass
    
    return stats


if HAS_FASTAPI:
    # --- API Endpoints ---
    
    @app.get("/")
    async def root():
        """API root - health check."""
        return {
            "name": "Symphony AI Mobile API",
            "version": "1.0.0",
            "status": "running",
            "timestamp": datetime.now().isoformat()
        }
    
    @app.get("/health")
    async def health():
        """Detailed health check."""
        return {
            "status": "healthy",
            "services": get_service_status(),
            "timestamp": datetime.now().isoformat()
        }

    @app.get("/templates/manual_digest_intake_pdf")
    async def get_manual_digest_intake_pdf():
        """Download fillable PDF intake template for clients/builders."""
        pdf_path = BASE_DIR / "docs" / "templates" / "MANUAL_DIGEST_CLIENT_BUILDER_INTAKE_FILLABLE.pdf"
        if not pdf_path.exists():
            raise HTTPException(status_code=404, detail="Template not found")
        return FileResponse(
            path=str(pdf_path),
            media_type="application/pdf",
            filename="Symphony_Project_Intake_Fillable.pdf",
        )
    
    @app.get("/dashboard")
    async def dashboard():
        """Main dashboard data."""
        return {
            "stats": get_quick_stats(),
            "services": get_service_status(),
            "timestamp": datetime.now().isoformat()
        }
    
    @app.get("/stats")
    async def stats():
        """Quick stats only."""
        return get_quick_stats()
    
    @app.get("/services")
    async def services():
        """Service status list."""
        return {"services": get_service_status()}
    
    # --- Bids ---
    
    @app.get("/bids")
    async def get_bids():
        """Get BuildingConnected bids."""
        result = run_command([
            "python3", str(BASE_DIR / "orchestrator" / "core" / "bob_orchestrator.py"),
            "bid_check"
        ])
        return result
    
    @app.get("/bids/list")
    async def list_bids():
        """List all bid invitations."""
        result = run_command([
            "python3", str(BASE_DIR / "orchestrator" / "core" / "bob_orchestrator.py"),
            "bid_list"
        ])
        return result
    
    # --- Proposals ---
    
    @app.get("/proposals")
    async def get_proposals():
        """List all proposals."""
        proposals = []
        proposals_dir = BASE_DIR / "knowledge" / "projects"
        if proposals_dir.exists():
            for proj in proposals_dir.iterdir():
                if proj.is_dir():
                    pf = proj / "proposal.json"
                    if pf.exists():
                        try:
                            data = json.loads(pf.read_text())
                            proposals.append({
                                "id": proj.name,
                                "client": data.get("client_name", "Unknown"),
                                "status": data.get("status", "draft"),
                                "total": data.get("total", 0),
                                "created": data.get("created", "")
                            })
                        except:
                            pass
        return {"proposals": proposals}
    
    class QuoteRequest(BaseModel):
        client: str
        description: str
    
    @app.post("/proposals/create")
    async def create_proposal(request: QuoteRequest):
        """Create a new proposal."""
        result = run_command([
            "python3", str(BASE_DIR / "tools" / "smart_proposal.py"),
            "--client", request.client,
            "--description", request.description
        ], timeout=60)
        return result
    
    # --- Knowledge ---
    
    @app.get("/cortex/stats")
    async def cortex_stats():
        """Get cortex statistics."""
        result = run_command([
            "bash", str(BASE_DIR / "tools" / "cortex_status.sh")
        ])
        return result
    
    class ResearchRequest(BaseModel):
        query: str
    
    class FactsLearnRequest(BaseModel):
        text: str
        category: str = "general"
        learn_now: bool = False
        curate_now: bool = True

    class CuratorRunRequest(BaseModel):
        limit: int = 0
        force: bool = False
        contains: Optional[str] = None

    class CuratorFactStatusRequest(BaseModel):
        fact_ids: List[int]
        status: str
    
    @app.post("/facts/learn")
    async def facts_learn(request: FactsLearnRequest):
        """Ingest pasted facts (e.g. C4 driver info) into cortex for learning."""
        try:
            from tools.facts_ingest import ingest, CATEGORIES
            cat = request.category if request.category in CATEGORIES else "general"
            result = ingest(request.text.strip(), category=cat, learn_now=request.learn_now)

            # Keep curation loop tight: new facts should be scored immediately.
            if result.get("success") and request.curate_now:
                try:
                    from tools.cortex_curator import run_curator
                    curated = run_curator(limit=1, contains=result.get("path"), force=True)
                    result["curator"] = {
                        "indexed_files": curated.get("indexed_files", 0),
                        "new_facts": curated.get("new_facts", 0),
                        "updated_facts": curated.get("updated_facts", 0),
                    }
                except Exception as ce:
                    result["curator_warning"] = str(ce)

            return result
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))
    
    @app.get("/facts/categories")
    async def facts_categories():
        """List available fact categories."""
        from tools.facts_ingest import CATEGORIES
        return {"categories": CATEGORIES}

    @app.post("/cortex/curator/run")
    async def run_cortex_curator(request: CuratorRunRequest):
        """Run Cortex Curator pipeline: dedupe, confidence scoring, contradiction checks."""
        try:
            from tools.cortex_curator import run_curator
            result = run_curator(
                limit=request.limit if request.limit > 0 else None,
                force=request.force,
                contains=request.contains,
            )
            return result
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.get("/cortex/curator/status")
    async def cortex_curator_status():
        """Get curator status, review queue, and trusted/review counts."""
        try:
            from tools.cortex_curator import get_curator_status
            return get_curator_status()
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.get("/cortex/curator/review")
    async def cortex_curator_review(
        status: str = "review",
        limit: int = 50,
        offset: int = 0,
        min_confidence: float = -1.0,
        min_professional: float = 0.25,
        subject: str = "",
    ):
        """List curation queue with smart-home reasoning/troubleshooting scores."""
        try:
            from tools.cortex_curator import list_review_facts
            status = status if status in {"review", "trusted"} else "review"
            return list_review_facts(
                status=status,
                limit=max(1, min(limit, 200)),
                offset=max(0, offset),
                min_confidence=(None if min_confidence < 0 else float(min_confidence)),
                min_professional_score=(None if min_professional < 0 else float(min_professional)),
                subject_contains=(subject or None),
            )
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/cortex/curator/facts/status")
    async def cortex_curator_set_status(request: CuratorFactStatusRequest):
        """Manually set fact status to trusted/review by IDs."""
        try:
            from tools.cortex_curator import set_fact_status
            desired = request.status.strip().lower()
            return set_fact_status(request.fact_ids, desired)
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/cortex/curator/promote")
    async def cortex_curator_promote(request: CuratorFactStatusRequest):
        """Promote selected facts to trusted."""
        try:
            from tools.cortex_curator import set_fact_status
            return set_fact_status(request.fact_ids, "trusted")
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/cortex/curator/demote")
    async def cortex_curator_demote(request: CuratorFactStatusRequest):
        """Demote selected facts to review."""
        try:
            from tools.cortex_curator import set_fact_status
            return set_fact_status(request.fact_ids, "review")
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.get("/memory_guard/status")
    async def memory_guard_status():
        """Get launchd status for smart memory guard."""
        try:
            return {
                "success": True,
                "timestamp": datetime.now().isoformat(),
                "job": get_launchd_job_status("com.symphony.memory-guard"),
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.get("/notes/pipeline_status")
    async def notes_pipeline_status():
        """Get health/status of the Notes ingestion pipeline."""
        try:
            notes_watcher_job = get_launchd_job_status("com.symphony.notes-watcher")
            incoming_tasks_job = get_launchd_job_status("com.symphony.incoming-tasks")
            notes_sync_job = get_launchd_job_status("com.symphony.notes-sync-photos")

            watcher_state = _read_json_file(BASE_DIR / "knowledge" / "state" / "notes_watcher_state.json")
            incoming_state = _read_json_file(BASE_DIR / "knowledge" / "state" / "incoming_tasks_state.json")
            processed_tasks = _read_json_file(BASE_DIR / "knowledge" / "state" / "processed_tasks.json")

            logs = {
                "notes_watcher_log": str(BASE_DIR / "logs" / "notes-watcher.log"),
                "notes_watcher_error_log": str(BASE_DIR / "logs" / "notes-watcher.error.log"),
                "incoming_tasks_log": str(BASE_DIR / "logs" / "incoming-tasks.log"),
                "incoming_tasks_error_log": str(BASE_DIR / "logs" / "incoming-tasks.error.log"),
                "notes_sync_log": str(BASE_DIR / "logs" / "notes-sync-photos.log"),
                "notes_sync_error_log": str(BASE_DIR / "logs" / "notes-sync-photos.error.log"),
            }
            log_mtimes = {}
            for key, file_path in logs.items():
                p = Path(file_path)
                if p.exists():
                    log_mtimes[key + "_updated_at"] = datetime.fromtimestamp(p.stat().st_mtime).isoformat()

            return {
                "success": True,
                "timestamp": datetime.now().isoformat(),
                "jobs": {
                    "notes_watcher": notes_watcher_job,
                    "incoming_tasks": incoming_tasks_job,
                    "notes_sync_photos": notes_sync_job,
                },
                "state": {
                    "notes_watcher_last_check": watcher_state.get("last_check"),
                    "notes_watcher_processed_count": len(watcher_state.get("processed_notes", {})),
                    "notes_watcher_known_projects": len(watcher_state.get("known_projects", [])),
                    "incoming_tasks_last_check": incoming_state.get("last_check"),
                    "incoming_tasks_processed_count": len(incoming_state.get("processed_notes", {})),
                    "incoming_tasks_completed_total": len(processed_tasks) if isinstance(processed_tasks, list) else 0,
                },
                "logs": {**logs, **log_mtimes},
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    class NotesProcessNowRequest(BaseModel):
        note_id: Optional[int] = None
        project_name: Optional[str] = None
        sync_media: Optional[bool] = True
        run_incoming_tasks: Optional[bool] = True

    class IMessageWatchlistRequest(BaseModel):
        numbers: List[str]
        monitor_all: Optional[bool] = False

    class IMessageAutomationRequest(BaseModel):
        create_service_invoice_drafts: Optional[bool] = None
        create_appointment_drafts: Optional[bool] = None

    class IMessageBackfillRequest(BaseModel):
        weeks: int = 4
        dry_run: bool = True
        limit: int = 5000

    class IMessageIntakeActionRequest(BaseModel):
        draft_id: str
        note: Optional[str] = ""
        proposed_start: Optional[str] = None
        duration_min: Optional[int] = 60
        send_confirmation: Optional[bool] = True
        confirmation_message: Optional[str] = None

    class IMessageConfirmationRequest(BaseModel):
        kind: str
        draft_id: str
        message: Optional[str] = None

    class IMessageRetryAllFailuresRequest(BaseModel):
        limit: int = 25

    class ContactsLookupRequest(BaseModel):
        query: str

    class ClientContactAddRequest(BaseModel):
        name: str
        phones: List[str] = []
        emails: List[str] = []
        notes: Optional[str] = ""
        auto_monitor: Optional[bool] = True

    class ClientContactRemoveRequest(BaseModel):
        client_id: str
        remove_from_watchlist: Optional[bool] = True

    class ContactsSeedWatchlistRequest(BaseModel):
        include_clients_registry: bool = True
        include_contacts_index: bool = True
        include_emails: bool = True
        overwrite: bool = False

    class OpsRecoveryRunRequest(BaseModel):
        apply: bool = False
        threshold: float = float(os.environ.get("AUTOPILOT_CONFIDENCE_THRESHOLD", "0.80"))
        playbook: Optional[str] = None

    class NetworkDropoutWatchStartRequest(BaseModel):
        gateway_ip: str = "192.168.1.1"
        wan_ip: str = "1.1.1.1"
        control4_ip: Optional[str] = None
        sonos_ip: Optional[str] = None
        interval_sec: float = 2.0

    class DropoutIncidentCreateRequest(BaseModel):
        event_timestamp: Optional[str] = None
        event_name: Optional[str] = None
        notes: Optional[str] = None
        priority: Optional[str] = "high"

    class InventoryRebuildRequest(BaseModel):
        force: Optional[bool] = False
        low_stock_limit: Optional[int] = 25
        top_limit: Optional[int] = 100

    class IntegrationBriefRequest(BaseModel):
        no_cache: Optional[bool] = False

    class EmployeeBotCreateRequest(BaseModel):
        worker_id: str
        name: str
        role: str
        emoji: Optional[str] = "🤖"
        skills: List[str]
        intro: Optional[str] = ""
        token_env: Optional[str] = ""

    class EmployeeBotWorkerRequest(BaseModel):
        worker_id: str

    class NotesProjectLinkUpsertRequest(BaseModel):
        match_text: str
        project_name: str
        enabled: Optional[bool] = True

    class NotesTaskApprovalRejectRequest(BaseModel):
        reason: Optional[str] = ""

    @app.post("/notes/process_now")
    async def notes_process_now(request: NotesProcessNowRequest):
        """Force immediate notes processing and optional media/task sync."""
        try:
            steps: List[Dict[str, Any]] = []

            if request.note_id is not None:
                notes_result = run_tool_endpoint(
                    "notes_watcher.py",
                    ["--process-note", str(request.note_id)],
                    timeout=120,
                )
                steps.append({"step": "notes_watcher_process_note", **notes_result})
            elif request.project_name:
                notes_result = run_tool_endpoint(
                    "notes_watcher.py",
                    ["--process-project", request.project_name],
                    timeout=150,
                )
                steps.append({"step": "notes_watcher_process_project", **notes_result})
            else:
                notes_result = run_tool_endpoint(
                    "notes_watcher.py",
                    ["--check"],
                    timeout=120,
                )
                steps.append({"step": "notes_watcher_check", **notes_result})

            if request.sync_media:
                media_result = run_tool_endpoint(
                    "notes_sync.py",
                    ["--sync-photos"],
                    timeout=180,
                )
                steps.append({"step": "notes_sync_photos", **media_result})

            if request.run_incoming_tasks:
                incoming_result = run_command(
                    [
                        "python3",
                        str(BASE_DIR / "orchestrator" / "incoming_task_processor.py"),
                        "--check",
                    ],
                    timeout=180,
                )
                steps.append({"step": "incoming_tasks_check", **incoming_result})

            success = all(step.get("success", False) for step in steps)
            return {
                "success": success,
                "timestamp": datetime.now().isoformat(),
                "note_id": request.note_id,
                "project_name": request.project_name,
                "steps": steps,
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.get("/notes/project_links")
    async def notes_project_links():
        """List explicit note->project linking rules used by notes_watcher."""
        try:
            data = _read_json_file(NOTES_PROJECT_LINKS_FILE)
            rules = data.get("rules", []) if isinstance(data, dict) else []
            return {"success": True, "count": len(rules), "rules": rules}
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/notes/project_links")
    async def notes_project_links_upsert(request: NotesProjectLinkUpsertRequest):
        """Add a note linking rule: if note contains match_text -> project_name."""
        try:
            match_text = (request.match_text or "").strip()
            project_name = (request.project_name or "").strip()
            if not match_text or not project_name:
                raise HTTPException(status_code=400, detail="match_text and project_name are required")

            NOTES_PROJECT_LINKS_FILE.parent.mkdir(parents=True, exist_ok=True)
            data = _read_json_file(NOTES_PROJECT_LINKS_FILE)
            rules = data.get("rules", []) if isinstance(data, dict) else []
            # Upsert by match_text exact lower.
            normalized = match_text.lower()
            updated = False
            for r in rules:
                if str(r.get("match_text", "")).strip().lower() == normalized:
                    r["project_name"] = project_name
                    r["enabled"] = bool(request.enabled)
                    r["updated_at"] = datetime.now().isoformat()
                    updated = True
                    break
            if not updated:
                rules.append(
                    {
                        "id": f"npl_{uuid4().hex[:10]}",
                        "match_text": match_text,
                        "project_name": project_name,
                        "enabled": bool(request.enabled),
                        "created_at": datetime.now().isoformat(),
                    }
                )
            NOTES_PROJECT_LINKS_FILE.write_text(json.dumps({"rules": rules}, indent=2), encoding="utf-8")
            return {"success": True, "count": len(rules), "rules": rules}
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.get("/notes/task_approvals")
    async def notes_task_approvals(status: str = "pending_approval", limit: int = 50):
        """List notes-to-task approvals queue."""
        try:
            sys.path.insert(0, str(BASE_DIR / "tools"))
            from notes_watcher import list_task_approvals
            return list_task_approvals(status=status, limit=max(1, min(limit, 200)))
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/notes/task_approvals/{approval_id}/approve")
    async def notes_task_approvals_approve(approval_id: str):
        """Approve note follow-up and create task board task."""
        try:
            sys.path.insert(0, str(BASE_DIR / "tools"))
            from notes_watcher import approve_task_approval
            return approve_task_approval(approval_id)
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/notes/task_approvals/{approval_id}/reject")
    async def notes_task_approvals_reject(approval_id: str, request: NotesTaskApprovalRejectRequest):
        """Reject note follow-up approval item."""
        try:
            sys.path.insert(0, str(BASE_DIR / "tools"))
            from notes_watcher import reject_task_approval
            return reject_task_approval(approval_id, reason=(request.reason or ""))
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.get("/imessages/status")
    async def imessages_status():
        """Get iMessage watcher status and monitoring configuration."""
        try:
            watcher_job = get_launchd_job_status("com.symphony.imessage-watcher")
            watcher_state = _read_json_file(IMESSAGE_WATCHER_STATE_FILE)
            script_result = run_tool_endpoint("imessage_watcher.py", ["--status"], timeout=30)
            script_json = _extract_json_from_output(script_result, fallback_key="output")
            return {
                "success": True,
                "timestamp": datetime.now().isoformat(),
                "job": watcher_job,
                "state": watcher_state,
                "watcher": script_json,
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.get("/imessages/watchlist")
    async def imessages_watchlist():
        """Get monitored numbers/emails for iMessage pipeline."""
        try:
            state = _read_json_file(IMESSAGE_WATCHER_STATE_FILE)
            return {
                "success": True,
                "watchlist": state.get("watchlist", []),
                "watchlist_count": len(state.get("watchlist", [])),
                "monitor_all": False,
                "mode": "watchlist_plus_keywords",
                "automation": state.get("automation", {}),
                "keyword_discovery_enabled": bool(state.get("keyword_discovery_enabled", True)),
                "work_signal_threshold": max(1, int(state.get("work_signal_threshold", 2))),
                "last_check": state.get("last_check"),
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/imessages/watchlist")
    async def imessages_set_watchlist(request: IMessageWatchlistRequest):
        """Set monitored numbers/emails for iMessage watcher."""
        try:
            numbers = request.numbers or []
            args: List[str] = []
            if numbers:
                args = ["--set-watchlist", ",".join(numbers)]
            else:
                args = ["--clear-watchlist"]
            result = run_tool_endpoint("imessage_watcher.py", args, timeout=30)
            payload = _extract_json_from_output(result, fallback_key="output")
            payload.setdefault("command_success", result.get("success", False))
            payload["mode"] = "watchlist_plus_keywords"
            payload["monitor_all"] = False
            return payload
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.get("/imessages/automation")
    async def imessages_automation_status():
        """Get iMessage automation toggles for invoice/scheduling drafts."""
        try:
            state = _read_json_file(IMESSAGE_WATCHER_STATE_FILE)
            automation = state.get("automation", {}) if isinstance(state, dict) else {}
            return {
                "success": True,
                "automation": {
                    "create_service_invoice_drafts": bool(automation.get("create_service_invoice_drafts", True)),
                    "create_appointment_drafts": bool(automation.get("create_appointment_drafts", True)),
                },
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/imessages/automation")
    async def imessages_set_automation(request: IMessageAutomationRequest):
        """Update iMessage automation toggles for draft generation."""
        try:
            args: List[str] = []
            if request.create_service_invoice_drafts is not None:
                args.extend(["--set-auto-invoice-drafts", "true" if request.create_service_invoice_drafts else "false"])
            if request.create_appointment_drafts is not None:
                args.extend(["--set-auto-appointment-drafts", "true" if request.create_appointment_drafts else "false"])
            if not args:
                return {"success": True, "message": "No changes provided"}
            result = run_tool_endpoint("imessage_watcher.py", args, timeout=30)
            payload = _extract_json_from_output(result, fallback_key="output")
            payload.setdefault("command_success", result.get("success", False))
            return payload
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/imessages/process_now")
    async def imessages_process_now():
        """Force immediate iMessage ingestion pass."""
        try:
            result = run_tool_endpoint("imessage_watcher.py", ["--check"], timeout=90)
            payload = _extract_json_from_output(result, fallback_key="output")
            payload.setdefault("command_success", result.get("success", False))
            return payload
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/imessages/backfill")
    async def imessages_backfill(request: IMessageBackfillRequest):
        """Process historical iMessages from the last N weeks."""
        try:
            weeks = max(1, min(int(request.weeks or 4), 26))
            limit = max(1, min(int(request.limit or 5000), 50000))
            args: List[str] = ["--backfill-weeks", str(weeks), "--limit", str(limit)]
            if bool(request.dry_run):
                args.append("--dry-run")
            result = run_tool_endpoint("imessage_watcher.py", args, timeout=180)
            payload = _extract_json_from_output(result, fallback_key="output")
            payload.setdefault("command_success", result.get("success", False))
            return payload
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.get("/imessages/recent")
    async def imessages_recent(limit: int = 20):
        """Recent parsed work texts with task mapping info for live feed."""
        try:
            items = _read_recent_jsonl(IMESSAGE_WORK_LOG_FILE, limit=limit)
            feed = []
            for item in items:
                feed.append(
                    {
                        "timestamp": item.get("timestamp"),
                        "rowid": item.get("rowid"),
                        "direction": item.get("direction"),
                        "handle": item.get("handle"),
                        "contact_name": item.get("contact_name"),
                        "linked_projects": item.get("linked_projects", []),
                        "text": item.get("text"),
                        "task_id": item.get("task_id"),
                    }
                )
            return {
                "success": True,
                "count": len(feed),
                "items": feed,
                "source_file": str(IMESSAGE_WORK_LOG_FILE),
                "timestamp": datetime.now().isoformat(),
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.get("/imessages/intake")
    async def imessages_intake(status: str = "draft", limit: int = 50):
        """Unified queue for invoice + appointment drafts with action metadata."""
        try:
            capped = max(1, min(int(limit), 500))
            target_status = (status or "draft").strip().lower()
            invoice_rows = _read_all_jsonl(IMESSAGE_INVOICE_DRAFTS_FILE)
            appt_rows = _read_all_jsonl(IMESSAGE_APPOINTMENT_DRAFTS_FILE)
            items: List[Dict[str, Any]] = []

            for row in invoice_rows:
                row_status = str(row.get("status", "draft")).strip().lower()
                if target_status != "all" and row_status != target_status:
                    continue
                items.append(
                    {
                        "kind": "invoice",
                        "draft_id": row.get("draft_id"),
                        "status": row.get("status", "draft"),
                        "created_at": row.get("created_at"),
                        "rowid": row.get("rowid"),
                        "handle_masked": row.get("handle_masked"),
                        "contact_name_masked": row.get("contact_name_masked"),
                        "project_hint": row.get("project_hint"),
                        "request_text_redacted": row.get("request_text_redacted"),
                        "linked_projects": row.get("linked_projects", []),
                        "last_action_at": row.get("last_action_at"),
                    }
                )

            for row in appt_rows:
                row_status = str(row.get("status", "draft")).strip().lower()
                if target_status != "all" and row_status != target_status:
                    continue
                items.append(
                    {
                        "kind": "appointment",
                        "draft_id": row.get("draft_id"),
                        "status": row.get("status", "draft"),
                        "created_at": row.get("created_at"),
                        "rowid": row.get("rowid"),
                        "handle_masked": row.get("handle_masked"),
                        "contact_name_masked": row.get("contact_name_masked"),
                        "project_hint": row.get("project_hint"),
                        "request_text_redacted": row.get("request_text_redacted"),
                        "linked_projects": row.get("linked_projects", []),
                        "proposed_start": row.get("proposed_start"),
                        "calendar_event_created_at": row.get("calendar_event_created_at"),
                        "last_action_at": row.get("last_action_at"),
                    }
                )

            items.sort(key=lambda x: str(x.get("created_at") or ""), reverse=True)
            items = items[:capped]
            return {
                "success": True,
                "count": len(items),
                "status_filter": target_status,
                "items": items,
                "invoice_file": str(IMESSAGE_INVOICE_DRAFTS_FILE),
                "appointment_file": str(IMESSAGE_APPOINTMENT_DRAFTS_FILE),
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/imessages/intake/invoice/approve")
    async def imessages_intake_invoice_approve(request: IMessageIntakeActionRequest):
        """Approve invoice draft and optionally send confirmation text."""
        try:
            payload = {
                "draft_id": request.draft_id,
                "note": request.note or "",
                "send_confirmation": bool(request.send_confirmation),
                "confirmation_message": request.confirmation_message,
            }
            result = _execute_invoice_approve(payload)
            _append_intake_audit(
                action="invoice_approve",
                kind="invoice",
                draft_id=str(request.draft_id or ""),
                success=bool(result.get("success")),
                payload=payload,
                result=result,
                error=str(result.get("error") or ""),
            )
            if not result.get("success"):
                failure_id = _enqueue_intake_failure(
                    action="invoice_approve",
                    kind="invoice",
                    draft_id=str(request.draft_id or ""),
                    payload=payload,
                    error=str(result.get("error") or "unknown failure"),
                )
                result["failure_id"] = failure_id
            return result
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/imessages/intake/appointment/calendar")
    async def imessages_intake_appointment_calendar(request: IMessageIntakeActionRequest):
        """Create a calendar event from appointment draft and optionally send confirmation."""
        try:
            payload = {
                "draft_id": request.draft_id,
                "proposed_start": request.proposed_start,
                "duration_min": int(request.duration_min or 60),
                "send_confirmation": bool(request.send_confirmation),
                "confirmation_message": request.confirmation_message,
            }
            result = _execute_appointment_schedule(payload)
            _append_intake_audit(
                action="appointment_calendar",
                kind="appointment",
                draft_id=str(request.draft_id or ""),
                success=bool(result.get("success")),
                payload=payload,
                result=result,
                error=str(result.get("error") or ""),
            )
            if not result.get("success"):
                failure_id = _enqueue_intake_failure(
                    action="appointment_calendar",
                    kind="appointment",
                    draft_id=str(request.draft_id or ""),
                    payload=payload,
                    error=str(result.get("error") or "unknown failure"),
                )
                result["failure_id"] = failure_id
            return result
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/imessages/intake/confirmation/send")
    async def imessages_intake_confirmation_send(request: IMessageConfirmationRequest):
        """Send a confirmation iMessage for a draft."""
        try:
            payload = {
                "kind": request.kind,
                "draft_id": request.draft_id,
                "message": request.message,
            }
            result = _execute_confirmation_send(payload)
            _append_intake_audit(
                action="confirmation_send",
                kind=str(request.kind or ""),
                draft_id=str(request.draft_id or ""),
                success=bool(result.get("success")),
                payload=payload,
                result=result,
                error=str(result.get("error") or ""),
            )
            if not result.get("success"):
                failure_id = _enqueue_intake_failure(
                    action="confirmation_send",
                    kind=str(request.kind or ""),
                    draft_id=str(request.draft_id or ""),
                    payload=payload,
                    error=str(result.get("error") or "unknown failure"),
                )
                result["failure_id"] = failure_id
            return result
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.get("/imessages/intake/audit")
    async def imessages_intake_audit(limit: int = 100):
        """Recent intake actions audit entries."""
        try:
            items = _read_recent_jsonl(IMESSAGE_INTAKE_AUDIT_FILE, limit=max(1, min(limit, 1000)))
            return {
                "success": True,
                "count": len(items),
                "items": items,
                "source_file": str(IMESSAGE_INTAKE_AUDIT_FILE),
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.get("/imessages/intake/failures")
    async def imessages_intake_failures(status: str = "pending", limit: int = 100):
        """List intake action failures queued for retry."""
        try:
            rows = _load_intake_failures()
            target = (status or "pending").strip().lower()
            if target != "all":
                rows = [r for r in rows if str(r.get("status", "")).lower() == target]
            rows = sorted(rows, key=lambda x: str(x.get("created_at") or ""), reverse=True)
            rows = rows[: max(1, min(limit, 1000))]
            return {"success": True, "count": len(rows), "status_filter": target, "items": rows}
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/imessages/intake/failures/{failure_id}/retry")
    async def imessages_intake_failures_retry(failure_id: str):
        """Retry a queued intake failure by id."""
        try:
            rows = _load_intake_failures()
            target = None
            for row in rows:
                if str(row.get("id", "")) == failure_id:
                    target = row
                    break
            if target is None:
                raise HTTPException(status_code=404, detail="Failure item not found")

            action = str(target.get("action", "")).strip().lower()
            payload = target.get("payload", {}) if isinstance(target.get("payload"), dict) else {}
            if action == "invoice_approve":
                result = _execute_invoice_approve(payload)
            elif action == "appointment_calendar":
                result = _execute_appointment_schedule(payload)
            elif action == "confirmation_send":
                result = _execute_confirmation_send(payload)
            else:
                result = {"success": False, "error": f"Unsupported action: {action}"}

            target["attempt_count"] = int(target.get("attempt_count", 0)) + 1
            target["last_attempt_at"] = datetime.now().isoformat()
            if result.get("success"):
                target["status"] = "resolved"
                target["resolved_at"] = datetime.now().isoformat()
                target["last_error"] = ""
            else:
                target["status"] = "pending"
                target["last_error"] = str(result.get("error") or "retry failed")
            _save_intake_failures(rows)

            _append_intake_audit(
                action=f"retry:{action}",
                kind=str(target.get("kind", "")),
                draft_id=str(target.get("draft_id", "")),
                success=bool(result.get("success")),
                payload=payload,
                result=result,
                error=str(result.get("error") or ""),
            )

            return {
                "success": bool(result.get("success")),
                "failure_id": failure_id,
                "failure_status": target.get("status"),
                "result": result,
            }
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/imessages/intake/failures/retry_all")
    async def imessages_intake_failures_retry_all(request: IMessageRetryAllFailuresRequest):
        """Retry multiple pending intake failures."""
        try:
            rows = _load_intake_failures()
            pending = [r for r in rows if str(r.get("status", "")).lower() == "pending"]
            pending = sorted(pending, key=lambda x: str(x.get("created_at") or ""))
            capped = pending[: max(1, min(int(request.limit or 25), 200))]
            retried = 0
            resolved = 0
            still_pending = 0
            results: List[Dict[str, Any]] = []

            for target in capped:
                action = str(target.get("action", "")).strip().lower()
                payload = target.get("payload", {}) if isinstance(target.get("payload"), dict) else {}
                if action == "invoice_approve":
                    result = _execute_invoice_approve(payload)
                elif action == "appointment_calendar":
                    result = _execute_appointment_schedule(payload)
                elif action == "confirmation_send":
                    result = _execute_confirmation_send(payload)
                else:
                    result = {"success": False, "error": f"Unsupported action: {action}"}

                target["attempt_count"] = int(target.get("attempt_count", 0)) + 1
                target["last_attempt_at"] = datetime.now().isoformat()
                retried += 1
                if result.get("success"):
                    target["status"] = "resolved"
                    target["resolved_at"] = datetime.now().isoformat()
                    target["last_error"] = ""
                    resolved += 1
                else:
                    target["status"] = "pending"
                    target["last_error"] = str(result.get("error") or "retry failed")
                    still_pending += 1

                _append_intake_audit(
                    action=f"retry_all:{action}",
                    kind=str(target.get("kind", "")),
                    draft_id=str(target.get("draft_id", "")),
                    success=bool(result.get("success")),
                    payload=payload,
                    result=result,
                    error=str(result.get("error") or ""),
                )
                results.append(
                    {
                        "failure_id": target.get("id"),
                        "action": action,
                        "success": bool(result.get("success")),
                        "error": str(result.get("error") or ""),
                    }
                )

            _save_intake_failures(rows)
            return {
                "success": True,
                "retried_count": retried,
                "resolved_count": resolved,
                "still_pending_count": still_pending,
                "results": results,
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.get("/contacts/status")
    async def contacts_status():
        """Get contacts sync status and launchd job status."""
        try:
            contacts_job = get_launchd_job_status("com.symphony.contacts-sync")
            status_result = run_tool_endpoint("contacts_sync.py", ["--status"], timeout=30)
            status_json = _extract_json_from_output(status_result, fallback_key="output")
            return {
                "success": True,
                "timestamp": datetime.now().isoformat(),
                "job": contacts_job,
                "contacts": status_json,
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/contacts/sync")
    async def contacts_sync_now():
        """Run contacts sync now to refresh phone/email/project linking index."""
        try:
            result = run_tool_endpoint("contacts_sync.py", ["--sync"], timeout=120)
            payload = _extract_json_from_output(result, fallback_key="output")
            payload.setdefault("command_success", result.get("success", False))
            return payload
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.get("/contacts/lookup")
    async def contacts_lookup(query: str):
        """Lookup contact by phone/email/name using current index."""
        try:
            result = run_tool_endpoint("contacts_sync.py", ["--lookup", query], timeout=30)
            payload = _extract_json_from_output(result, fallback_key="output")
            payload.setdefault("command_success", result.get("success", False))
            return payload
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.get("/contacts/list")
    async def contacts_list(query: str = "", limit: int = 200):
        """List synced contacts for quick in-app picker/search."""
        try:
            data = _read_json_file(CONTACTS_INDEX_FILE)
            contacts = data.get("contacts", []) if isinstance(data, dict) else []
            q = (query or "").strip().lower()
            if q:
                filtered = []
                for c in contacts:
                    name = str(c.get("name", "")).lower()
                    phones = " ".join(c.get("phones", []))
                    emails = " ".join(c.get("emails", []))
                    if q in name or q in phones or q in emails.lower():
                        filtered.append(c)
                contacts = filtered
            contacts = contacts[: max(1, min(limit, 1000))]
            items = []
            for c in contacts:
                name = str(c.get("name", "")).strip() or "Unknown"
                phones = [str(p) for p in c.get("phones", [])]
                emails = [str(e) for e in c.get("emails", [])]
                linked_projects = [str(p) for p in c.get("linked_projects", [])]
                stable = f"{name}|{'|'.join(phones[:2])}|{'|'.join(emails[:1])}"
                item_id = _safe_export_filename(stable)[:120]
                items.append(
                    {
                        "id": item_id,
                        "name": name,
                        "phones": phones,
                        "emails": emails,
                        "linked_projects": linked_projects,
                    }
                )
            return {"success": True, "count": len(items), "contacts": items, "timestamp": datetime.now().isoformat()}
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.get("/contacts/clients")
    async def contacts_clients():
        """List manually added client contacts."""
        try:
            data = _read_json_file(CLIENTS_REGISTRY_FILE)
            clients = data.get("clients", []) if isinstance(data, dict) else []
            return {"success": True, "count": len(clients), "clients": clients}
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/contacts/clients/add")
    async def contacts_clients_add(request: ClientContactAddRequest):
        """
        Add a new client contact record and optionally append phones to iMessage watchlist.
        """
        try:
            name = (request.name or "").strip()
            if not name:
                raise HTTPException(status_code=400, detail="Client name is required")

            CLIENTS_REGISTRY_FILE.parent.mkdir(parents=True, exist_ok=True)
            data = _read_json_file(CLIENTS_REGISTRY_FILE)
            clients = data.get("clients", []) if isinstance(data, dict) else []

            phones = [str(p).strip() for p in (request.phones or []) if str(p).strip()]
            emails = [str(e).strip().lower() for e in (request.emails or []) if str(e).strip()]
            client_id = f"client_{uuid4().hex[:10]}"
            record = {
                "id": client_id,
                "name": name,
                "phones": phones,
                "emails": emails,
                "notes": (request.notes or "").strip(),
                "created_at": datetime.now().isoformat(),
            }
            clients.append(record)
            CLIENTS_REGISTRY_FILE.write_text(json.dumps({"clients": clients}, indent=2), encoding="utf-8")

            monitor_result = None
            if request.auto_monitor and phones:
                state = _read_json_file(IMESSAGE_WATCHER_STATE_FILE)
                current = [str(x) for x in state.get("watchlist", [])] if isinstance(state, dict) else []
                merged = sorted(set(current + phones))
                set_result = run_tool_endpoint(
                    "imessage_watcher.py",
                    ["--set-watchlist", ",".join(merged)],
                    timeout=30,
                )
                monitor_result = _extract_json_from_output(set_result, fallback_key="output")
                monitor_result.setdefault("command_success", set_result.get("success", False))

            return {
                "success": True,
                "client": record,
                "clients_count": len(clients),
                "watchlist_updated": bool(request.auto_monitor and phones),
                "monitor_result": monitor_result,
            }
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/contacts/clients/remove")
    async def contacts_clients_remove(request: ClientContactRemoveRequest):
        """Remove a client contact record and optionally remove its handles from iMessage watchlist."""
        try:
            client_id = (request.client_id or "").strip()
            if not client_id:
                raise HTTPException(status_code=400, detail="client_id is required")

            data = _read_json_file(CLIENTS_REGISTRY_FILE)
            clients = data.get("clients", []) if isinstance(data, dict) else []
            target = None
            kept = []
            for item in clients:
                if str(item.get("id", "")) == client_id and target is None:
                    target = item
                    continue
                kept.append(item)
            if target is None:
                raise HTTPException(status_code=404, detail="Client not found")

            CLIENTS_REGISTRY_FILE.parent.mkdir(parents=True, exist_ok=True)
            CLIENTS_REGISTRY_FILE.write_text(json.dumps({"clients": kept}, indent=2), encoding="utf-8")

            monitor_result = None
            if bool(request.remove_from_watchlist):
                state = _read_json_file(IMESSAGE_WATCHER_STATE_FILE)
                current = [str(x) for x in state.get("watchlist", [])] if isinstance(state, dict) else []
                removal = {
                    normalize_contact_identifier(str(x))
                    for x in ([*(target.get("phones", []) or []), *(target.get("emails", []) or [])])
                    if str(x).strip()
                }
                merged = [x for x in current if normalize_contact_identifier(x) not in removal]
                set_result = run_tool_endpoint(
                    "imessage_watcher.py",
                    ["--set-watchlist", ",".join(merged)] if merged else ["--clear-watchlist"],
                    timeout=30,
                )
                monitor_result = _extract_json_from_output(set_result, fallback_key="output")
                monitor_result.setdefault("command_success", set_result.get("success", False))

            return {
                "success": True,
                "removed_client_id": client_id,
                "removed_client_name": target.get("name"),
                "clients_count": len(kept),
                "watchlist_updated": bool(request.remove_from_watchlist),
                "monitor_result": monitor_result,
            }
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/contacts/watchlist/seed")
    async def contacts_seed_watchlist(request: ContactsSeedWatchlistRequest):
        """Seed iMessage scan list from saved client/contact records."""
        try:
            candidates: List[str] = []

            if bool(request.include_clients_registry):
                data = _read_json_file(CLIENTS_REGISTRY_FILE)
                clients = data.get("clients", []) if isinstance(data, dict) else []
                for client in clients:
                    phones = [str(x).strip() for x in (client.get("phones", []) or []) if str(x).strip()]
                    emails = [str(x).strip().lower() for x in (client.get("emails", []) or []) if str(x).strip()]
                    candidates.extend(phones)
                    if bool(request.include_emails):
                        candidates.extend(emails)

            if bool(request.include_contacts_index):
                idx = _read_json_file(CONTACTS_INDEX_FILE)
                contacts = idx.get("contacts", []) if isinstance(idx, dict) else []
                for contact in contacts:
                    phones = [str(x).strip() for x in (contact.get("phones", []) or []) if str(x).strip()]
                    emails = [str(x).strip().lower() for x in (contact.get("emails", []) or []) if str(x).strip()]
                    candidates.extend(phones)
                    if bool(request.include_emails):
                        candidates.extend(emails)

            unique: List[str] = []
            seen: set[str] = set()
            for raw in candidates:
                norm = normalize_contact_identifier(raw)
                if not norm or norm in seen:
                    continue
                seen.add(norm)
                unique.append(raw)

            state = _read_json_file(IMESSAGE_WATCHER_STATE_FILE)
            existing = [str(x).strip() for x in (state.get("watchlist", []) if isinstance(state, dict) else []) if str(x).strip()]
            existing_norm = {normalize_contact_identifier(x) for x in existing}

            if bool(request.overwrite):
                final = unique
            else:
                final = list(existing)
                for raw in unique:
                    norm = normalize_contact_identifier(raw)
                    if norm in existing_norm:
                        continue
                    final.append(raw)
                    existing_norm.add(norm)

            args = ["--set-watchlist", ",".join(final)] if final else ["--clear-watchlist"]
            set_result = run_tool_endpoint("imessage_watcher.py", args, timeout=45)
            payload = _extract_json_from_output(set_result, fallback_key="output")
            payload.setdefault("command_success", set_result.get("success", False))
            payload["seeded_count"] = len(unique)
            payload["final_watchlist_count"] = len(final)
            payload["mode"] = "watchlist_plus_keywords"
            return payload
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.get("/ops/health")
    async def ops_health():
        """Unified operations health across APIs, pipelines, automations, and iOS guardian."""
        try:
            services = {
                "mobile_api": {
                    "healthy": True,
                    "url": "http://127.0.0.1:8420/health",
                }
            }
            # External local dependencies.
            try:
                trading_health = _http_get_json("http://127.0.0.1:8421/health", timeout=4)
                services["trading_api"] = {
                    "healthy": trading_health.get("status") == "healthy",
                    "url": "http://127.0.0.1:8421/health",
                }
            except Exception as exc:
                services["trading_api"] = {
                    "healthy": False,
                    "url": "http://127.0.0.1:8421/health",
                    "error": str(exc),
                }

            # Notes pipeline and launchd jobs.
            jobs = {
                "memory_guard": get_launchd_job_status("com.symphony.memory-guard"),
                "notes_watcher": get_launchd_job_status("com.symphony.notes-watcher"),
                "incoming_tasks": get_launchd_job_status("com.symphony.incoming-tasks"),
                "notes_sync_photos": get_launchd_job_status("com.symphony.notes-sync-photos"),
                "ios_build_guardian": get_launchd_job_status("com.symphony.ios-build-guardian"),
                "autonomous_recovery": get_launchd_job_status("com.symphony.autonomous-recovery"),
            }

            notes_state = _read_json_file(BASE_DIR / "knowledge" / "state" / "notes_watcher_state.json")
            incoming_state = _read_json_file(BASE_DIR / "knowledge" / "state" / "incoming_tasks_state.json")
            ios_guardian = _read_json_file(BASE_DIR / "data" / "ios_build_guardian_status.json")
            recovery_state = _read_json_file(BASE_DIR / "data" / "autonomous_recovery_last.json")

            # Trading automation snapshot (if trading API available).
            trading_automation = {}
            if services.get("trading_api", {}).get("healthy"):
                try:
                    trading_automation = _http_get_json("http://127.0.0.1:8421/automation/health", timeout=5)
                except Exception as exc:
                    trading_automation = {"success": False, "error": str(exc)}

            problems: List[str] = []
            if not services["trading_api"]["healthy"]:
                problems.append("trading_api_unhealthy")
            if ios_guardian and not ios_guardian.get("overall_ok", True):
                problems.append("ios_build_failures")
            for key, job in jobs.items():
                if key in {"autonomous_recovery"}:
                    continue
                if job.get("loaded") and not job.get("running") and job.get("last_exit_code", 0) not in (0, None):
                    problems.append(f"{key}_not_running")

            status = "healthy" if not problems else "degraded"
            return {
                "success": True,
                "status": status,
                "timestamp": datetime.now().isoformat(),
                "services": services,
                "jobs": jobs,
                "notes_pipeline": {
                    "notes_watcher_last_check": notes_state.get("last_check"),
                    "notes_processed_count": len(notes_state.get("processed_notes", {})),
                    "incoming_last_check": incoming_state.get("last_check"),
                    "incoming_processed_count": len(incoming_state.get("processed_notes", {})),
                },
                "ios_build_guardian": ios_guardian,
                "autonomous_recovery": recovery_state,
                "trading_automation": trading_automation,
                "problems": problems,
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/ops/recovery/run")
    async def ops_recovery_run(request: OpsRecoveryRunRequest):
        """Run autonomous recovery scan/apply with confidence gating."""
        try:
            threshold = max(0.0, min(1.0, float(request.threshold)))
            args = []
            if request.apply:
                args.append("--apply")
            else:
                args.append("--scan")
            args.extend(["--threshold", str(threshold)])
            if request.playbook:
                args.extend(["--playbook", request.playbook])

            result = run_tool_endpoint("autonomous_recovery.py", args, timeout=180)
            payload = _extract_json_from_output(result, fallback_key="output")
            payload.setdefault("command_success", result.get("success", False))
            return payload
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.get("/ops/inventory/summary")
    async def ops_inventory_summary(low_stock_limit: int = 25, top_limit: int = 100):
        """Inventory intelligence summary for Ops app (low stock, reorder candidates, manuals queue)."""
        try:
            summary = _build_inventory_summary(low_stock_limit=low_stock_limit, top_limit=top_limit)
            # If reports are missing, generate once automatically then return summary.
            if summary["counts"]["inventory_rows"] == 0:
                run_tool_endpoint("bob_build_inventory.py", [], timeout=240)
                summary = _build_inventory_summary(low_stock_limit=low_stock_limit, top_limit=top_limit)
            return summary
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/ops/inventory/rebuild")
    async def ops_inventory_rebuild(request: InventoryRebuildRequest):
        """Rebuild inventory reports from extracted knowledge and return fresh summary."""
        try:
            result = run_tool_endpoint("bob_build_inventory.py", [], timeout=300)
            payload = _extract_json_from_output(result, fallback_key="output")
            summary = _build_inventory_summary(
                low_stock_limit=max(1, min(int(request.low_stock_limit or 25), 200)),
                top_limit=max(1, min(int(request.top_limit or 100), 500)),
            )
            return {
                "success": bool(result.get("success", False)),
                "command_success": bool(result.get("success", False)),
                "command_output": str(result.get("output", ""))[-2000:],
                "command_error": result.get("error"),
                "command_result": payload,
                "summary": summary,
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.get("/ops/turnkey/status")
    async def ops_turnkey_status():
        """Turnkey readiness snapshot for app-driven operations."""
        try:
            env_checks = {
                "PERPLEXITY_API_KEY": _has_env_key("PERPLEXITY_API_KEY"),
                "DTOOLS_API_KEY": _has_env_key("DTOOLS_API_KEY"),
                "OPENAI_API_KEY": _has_env_key("OPENAI_API_KEY"),
                "SYMPHONY_API_TOKEN": _has_env_key("SYMPHONY_API_TOKEN"),
                "TELEGRAM_BOT_TOKEN": _has_env_key("TELEGRAM_BOT_TOKEN"),
            }
            jobs = {
                "mobile_api": get_launchd_job_status("com.symphony.mobile-api"),
                "markup_app": get_launchd_job_status("com.symphony.markup-app"),
                "incoming_tasks": get_launchd_job_status("com.symphony.incoming-tasks"),
                "notes_watcher": get_launchd_job_status("com.symphony.notes-watcher"),
            }
            all_env_ok = all(env_checks.values())
            job_issues = [
                key for key, job in jobs.items()
                if job.get("loaded") and not job.get("running") and job.get("last_exit_code", 0) not in (0, None)
            ]
            ready = all_env_ok and not job_issues
            return {
                "success": True,
                "ready": ready,
                "timestamp": datetime.now().isoformat(),
                "env": env_checks,
                "jobs": jobs,
                "missing_env": [k for k, ok in env_checks.items() if not ok],
                "job_issues": job_issues,
                "next_steps": (
                    [] if ready else [
                        "Add missing API keys in .env",
                        "Restart failed launchd services",
                        "Run turnkey brief from Ops app and follow plan",
                    ]
                ),
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/ops/integration/brief")
    async def ops_integration_brief(request: IntegrationBriefRequest):
        """Generate unified Ops integration brief via Perplexity/local fallback."""
        try:
            args = [str(BASE_DIR / "tools" / "perplexity_integration_brief.py"), "--out", "docs/OPS_UNIFIED_BRIEF.md"]
            if request.no_cache:
                args.append("--no-cache")
            result = run_command(["python3", *args], timeout=300)
            out_path = BASE_DIR / "docs" / "OPS_UNIFIED_BRIEF.md"
            preview = ""
            if out_path.exists():
                try:
                    preview = out_path.read_text(encoding="utf-8", errors="ignore")[:4000]
                except Exception:
                    preview = ""
            return {
                "success": bool(result.get("success", False)),
                "command_success": bool(result.get("success", False)),
                "command_output": str(result.get("output", ""))[-2000:],
                "command_error": result.get("error"),
                "brief_path": str(out_path),
                "preview": preview,
                "timestamp": datetime.now().isoformat(),
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.get("/ops/employee_bots")
    async def ops_employee_bots():
        """List employee bot profiles and launch hints."""
        try:
            payload = _load_employee_bots()
            workers = payload.get("workers", {}) if isinstance(payload, dict) else {}
            items: List[Dict[str, Any]] = []
            for worker_id, cfg in workers.items():
                if not isinstance(cfg, dict):
                    continue
                token_env = str(cfg.get("token_env") or f"{worker_id.upper()}_BOT_TOKEN")
                items.append({
                    "worker_id": worker_id,
                    "name": str(cfg.get("name") or worker_id.title()),
                    "role": str(cfg.get("role") or "Employee Bot"),
                    "emoji": str(cfg.get("emoji") or "🤖"),
                    "skills": [str(s) for s in cfg.get("skills", [])],
                    "intro": str(cfg.get("intro") or ""),
                    "token_env": token_env,
                    "token_configured": _has_env_key(token_env),
                    "start_command": f"EMPLOYEE={worker_id} TELEGRAM_BOT_TOKEN=${token_env} python3 telegram-bob-remote/employee_bot.py",
                    **_employee_bot_runtime_status(worker_id),
                })
            items.sort(key=lambda x: x["worker_id"])
            return {
                "success": True,
                "count": len(items),
                "workers": items,
                "path": str(EMPLOYEE_BOTS_FILE),
                "valid_task_types": _valid_task_types(),
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/ops/employee_bots/create")
    async def ops_employee_bots_create(request: EmployeeBotCreateRequest):
        """Create or update an employee bot profile."""
        try:
            worker_id = _slugify(request.worker_id, fallback="")
            if not worker_id:
                raise HTTPException(status_code=400, detail="worker_id is required")
            valid_types = set(_valid_task_types())
            skills = [str(s).strip().lower() for s in request.skills if str(s).strip()]
            bad = [s for s in skills if s not in valid_types]
            if bad:
                raise HTTPException(status_code=400, detail=f"Invalid skills: {', '.join(bad)}")
            if not skills:
                raise HTTPException(status_code=400, detail="At least one skill is required")
            payload = _load_employee_bots()
            workers = payload.setdefault("workers", {})
            token_env = (request.token_env or "").strip() or f"{worker_id.upper()}_BOT_TOKEN"
            workers[worker_id] = {
                "name": request.name.strip() or worker_id.title(),
                "role": request.role.strip() or "Employee Bot",
                "emoji": (request.emoji or "🤖").strip() or "🤖",
                "skills": skills,
                "intro": (request.intro or "").strip(),
                "token_env": token_env,
            }
            _save_employee_bots(payload)
            profile = {"worker_id": worker_id, **workers[worker_id]}
            return {
                "success": True,
                "worker_id": worker_id,
                "profile": profile,
                "token_configured": _has_env_key(token_env),
                "start_command": f"EMPLOYEE={worker_id} TELEGRAM_BOT_TOKEN=${token_env} python3 telegram-bob-remote/employee_bot.py",
                "notes": "Set token env in .env and restart worker/bot services to apply.",
            }
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.get("/ops/employee_bots/runtime")
    async def ops_employee_bots_runtime(worker_id: str):
        """Get runtime status for a specific employee bot."""
        try:
            worker = _slugify(worker_id, fallback="")
            if not worker:
                raise HTTPException(status_code=400, detail="worker_id is required")
            return {"success": True, **_employee_bot_runtime_status(worker)}
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/ops/employee_bots/start")
    async def ops_employee_bots_start(request: EmployeeBotWorkerRequest):
        """Start an employee Telegram bot in background."""
        try:
            worker = _slugify(request.worker_id, fallback="")
            payload = _load_employee_bots()
            workers = payload.get("workers", {})
            if worker not in workers:
                raise HTTPException(status_code=404, detail=f"Unknown worker_id: {worker}")
            status = _employee_bot_runtime_status(worker)
            if status.get("running"):
                return {"success": True, "message": "Already running", **status}

            cfg = workers.get(worker, {})
            token_env = str(cfg.get("token_env") or f"{worker.upper()}_BOT_TOKEN")
            token = (os.environ.get(token_env, "") or "").strip()
            if not token:
                raise HTTPException(status_code=400, detail=f"Missing token env: {token_env}")

            py = _pick_python_with_module("telegram")
            if not py:
                raise HTTPException(status_code=500, detail="No python interpreter with telegram module found")

            paths = _employee_bot_paths(worker)
            paths["log"].parent.mkdir(parents=True, exist_ok=True)
            script = BASE_DIR / "telegram-bob-remote" / "employee_bot.py"
            env = os.environ.copy()
            env["EMPLOYEE"] = worker
            env["TELEGRAM_BOT_TOKEN"] = token
            with paths["log"].open("a", encoding="utf-8") as out, paths["err"].open("a", encoding="utf-8") as err:
                proc = subprocess.Popen(
                    [py, str(script)],
                    cwd=BASE_DIR,
                    stdout=out,
                    stderr=err,
                    env=env,
                    start_new_session=True,
                )
            paths["pid"].write_text(str(proc.pid), encoding="utf-8")
            return {"success": True, "message": "Started", **_employee_bot_runtime_status(worker)}
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/ops/employee_bots/stop")
    async def ops_employee_bots_stop(request: EmployeeBotWorkerRequest):
        """Stop an employee Telegram bot."""
        try:
            worker = _slugify(request.worker_id, fallback="")
            if not worker:
                raise HTTPException(status_code=400, detail="worker_id is required")
            paths = _employee_bot_paths(worker)
            pid = _employee_bot_pid(worker)
            if not _pid_is_alive(pid):
                try:
                    paths["pid"].unlink(missing_ok=True)
                except Exception:
                    pass
                return {"success": True, "message": "Already stopped", **_employee_bot_runtime_status(worker)}
            os.kill(pid, signal.SIGTERM)
            for _ in range(12):
                if not _pid_is_alive(pid):
                    break
                time.sleep(0.1)
            if _pid_is_alive(pid):
                os.kill(pid, signal.SIGKILL)
            try:
                paths["pid"].unlink(missing_ok=True)
            except Exception:
                pass
            return {"success": True, "message": "Stopped", **_employee_bot_runtime_status(worker)}
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/ops/employee_bots/restart")
    async def ops_employee_bots_restart(request: EmployeeBotWorkerRequest):
        """Restart an employee bot."""
        await ops_employee_bots_stop(request)
        return await ops_employee_bots_start(request)

    @app.post("/ops/employee_bots/test")
    async def ops_employee_bots_test(request: EmployeeBotWorkerRequest):
        """Validate employee bot token by calling Telegram getMe."""
        try:
            worker = _slugify(request.worker_id, fallback="")
            payload = _load_employee_bots()
            workers = payload.get("workers", {})
            if worker not in workers:
                raise HTTPException(status_code=404, detail=f"Unknown worker_id: {worker}")
            cfg = workers.get(worker, {})
            token_env = str(cfg.get("token_env") or f"{worker.upper()}_BOT_TOKEN")
            token = (os.environ.get(token_env, "") or "").strip()
            if not token:
                raise HTTPException(status_code=400, detail=f"Missing token env: {token_env}")
            url = f"https://api.telegram.org/bot{token}/getMe"
            result = _http_get_json(url, timeout=8)
            ok = bool(result.get("ok"))
            return {
                "success": ok,
                "worker_id": worker,
                "token_env": token_env,
                "telegram_ok": ok,
                "bot": result.get("result") if ok else None,
                "raw": result if not ok else None,
            }
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.get("/network/dropout/status")
    async def network_dropout_status():
        """Get live state of the network dropout watcher."""
        try:
            pid = _network_watch_pid()
            running = _pid_is_alive(pid)
            status = _read_json_file(NETWORK_WATCH_STATUS_FILE)
            recent_events = _read_recent_jsonl(NETWORK_WATCH_EVENTS_FILE, limit=30)
            return {
                "success": True,
                "running": running,
                "pid": pid,
                "status_file": str(NETWORK_WATCH_STATUS_FILE),
                "events_file": str(NETWORK_WATCH_EVENTS_FILE),
                "status": status,
                "recent_events": recent_events,
                "timestamp": datetime.now().isoformat(),
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/network/dropout/start")
    async def network_dropout_start(request: NetworkDropoutWatchStartRequest):
        """Start network dropout watcher process."""
        try:
            existing_pid = _network_watch_pid()
            if _pid_is_alive(existing_pid):
                return {
                    "success": True,
                    "running": True,
                    "pid": existing_pid,
                    "message": "Watcher already running",
                }

            python_bin = BASE_DIR / ".venv" / "bin" / "python3"
            args = [
                str(python_bin if python_bin.exists() else Path(sys.executable)),
                str(BASE_DIR / "tools" / "network_dropout_watch.py"),
                "--watch",
                "--gateway-ip",
                request.gateway_ip.strip(),
                "--wan-ip",
                request.wan_ip.strip(),
                "--interval-sec",
                str(max(0.5, min(float(request.interval_sec), 10.0))),
                "--state-dir",
                str(NETWORK_WATCH_DIR),
            ]
            if request.control4_ip and request.control4_ip.strip():
                args.extend(["--control4-ip", request.control4_ip.strip()])
            if request.sonos_ip and request.sonos_ip.strip():
                args.extend(["--sonos-ip", request.sonos_ip.strip()])

            with NETWORK_WATCH_STDOUT.open("a", encoding="utf-8") as out, NETWORK_WATCH_STDERR.open("a", encoding="utf-8") as err:
                proc = subprocess.Popen(
                    args,
                    cwd=BASE_DIR,
                    stdout=out,
                    stderr=err,
                    start_new_session=True,
                )
            NETWORK_WATCH_PID_FILE.write_text(str(proc.pid), encoding="utf-8")

            return {
                "success": True,
                "running": True,
                "pid": proc.pid,
                "message": "Watcher started",
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/network/dropout/stop")
    async def network_dropout_stop():
        """Stop network dropout watcher process if running."""
        try:
            pid = _network_watch_pid()
            if not _pid_is_alive(pid):
                try:
                    NETWORK_WATCH_PID_FILE.unlink(missing_ok=True)
                except Exception:
                    pass
                return {"success": True, "running": False, "message": "Watcher already stopped"}

            os.kill(pid, signal.SIGTERM)
            # Give it a brief moment to exit cleanly.
            for _ in range(10):
                if not _pid_is_alive(pid):
                    break
                time.sleep(0.1)
            if _pid_is_alive(pid):
                os.kill(pid, signal.SIGKILL)
            try:
                NETWORK_WATCH_PID_FILE.unlink(missing_ok=True)
            except Exception:
                pass
            return {"success": True, "running": False, "message": "Watcher stopped"}
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.get("/network/dropout/events")
    async def network_dropout_events(limit: int = 100):
        """Tail watcher events for recent dropout transitions."""
        try:
            events = _read_recent_jsonl(NETWORK_WATCH_EVENTS_FILE, limit=max(1, min(limit, 500)))
            return {
                "success": True,
                "count": len(events),
                "events": events,
                "timestamp": datetime.now().isoformat(),
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/tasks/incidents/from_dropout")
    async def create_incident_from_dropout(request: DropoutIncidentCreateRequest):
        """Create a troubleshooting task from the most recent dropout event."""
        try:
            sys.path.insert(0, str(BASE_DIR))
            from orchestrator.task_board import add_task

            events = _read_recent_jsonl(NETWORK_WATCH_EVENTS_FILE, limit=200)
            selected_event: Optional[Dict[str, Any]] = None
            requested_ts = (request.event_timestamp or "").strip()
            requested_name = (request.event_name or "").strip()

            if requested_ts:
                for event in events:
                    if str(event.get("timestamp", "")).strip() != requested_ts:
                        continue
                    if requested_name and str(event.get("event", "")).strip() != requested_name:
                        continue
                    selected_event = event
                    break
            elif events:
                selected_event = events[0]

            if not selected_event:
                return {
                    "success": False,
                    "error": "No dropout event available to create incident.",
                    "task_id": None,
                }

            event_name = str(selected_event.get("event") or "dropout").strip() or "dropout"
            from_state = str(selected_event.get("from") or "unknown").strip() or "unknown"
            to_state = str(selected_event.get("to") or selected_event.get("health") or "unknown").strip() or "unknown"
            timestamp = str(selected_event.get("timestamp") or datetime.now().isoformat()).strip()
            notes = (request.notes or "").strip()
            priority = (request.priority or "high").strip().lower()
            if priority not in {"critical", "high", "medium", "low"}:
                priority = "high"

            title = f"Network incident: {event_name} ({from_state} -> {to_state})"
            description_lines = [
                "Generated from network dropout watcher.",
                f"Event: {event_name}",
                f"Transition: {from_state} -> {to_state}",
                f"Timestamp: {timestamp}",
            ]
            if notes:
                description_lines.append(f"Operator notes: {notes}")

            task_id = add_task(
                title=title[:220],
                description="\n".join(description_lines),
                task_type="troubleshooting",
                priority=priority,
                source="network_dropout_watch",
                source_id=f"dropout:{timestamp}:{event_name}",
                metadata={"dropout_event": selected_event},
            )

            return {
                "success": True,
                "task_id": task_id,
                "title": title,
                "priority": priority,
                "event_timestamp": timestamp,
                "message": "Incident task created from dropout event.",
            }
        except Exception as e:
            return {"success": False, "task_id": None, "error": str(e)}

    @app.post("/trading/fix_api")
    async def trading_fix_api():
        """Run watchdog to clear stale :8421 listener and restart trading API."""
        try:
            status = run_trading_api_watchdog()
            return {
                "success": status.get("success", False),
                "output": json.dumps(status, indent=2),
                "error": None if status.get("success") else "Trading API watchdog failed",
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))
    
    @app.post("/research")
    async def research(request: ResearchRequest):
        """Search knowledge base."""
        result = run_command([
            "python3", str(BASE_DIR / "tools" / "smart_research.py"),
            "--query", request.query
        ], timeout=60)
        return result
    
    # --- Website ---
    
    @app.get("/website/status")
    async def website_status():
        """Check website health."""
        result = run_command([
            "python3", str(BASE_DIR / "tools" / "website_monitor.py"),
            "--json"
        ], timeout=60)
        return _extract_json_from_output(result, "website_status")
    
    # --- Subscriptions ---
    
    @app.get("/subscriptions")
    async def get_subscriptions():
        """Get all subscriptions."""
        subs_file = BASE_DIR / "knowledge" / "subscriptions.json"
        if subs_file.exists():
            return json.loads(subs_file.read_text())
        return {"subscriptions": []}
    
    # --- Morning Checklist ---
    
    @app.get("/morning")
    async def morning_checklist():
        """Run morning checklist."""
        result = run_command([
            "python3", str(BASE_DIR / "orchestrator" / "morning_checklist.py"),
            "--quick", "--dry"
        ], timeout=60)
        return result
    
    # --- Dealer Forms ---
    
    @app.get("/dealers")
    async def list_dealers():
        """List known dealer application forms."""
        result = run_command([
            "python3", str(BASE_DIR / "tools" / "dealer_forms.py"),
            "--list"
        ])
        return result
    
    # --- Usage Monitor ---
    
    @app.get("/usage")
    async def get_usage():
        """Get usage across all metered services."""
        result = run_command([
            "python3", str(BASE_DIR / "tools" / "usage_monitor.py"),
            "--json"
        ])
        if result.get("success") and result.get("output"):
            try:
                return json.loads(result["output"])
            except:
                pass
        return result
    
    @app.get("/usage/alerts")
    async def get_usage_alerts():
        """Get usage alerts only."""
        result = run_command([
            "python3", str(BASE_DIR / "tools" / "usage_monitor.py"),
            "--json", "--alerts"
        ])
        if result.get("success") and result.get("output"):
            try:
                return json.loads(result["output"])
            except:
                pass
        return result
    
    class UsageUpdate(BaseModel):
        service: str
        auto_pct: Optional[float] = None
        api_pct: Optional[float] = None
        spent: Optional[float] = None
    
    @app.post("/usage/update")
    async def update_usage(request: UsageUpdate):
        """Update usage for a service manually."""
        if request.service == "cursor":
            cmd = [
                "python3", str(BASE_DIR / "tools" / "usage_monitor.py"),
                "--update-cursor", str(request.auto_pct or 0), str(request.api_pct or 0)
            ]
        elif request.service == "openai":
            cmd = [
                "python3", str(BASE_DIR / "tools" / "usage_monitor.py"),
                "--update-openai", str(request.spent or 0)
            ]
        else:
            return {"success": False, "error": "Unknown service"}
        return run_command(cmd)
    
    # --- AI Markup Tool ---
    
    class MarkupRequest(BaseModel):
        project_name: str
        description: str
        rooms: Optional[List[str]] = None

    class ProductApprovalItem(BaseModel):
        brand: Optional[str] = None
        model: str
        part_number: Optional[str] = None
        category: Optional[str] = None
        short_description: Optional[str] = None
        keywords: Optional[str] = None
        unit_price: Optional[float] = None
        unit_cost: Optional[float] = None
        msrp: Optional[float] = None
        supplier: Optional[str] = None

    class ProductApproveStoreRequest(BaseModel):
        products: List[ProductApprovalItem]
        source_file: Optional[str] = ""
        parse_profile: Optional[str] = "auto"
        dealer_tier: Optional[str] = "standard"
        notes: Optional[str] = ""

    @app.post("/projects/manual_digest")
    async def projects_manual_digest(
        project_name: str = Form(""),
        run_ai_summary: bool = Form(True),
        files: List[UploadFile] = File(...),
    ):
        """
        Upload project manuals/docs and build a structured digest for project planning.
        """
        if not files:
            return {"success": False, "error": "No files uploaded."}

        ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        project_slug = _slugify(project_name or "new-project")
        batch_dir = MANUAL_DIGEST_DIR / project_slug / ts
        batch_dir.mkdir(parents=True, exist_ok=True)

        file_summaries: List[Dict[str, Any]] = []
        per_file_signals: List[Dict[str, Any]] = []

        for upload in files:
            safe_name = "".join(ch if ch.isalnum() or ch in "._-" else "_" for ch in (upload.filename or "file"))
            save_path = batch_dir / safe_name
            content = await upload.read()
            save_path.write_bytes(content)

            extracted_text = _extract_text_from_path(save_path)
            signals = _extract_manual_digest_signals(extracted_text) if extracted_text else {}
            if signals:
                per_file_signals.append(signals)

            file_summaries.append(
                {
                    "filename": safe_name,
                    "path": str(save_path),
                    "size_bytes": len(content),
                    "text_chars": len(extracted_text),
                    "supported": bool(extracted_text),
                }
            )

        merged = _merge_digest_signals(per_file_signals)
        ai_summary = _ai_project_digest_summary(project_name or project_slug, merged) if run_ai_summary else {}

        digest = {
            "success": True,
            "project_name": project_name or project_slug,
            "project_slug": project_slug,
            "batch_timestamp": ts,
            "files": file_summaries,
            "digest": merged,
            "ai_summary": ai_summary,
            "output_dir": str(batch_dir),
            "timestamp": datetime.now().isoformat(),
        }

        (batch_dir / "digest.json").write_text(json.dumps(digest, indent=2))
        return digest

    @app.post("/projects/room_modeler")
    async def projects_room_modeler(
        project_name: str = Form(""),
        system_profile: str = Form("control4"),
        markup_file: UploadFile = File(...),
    ):
        """
        Parse a .symphony markup file and generate per-room system modeling guidance.
        """
        try:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
            project_slug = _slugify(project_name or "room-modeler")
            batch_dir = MANUAL_DIGEST_DIR / "room_modeler" / project_slug / ts
            batch_dir.mkdir(parents=True, exist_ok=True)

            upload_name = markup_file.filename or "markup.symphony"
            safe_name = "".join(ch if ch.isalnum() or ch in "._-" else "_" for ch in upload_name)
            save_path = batch_dir / safe_name

            content = await markup_file.read()
            if not content:
                return {"success": False, "error": "Uploaded markup file was empty."}
            save_path.write_bytes(content)

            try:
                payload = json.loads(content.decode("utf-8", errors="ignore"))
            except Exception:
                return {
                    "success": False,
                    "error": "Could not parse .symphony JSON payload.",
                    "file_path": str(save_path),
                }

            extracted = _extract_rooms_from_symphony_markup(payload if isinstance(payload, dict) else {})
            rooms = extracted.get("rooms", [])
            if not rooms:
                return {
                    "success": False,
                    "error": "No rooms detected in markup file.",
                    "file_path": str(save_path),
                }

            profile = _room_model_profile(system_profile)
            room_models: List[Dict[str, Any]] = []
            for row in rooms:
                room_name = str(row.get("room", "")).strip()
                symbol_count = int(row.get("symbol_count", 0) or 0)
                room_models.append(
                    {
                        "room": room_name,
                        "pages": row.get("pages", []),
                        "detected_symbols": symbol_count,
                        "recommended_scope": profile.get("scope", []),
                        "checks": profile.get("checks", []),
                        "priority": "high" if symbol_count >= 6 else "medium",
                    }
                )

            result = {
                "success": True,
                "project_name": project_name or payload.get("projectName", project_slug),
                "system_profile": (system_profile or "control4").strip().lower(),
                "project_slug": project_slug,
                "batch_timestamp": ts,
                "file_name": safe_name,
                "file_path": str(save_path),
                "rooms_count": len(room_models),
                "rooms": room_models,
                "summary": {
                    "high_priority_rooms": sum(1 for r in room_models if r.get("priority") == "high"),
                    "total_detected_symbols": sum(int(r.get("detected_symbols", 0) or 0) for r in room_models),
                },
                "timestamp": datetime.now().isoformat(),
            }
            (batch_dir / "room_modeler.json").write_text(json.dumps(result, indent=2), encoding="utf-8")
            return result
        except Exception as e:
            return {"success": False, "error": str(e)}

    @app.post("/projects/proposal_scope")
    async def projects_proposal_scope(
        project_name: str = Form(""),
        client_name: str = Form(""),
        run_ai_summary: bool = Form(True),
        proposal_file: UploadFile = File(...),
    ):
        """
        Upload a finished proposal and return a structured scope-of-work summary.
        """
        try:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
            project_slug = _slugify(project_name or client_name or "proposal-scope")
            batch_dir = MANUAL_DIGEST_DIR / "proposal_scope" / project_slug / ts
            batch_dir.mkdir(parents=True, exist_ok=True)

            safe_name = "".join(ch if ch.isalnum() or ch in "._-" else "_" for ch in (proposal_file.filename or "proposal.pdf"))
            save_path = batch_dir / safe_name
            content = await proposal_file.read()
            if not content:
                return {"success": False, "error": "Uploaded proposal file was empty."}
            save_path.write_bytes(content)

            extracted_text = _extract_text_from_path(save_path)
            text_chars = len(extracted_text or "")
            if text_chars == 0:
                return {
                    "success": False,
                    "error": "Could not extract proposal text. Upload a text-based PDF or supported document.",
                    "file_path": str(save_path),
                }

            scope = _extract_proposal_scope_signals(extracted_text)
            qv = _extract_dtools_qv(extracted_text, fallback_text=safe_name)

            ai_summary: Dict[str, Any] = {}
            if run_ai_summary:
                digest_payload = {
                    "recommended_notes": scope.get("included_items", []),
                    "scope_notes": scope.get("scope_of_work", []),
                    "risk_notes": scope.get("allowances", []) + scope.get("assumptions", []),
                    "open_questions": scope.get("open_questions", []),
                    "detected_skus": scope.get("detected_skus", []),
                    "detected_brands": scope.get("detected_brands", []),
                }
                ai_summary = _ai_project_digest_summary(project_name or project_slug, digest_payload)

            result = {
                "success": True,
                "project_name": project_name or project_slug,
                "client_name": client_name,
                "project_slug": project_slug,
                "batch_timestamp": ts,
                "file_name": safe_name,
                "file_path": str(save_path),
                "text_chars": text_chars,
                "dtools_quote_id": qv.get("quote_id", ""),
                "dtools_version": qv.get("version", ""),
                "dtools_quote_version": qv.get("quote_version", ""),
                "scope": scope,
                "ai_summary": ai_summary,
                "timestamp": datetime.now().isoformat(),
            }
            (batch_dir / "proposal_scope.json").write_text(json.dumps(result, indent=2), encoding="utf-8")
            return result
        except Exception as e:
            return {"success": False, "error": str(e)}

    @app.post("/dtools/products/import")
    async def dtools_products_import(
        file: UploadFile = File(...),
        create_in_dtools: bool = Form(False),
        max_products: int = Form(25),
        dealer_tier: str = Form("standard"),
        parse_profile: str = Form("auto"),
        expected_columns_json: str = Form(""),
        dry_run: bool = Form(True),
    ):
        """
        Upload a pricing/data sheet, parse products, and optionally create them in D-Tools Cloud.
        """
        try:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            upload_name = file.filename or f"upload_{ts}.pdf"
            safe_name = "".join(ch if ch.isalnum() or ch in "._-" else "_" for ch in upload_name)
            upload_path = DTOOLS_PRODUCT_AGENT_DIR / f"{ts}_{safe_name}"

            content = await file.read()
            upload_path.write_bytes(content)

            custom_columns: Optional[List[str]] = None
            if expected_columns_json:
                try:
                    maybe = json.loads(expected_columns_json)
                    if isinstance(maybe, list):
                        custom_columns = [str(x).strip() for x in maybe if str(x).strip()]
                except Exception:
                    custom_columns = None

            products = _parse_sheet_file_with_profile(
                upload_path,
                dealer_tier=dealer_tier,
                parse_profile=parse_profile,
                custom_columns=custom_columns,
            )
            if not products:
                return {
                    "success": False,
                    "error": "No products parsed from file.",
                    "file": str(upload_path),
                }

            limited = products[: max(1, min(max_products, 500))]
            results: List[Dict[str, Any]] = []
            created_count = 0

            if create_in_dtools and not dry_run:
                from agents.dtools_browser_agent import DToolsBrowserAgent

                agent = DToolsBrowserAgent(headless=True)
                if not await agent.start():
                    return {
                        "success": False,
                        "error": "Failed to start browser agent (Playwright not ready).",
                        "parsed_count": len(products),
                        "products": limited,
                    }
                try:
                    if not await asyncio.wait_for(agent.login(), timeout=45):
                        return {
                            "success": False,
                            "error": "Failed to login to D-Tools Cloud. Check DTOOLS credentials.",
                            "parsed_count": len(products),
                            "products": limited,
                        }
                    for prod in limited:
                        out = await agent.create_product(prod)
                        results.append(out)
                        if out.get("success"):
                            created_count += 1
                finally:
                    await agent.stop()
            else:
                results = [{"success": True, "mode": "dry_run", "product": p} for p in limited]

            summary = {
                "success": True,
                "file": str(upload_path),
                "parsed_count": len(products),
                "attempted_count": len(limited),
                "created_count": created_count,
                "failed_count": max(0, len(limited) - created_count) if create_in_dtools and not dry_run else 0,
                "create_in_dtools": bool(create_in_dtools and not dry_run),
                "dealer_tier": dealer_tier,
                "parse_profile": parse_profile,
                "expected_columns": custom_columns or PARSE_PROFILES.get(parse_profile, {}).get("expected_columns", []),
                "results": results,
                "products": limited,
                "timestamp": datetime.now().isoformat(),
            }

            out_file = DTOOLS_PRODUCT_AGENT_DIR / f"product_import_{ts}.json"
            out_file.write_text(json.dumps(summary, indent=2))
            summary["output_file"] = str(out_file)
            return summary
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/dtools/products/approve_store")
    async def dtools_products_approve_store(request: ProductApproveStoreRequest):
        """
        Approve parsed products and persist them to local SQLite product database.
        """
        if not request.products:
            return {"success": False, "error": "No products provided."}

        conn = sqlite3.connect(str(DTOOLS_PRODUCT_DB))
        try:
            _init_product_db(conn)
            now = datetime.now().isoformat()
            total = len(request.products)
            cur = conn.cursor()
            cur.execute(
                """
                INSERT INTO approved_batches (created_at, source_file, parse_profile, dealer_tier, notes, total_products)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    now,
                    request.source_file or "",
                    request.parse_profile or "auto",
                    request.dealer_tier or "standard",
                    request.notes or "",
                    total,
                ),
            )
            batch_id = int(cur.lastrowid)

            keys = [
                _product_unique_key(
                    item.brand or "",
                    item.part_number or "",
                    item.model or "",
                )
                for item in request.products
                if (item.model or "").strip()
            ]
            existing_keys: set[str] = set()
            if keys:
                placeholders = ",".join(["?"] * len(keys))
                rows = conn.execute(
                    f"SELECT unique_key FROM approved_products WHERE unique_key IN ({placeholders})",
                    keys,
                ).fetchall()
                existing_keys = {str(r[0]) for r in rows}

            inserted_count = 0
            updated_count = 0
            saved_count = 0

            for item in request.products:
                model = (item.model or "").strip()
                if not model:
                    continue
                brand = (item.brand or "Unknown").strip()
                part_number = (item.part_number or "").strip()
                key = _product_unique_key(brand, part_number, model)
                if key in existing_keys:
                    updated_count += 1
                else:
                    inserted_count += 1

                conn.execute(
                    """
                    INSERT INTO approved_products (
                        batch_id, unique_key, brand, model, part_number, category, short_description,
                        keywords, unit_price, unit_cost, msrp, supplier, approved_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(unique_key) DO UPDATE SET
                        batch_id=excluded.batch_id,
                        brand=excluded.brand,
                        model=excluded.model,
                        part_number=excluded.part_number,
                        category=excluded.category,
                        short_description=excluded.short_description,
                        keywords=excluded.keywords,
                        unit_price=excluded.unit_price,
                        unit_cost=excluded.unit_cost,
                        msrp=excluded.msrp,
                        supplier=excluded.supplier,
                        approved_at=excluded.approved_at
                    """,
                    (
                        batch_id,
                        key,
                        brand,
                        model,
                        part_number,
                        item.category or "",
                        item.short_description or "",
                        item.keywords or "",
                        item.unit_price,
                        item.unit_cost,
                        item.msrp,
                        item.supplier or "",
                        now,
                    ),
                )
                saved_count += 1

            conn.commit()
            return {
                "success": True,
                "database": str(DTOOLS_PRODUCT_DB),
                "batch_id": batch_id,
                "saved_count": saved_count,
                "inserted_count": inserted_count,
                "updated_count": updated_count,
                "timestamp": now,
            }
        except Exception as e:
            conn.rollback()
            raise HTTPException(status_code=500, detail=str(e))
        finally:
            conn.close()

    @app.post("/dtools/products/retry_create")
    async def dtools_products_retry_create(request: ProductApproveStoreRequest):
        """
        Retry creating selected products in D-Tools after manual fixes.
        """
        if not request.products:
            return {"success": False, "error": "No products provided."}

        try:
            from agents.dtools_browser_agent import DToolsBrowserAgent

            agent = DToolsBrowserAgent(headless=True)
            if not await agent.start():
                return {"success": False, "error": "Failed to start browser agent (Playwright not ready)."}

            created_count = 0
            failed_count = 0
            results: List[Dict[str, Any]] = []
            try:
                if not await asyncio.wait_for(agent.login(), timeout=45):
                    return {"success": False, "error": "Failed to login to D-Tools Cloud. Check credentials."}

                for item in request.products:
                    payload = {
                        "brand": item.brand or "Generic",
                        "model": item.model,
                        "part_number": item.part_number or "",
                        "category": item.category or "Power Management",
                        "short_description": item.short_description or "",
                        "keywords": item.keywords or "",
                        "unit_price": item.unit_price,
                        "unit_cost": item.unit_cost,
                        "msrp": item.msrp,
                        "supplier": item.supplier or "",
                    }
                    out = await agent.create_product(payload)
                    results.append(out)
                    if out.get("success"):
                        created_count += 1
                    else:
                        failed_count += 1
            finally:
                await agent.stop()

            return {
                "success": True,
                "attempted_count": len(request.products),
                "created_count": created_count,
                "failed_count": failed_count,
                "results": results,
                "timestamp": datetime.now().isoformat(),
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.get("/dtools/products/auth_check")
    async def dtools_products_auth_check():
        """
        Quick browser-auth smoke test for D-Tools product agent.
        Helps diagnose selector drift / CAPTCHA / credential issues before import.
        """
        try:
            from agents.dtools_browser_agent import DToolsBrowserAgent

            agent = DToolsBrowserAgent(headless=True)
            if not await agent.start():
                return {"success": False, "error": "Failed to start browser agent (Playwright not ready)."}
            try:
                ok = await asyncio.wait_for(agent.login(), timeout=45)
                return {
                    "success": bool(ok),
                    "message": "Browser login succeeded" if ok else "Browser login failed (credentials/selectors/challenge).",
                    "timestamp": datetime.now().isoformat(),
                }
            except asyncio.TimeoutError:
                return {
                    "success": False,
                    "error": "Browser auth check timed out after 45s.",
                    "timestamp": datetime.now().isoformat(),
                }
            finally:
                await agent.stop()
        except Exception as e:
            return {"success": False, "error": str(e)}

    @app.get("/dtools/products/export")
    async def dtools_products_export(
        batch_id: Optional[int] = None,
        export_format: str = "csv",
        include_all_batches: bool = False,
    ):
        """
        Export approved products from local DB.
        - batch_id: specific batch to export
        - export_format: csv or json
        - include_all_batches: export across all batches (ignores batch_id)
        """
        fmt = (export_format or "csv").strip().lower()
        if fmt not in {"csv", "json"}:
            raise HTTPException(status_code=400, detail="export_format must be 'csv' or 'json'")

        conn = sqlite3.connect(str(DTOOLS_PRODUCT_DB))
        conn.row_factory = sqlite3.Row
        try:
            _init_product_db(conn)
            params: List[Any] = []
            where_clause = ""
            effective_batch: Optional[int] = None

            if include_all_batches:
                where_clause = ""
            elif batch_id is not None:
                where_clause = "WHERE p.batch_id = ?"
                params.append(int(batch_id))
                effective_batch = int(batch_id)
            else:
                row = conn.execute("SELECT id FROM approved_batches ORDER BY id DESC LIMIT 1").fetchone()
                if not row:
                    raise HTTPException(status_code=404, detail="No approved product batches found")
                effective_batch = int(row["id"])
                where_clause = "WHERE p.batch_id = ?"
                params.append(effective_batch)

            rows = conn.execute(
                f"""
                SELECT
                    p.id,
                    p.batch_id,
                    b.created_at AS batch_created_at,
                    b.source_file AS batch_source_file,
                    b.parse_profile AS batch_parse_profile,
                    b.dealer_tier AS batch_dealer_tier,
                    p.brand,
                    p.model,
                    p.part_number,
                    p.category,
                    p.short_description,
                    p.keywords,
                    p.unit_price,
                    p.unit_cost,
                    p.msrp,
                    p.supplier,
                    p.approved_at
                FROM approved_products p
                LEFT JOIN approved_batches b ON b.id = p.batch_id
                {where_clause}
                ORDER BY p.batch_id DESC, p.id ASC
                """,
                params,
            ).fetchall()

            if not rows:
                if effective_batch is not None:
                    raise HTTPException(status_code=404, detail=f"No approved products found for batch_id={effective_batch}")
                raise HTTPException(status_code=404, detail="No approved products found")

            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            batch_label = "all" if include_all_batches else str(effective_batch or "latest")
            base_name = _safe_export_filename(f"dtools_products_batch_{batch_label}_{ts}")
            export_dir = Path(tempfile.gettempdir()) / "symphony_exports"
            export_dir.mkdir(parents=True, exist_ok=True)

            if fmt == "json":
                payload = {
                    "success": True,
                    "timestamp": datetime.now().isoformat(),
                    "batch_id": effective_batch,
                    "include_all_batches": include_all_batches,
                    "count": len(rows),
                    "products": [dict(row) for row in rows],
                }
                out_path = export_dir / f"{base_name}.json"
                out_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
                return FileResponse(
                    path=str(out_path),
                    media_type="application/json",
                    filename=out_path.name,
                )

            # CSV export
            out_path = export_dir / f"{base_name}.csv"
            fieldnames = [
                "id",
                "batch_id",
                "batch_created_at",
                "batch_source_file",
                "batch_parse_profile",
                "batch_dealer_tier",
                "brand",
                "model",
                "part_number",
                "category",
                "short_description",
                "keywords",
                "unit_price",
                "unit_cost",
                "msrp",
                "supplier",
                "approved_at",
            ]
            with out_path.open("w", encoding="utf-8", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                for row in rows:
                    writer.writerow(dict(row))

            return FileResponse(
                path=str(out_path),
                media_type="text/csv",
                filename=out_path.name,
            )
        finally:
            conn.close()
    
    @app.post("/markup/generate")
    async def generate_markup(request: MarkupRequest):
        """Generate AI-powered project markup/proposal."""
        result = run_command([
            "python3", str(BASE_DIR / "tools" / "smart_proposal.py"),
            "--client", request.project_name,
            "--description", request.description,
            "--ai-markup"
        ], timeout=120)
        return result
    
    @app.get("/markup/templates")
    async def get_markup_templates():
        """Get available markup templates."""
        templates = []
        template_dir = BASE_DIR / "knowledge" / "templates"
        if template_dir.exists():
            for f in template_dir.glob("*.json"):
                try:
                    data = json.loads(f.read_text())
                    templates.append({
                        "id": f.stem,
                        "name": data.get("name", f.stem),
                        "description": data.get("description", ""),
                        "rooms": data.get("rooms", [])
                    })
                except:
                    pass
        return {"templates": templates}
    
    @app.get("/markup/exports")
    async def get_markup_exports():
        """Get recent markup exports."""
        exports = []
        export_dir = BASE_DIR / "knowledge" / "markup_exports"
        if export_dir.exists():
            files = sorted(export_dir.glob("*.json"), key=lambda x: x.stat().st_mtime, reverse=True)[:20]
            for f in files:
                try:
                    data = json.loads(f.read_text())
                    exports.append({
                        "id": f.stem,
                        "filename": f.name,
                        "project": data.get("project", f.stem),
                        "symbols_count": len(data.get("symbols", [])),
                        "modified": datetime.fromtimestamp(f.stat().st_mtime).isoformat()
                    })
                except:
                    pass
        return {"exports": exports}
    
    @app.get("/markup/url")
    async def get_markup_url(request: Request):
        """Get URL for Symphony Markup app with safe fallback when HTTPS URL is unreachable."""
        import socket
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            local_ip = s.getsockname()[0]
            s.close()
        except:
            local_ip = "localhost"

        request_host = request.url.hostname or local_ip
        request_scheme = request.url.scheme or "http"
        derived_url = f"{request_scheme}://{request_host}:8091"

        https_url = os.environ.get("MARKUP_HTTPS_URL", "").strip()
        preferred_https = https_url or None
        https_reachable = bool(https_url and is_http_url_reachable(https_url))
        # Use HTTPS only when it is actually reachable; otherwise return the
        # derived host URL to avoid client-side open timeouts.
        active_url = https_url if https_reachable else derived_url
        warning = None
        if https_url and not https_reachable:
            warning = "MARKUP_HTTPS_URL configured but not currently reachable from API host."

        return {
            "url": active_url,
            "httpsUrl": (https_url if https_reachable else None),
            "configuredHttpsUrl": preferred_https,
            "derivedUrl": derived_url,
            "localhost": "http://localhost:8091",
            "status": "running" if is_port_open(8091) else "stopped",
            "warning": warning,
        }


def is_port_open(port: int) -> bool:
    """Check if a port is open."""
    import socket
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    sock.settimeout(1)
    result = sock.connect_ex(('127.0.0.1', port))
    sock.close()
    return result == 0


def is_http_url_reachable(url: str, timeout: int = 3) -> bool:
    """Best-effort URL probe used for MARKUP_HTTPS_URL fallback logic."""
    try:
        req = urllib.request.Request(url, method="GET")
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return 200 <= resp.status < 500
    except Exception:
        return False


# ============================================================================
# Leads Endpoints
# ============================================================================

def run_lead_tool(args: list, timeout: int = 90) -> dict:
    """Run lead_finder or outreach tool."""
    try:
        result = subprocess.run(
            ["python3", str(BASE_DIR / "tools" / "lead_finder.py")] + args,
            capture_output=True, text=True, timeout=timeout
        )
        return {
            "success": result.returncode == 0,
            "output": result.stdout or result.stderr
        }
    except subprocess.TimeoutExpired:
        return {"success": False, "error": "Scan timed out"}
    except Exception as e:
        return {"success": False, "error": str(e)}


def run_outreach_tool(args: list, timeout: int = 60) -> dict:
    """Run outreach automation tool."""
    try:
        result = subprocess.run(
            ["python3", str(BASE_DIR / "tools" / "outreach_automation.py")] + args,
            capture_output=True, text=True, timeout=timeout
        )
        return {
            "success": result.returncode == 0,
            "output": result.stdout or result.stderr
        }
    except subprocess.TimeoutExpired:
        return {"success": False, "error": "Operation timed out"}
    except Exception as e:
        return {"success": False, "error": str(e)}


if HAS_FASTAPI:
    @app.get("/leads/builders")
    async def scan_builders():
        """Scan for custom home builders."""
        return run_lead_tool(["--builders"])
    
    @app.get("/leads/realtors")
    async def scan_realtors():
        """Scan for luxury realtors."""
        return run_lead_tool(["--realtors"])
    
    @app.get("/leads/listings")
    async def scan_listings():
        """Scan for luxury listings."""
        return run_lead_tool(["--listings"])
    
    @app.get("/leads/property")
    async def scan_property_managers():
        """Scan for property management companies."""
        return run_lead_tool(["--property-managers"])
    
    @app.get("/leads/recent")
    async def get_recent_leads():
        """Get recently scanned leads."""
        return run_lead_tool(["--recent"])
    
    @app.get("/leads/outreach/queue")
    async def get_outreach_queue():
        """Get current outreach queue."""
        return run_outreach_tool(["--queue"])
    
    @app.get("/leads/outreach/generate")
    async def generate_outreach():
        """Generate outreach drafts from leads."""
        return run_outreach_tool(["--run"])
    
    # ========================================================================
    # AI Chat Endpoints
    # ========================================================================
    
    @app.get("/ai/status")
    async def ai_status():
        """Which AI backends are available (cortex, ollama, lm_studio, openai, perplexity)."""
        try:
            import sys
            sys.path.insert(0, str(BASE_DIR / "tools"))
            from ai_router import get_backend_status
            return get_backend_status()
        except Exception as e:
            return {"error": str(e)}

    @app.get("/ai/verify/ollama")
    async def ai_verify_ollama():
        """Verify Ollama (Betty) is reachable. Returns ok + message."""
        try:
            import sys
            sys.path.insert(0, str(BASE_DIR / "tools"))
            from ai_router import verify_backend, OLLAMA_URL
            ok, msg = verify_backend(OLLAMA_URL, "/api/tags", "Ollama (Betty)")
            return {"ok": ok, "message": msg}
        except Exception as e:
            return {"ok": False, "message": str(e)}

    @app.get("/ai/verify/lm_studio")
    async def ai_verify_lm_studio():
        """Verify LM Studio is reachable. Returns ok + message."""
        try:
            import sys
            sys.path.insert(0, str(BASE_DIR / "tools"))
            from ai_router import verify_backend, LM_STUDIO_URL
            ok, msg = verify_backend(LM_STUDIO_URL, "/v1/models", "LM Studio")
            return {"ok": ok, "message": msg}
        except Exception as e:
            return {"ok": False, "message": str(e)}

    class ChatRequest(BaseModel):
        question: str
        source: Optional[str] = "auto"
        session_id: Optional[str] = "default"

    @app.post("/ai/chat")
    async def ai_chat(body: ChatRequest):
        """
        Smart AI routing. Optional "source" to force: auto, cortex, ollama, lm_studio,
        gpt-4o-mini, perplexity.
        """
        try:
            question = body.question or ""
            source = body.source or "auto"
            session_id = body.session_id or "default"

            if not question:
                return {"success": False, "error": "No question provided"}
            
            import sys
            sys.path.insert(0, str(BASE_DIR / "tools"))
            from ai_router import ask, classify_question
            
            # Build retained memory context from recent Ask Bob exchanges.
            recent_memory = _load_recent_ask_bob_memory(session_id=session_id, limit=8)
            memory_lines: List[str] = []
            for item in recent_memory:
                q = (item.get("question", "") or "").strip()
                a = (item.get("answer", "") or "").strip()
                if q and a:
                    memory_lines.append(f"Q: {q}\nA: {a[:500]}")

            project_hint = _extract_project_hint(question)
            project_ctx = _collect_project_file_context(project_hint, question) if project_hint else {
                "project_hint": "",
                "files": [],
                "context": "",
            }

            augment_blocks: List[str] = []
            if memory_lines:
                augment_blocks.append(
                    "Recent conversation memory:\n" + "\n\n".join(memory_lines[-6:])
                )
            if project_ctx.get("context"):
                augment_blocks.append(
                    "Project file context:\n" + project_ctx["context"]
                )

            augmented_question = question
            if augment_blocks:
                augmented_question = (
                    f"{question}\n\n"
                    "Use the context below to answer accurately. "
                    "If context is insufficient, say what is missing.\n\n"
                    + "\n\n".join(augment_blocks)
                )

            complexity = classify_question(question)
            answer, used_source, cost = ask(augmented_question, source=source)

            files_scanned = project_ctx.get("files", [])
            if files_scanned:
                top_files = files_scanned[:4]
                more_count = max(0, len(files_scanned) - len(top_files))
                footer = "Sources: " + ", ".join(top_files)
                if more_count > 0:
                    footer += f", +{more_count} more"
                if footer not in answer:
                    answer = f"{answer}\n\n{footer}"

            # Persist memory for future asks.
            _append_ask_bob_memory(
                session_id=session_id,
                question=question,
                answer=answer,
                source=used_source,
            )
            
            return {
                "success": True,
                "output": answer,
                "source": used_source,
                "complexity": complexity,
                "cost_usd": cost,
                "session_id": session_id,
                "memory_used": bool(memory_lines),
                "project_context_used": bool(project_ctx.get("context")),
                "project_hint": project_ctx.get("project_hint", ""),
                "project_files_scanned": files_scanned,
            }
        except Exception as e:
            import traceback
            return {"success": False, "error": str(e), "trace": traceback.format_exc()}
    
    @app.get("/ai/costs")
    async def get_ai_costs():
        """Get AI usage cost summary."""
        try:
            import sys
            sys.path.insert(0, str(BASE_DIR / "tools"))
            from ai_router import get_cost_summary
            return get_cost_summary()
        except Exception as e:
            return {"error": str(e)}
    
    @app.post("/ai/log")
    async def log_ai_query(request: Request):
        """Log AI queries for knowledge building."""
        try:
            body = await request.json()
            question = body.get("question", "")
            source = body.get("source", "unknown")
            
            log_file = BASE_DIR / "data" / "ai_query_log.jsonl"
            log_file.parent.mkdir(parents=True, exist_ok=True)
            
            import json
            with open(log_file, "a") as f:
                f.write(json.dumps({
                    "timestamp": datetime.now().isoformat(),
                    "question": question,
                    "source": source
                }) + "\n")
            
            return {"success": True, "logged": True}
        except Exception as e:
            return {"success": False, "error": str(e)}

    # ─────────────────────────────────────────────────────────────
    # Claude Approval (Bridge: Task Board → Email → iOS Approve → Bob)
    # ─────────────────────────────────────────────────────────────
    
    class AddTaskRequest(BaseModel):
        title: str
        description: Optional[str] = ""
        task_type: Optional[str] = "research"
        priority: Optional[str] = "medium"

    class ProjectWatchRequest(BaseModel):
        project_name: str = ""
        client_name: str = ""
        address_line: str = ""
        location_name: str = ""
        folder_path: str
    
    @app.post("/tasks")
    async def add_task(request: AddTaskRequest):
        """Add a task to the board. Use task_type='claude' for approval workflow."""
        try:
            sys.path.insert(0, str(BASE_DIR))
            from orchestrator.task_board import add_task as tb_add_task
            task_id = tb_add_task(
                title=request.title,
                description=request.description or "",
                task_type=request.task_type or "research",
                priority=request.priority or "medium",
            )
            return {"success": True, "task_id": task_id}
        except Exception as e:
            return {"success": False, "error": str(e)}

    @app.post("/tasks/upload_intake")
    async def upload_intake_task(
        category: str = Form("document"),
        project_name: str = Form(""),
        client_name: str = Form(""),
        address_line: str = Form(""),
        location_name: str = Form(""),
        discipline: str = Form(""),
        sheet_number: str = Form(""),
        revision: str = Form(""),
        issue_date: str = Form(""),
        title: str = Form(""),
        description: str = Form(""),
        priority: str = Form("high"),
        file: UploadFile = File(...),
    ):
        """
        Upload a proposal/drawing/image/document, store it with a strict naming scheme,
        and create a categorized task for team processing.
        """
        try:
            allowed = {"proposal", "drawing", "image", "document"}
            category_norm = (category or "document").strip().lower()
            if category_norm not in allowed:
                category_norm = "document"

            source_filename = (file.filename or "upload.bin").strip()
            upload_id = uuid4().hex[:12]
            stored_filename = _build_intake_filename(
                category=category_norm,
                project_name=project_name,
                client_name=client_name,
                original_filename=source_filename,
                address_line=address_line,
                location_name=location_name,
                discipline=discipline,
                sheet_number=sheet_number,
                revision=revision,
                issue_date=issue_date,
            )

            day_dir = datetime.now().strftime("%Y-%m-%d")
            target_dir = TASK_UPLOAD_DIR / category_norm / day_dir
            target_dir.mkdir(parents=True, exist_ok=True)
            stored_path = target_dir / stored_filename

            data = await file.read()
            if not data:
                return {"success": False, "error": "Uploaded file was empty"}
            stored_path.write_bytes(data)

            # Attempt to detect D-Tools quote/version tags from title-block text.
            extracted_text = ""
            try:
                suffix = stored_path.suffix.lower()
                if suffix == ".dwg":
                    extracted_text = _extract_text_from_dwg(stored_path)
                elif suffix in {".png", ".jpg", ".jpeg", ".webp"}:
                    extracted_text = _ocr_titleblock_qv_from_image(stored_path)
                else:
                    extracted_text = _extract_text_from_path(stored_path)
            except Exception:
                extracted_text = ""
            dtools_qv = _extract_dtools_qv(extracted_text, fallback_text=f"{source_filename} {stored_filename}")

            default_title = f"Review {category_norm}: {project_name or client_name or source_filename}"
            task_title = (title or default_title).strip()

            task_type_map = {
                "proposal": "proposal",
                "drawing": "documentation",
                "image": "documentation",
                "document": "documentation",
            }
            task_type = task_type_map.get(category_norm, "documentation")

            desc_parts = [
                description.strip() if description else "",
                f"Category: {category_norm}",
                f"Project: {project_name or 'n/a'}",
                f"Client: {client_name or 'n/a'}",
                f"Address: {address_line or 'n/a'}",
                f"Location: {location_name or 'n/a'}",
                f"Discipline: {discipline or 'n/a'}",
                f"Sheet: {sheet_number or 'n/a'}",
                f"Revision: {revision or 'n/a'}",
                f"Issue date: {issue_date or 'n/a'}",
                f"D-Tools quote: {dtools_qv.get('quote_id') or 'n/a'}",
                f"D-Tools version: {dtools_qv.get('version') or 'n/a'}",
                f"Original filename: {source_filename}",
                f"Stored path: {stored_path}",
                "Naming scheme: YYYYMMDD-HHMMSS--category--project--client--original.ext",
                "Naming scheme v2: <ClientLast>-<AddressNumberStreet>-<Location> + metadata suffix",
            ]
            task_description = "\n".join([x for x in desc_parts if x])

            sys.path.insert(0, str(BASE_DIR))
            from orchestrator.task_board import add_task as tb_add_task

            task_id = tb_add_task(
                title=task_title,
                description=task_description,
                task_type=task_type,
                priority=(priority or "high"),
            )

            extraction_queued = False
            findings_path: Optional[Path] = None
            findings_preview: Dict[str, Any] = {}
            if category_norm == "drawing":
                queue_payload = {
                    "upload_id": upload_id,
                    "task_id": task_id,
                    "file_path": str(stored_path),
                    "file_name": stored_filename,
                    "project_name": project_name,
                    "client_name": client_name,
                    "address_line": address_line,
                    "location_name": location_name,
                    "discipline": discipline,
                    "sheet_number": sheet_number,
                    "revision": revision,
                    "issue_date": issue_date,
                    "dtools_quote_id": dtools_qv.get("quote_id", ""),
                    "dtools_version": dtools_qv.get("version", ""),
                    "dtools_quote_version": dtools_qv.get("quote_version", ""),
                    "status": "queued",
                    "queued_at": datetime.now().isoformat(),
                }
                queue_file = TASK_UPLOAD_EXTRACTION_QUEUE_DIR / f"{upload_id}.json"
                queue_file.write_text(json.dumps(queue_payload, indent=2), encoding="utf-8")
                extraction_queued = True
                try:
                    findings = _first_pass_drawing_findings(
                        stored_path=stored_path,
                        project_name=project_name,
                        client_name=client_name,
                    )
                    findings["upload_id"] = upload_id
                    findings["task_id"] = task_id
                    findings["metadata"] = {
                        "discipline": discipline,
                        "sheet_number": sheet_number,
                        "revision": revision,
                        "issue_date": issue_date,
                    }
                    findings_path = TASK_UPLOAD_FINDINGS_DIR / f"{upload_id}__findings.json"
                    findings_path.write_text(json.dumps(findings, indent=2), encoding="utf-8")
                    findings_preview = {
                        "legend_terms_count": len(findings.get("legend_terms", [])),
                        "symbol_matches_count": len(findings.get("symbol_matches", [])),
                        "sheet_references_count": len(findings.get("sheet_references", [])),
                    }
                    queue_payload["status"] = "first_pass_complete"
                    queue_payload["findings_path"] = str(findings_path)
                    queue_file.write_text(json.dumps(queue_payload, indent=2), encoding="utf-8")
                except Exception as extract_error:
                    queue_payload["status"] = "first_pass_failed"
                    queue_payload["error"] = str(extract_error)
                    queue_file.write_text(json.dumps(queue_payload, indent=2), encoding="utf-8")

            queue_record = {
                "upload_id": upload_id,
                "created_at": datetime.now().isoformat(),
                "category": category_norm,
                "project_name": project_name,
                "client_name": client_name,
                "address_line": address_line,
                "location_name": location_name,
                "discipline": discipline,
                "sheet_number": sheet_number,
                "revision": revision,
                "issue_date": issue_date,
                "dtools_quote_id": dtools_qv.get("quote_id", ""),
                "dtools_version": dtools_qv.get("version", ""),
                "dtools_quote_version": dtools_qv.get("quote_version", ""),
                "source_filename": source_filename,
                "stored_filename": stored_filename,
                "stored_path": str(stored_path),
                "task_id": task_id,
            }
            if findings_path:
                queue_record["findings_path"] = str(findings_path)
            _append_upload_queue_record(queue_record)

            return {
                "success": True,
                "upload_id": upload_id,
                "task_id": task_id,
                "category": category_norm,
                "stored_filename": stored_filename,
                "stored_path": str(stored_path),
                "naming_scheme": "YYYYMMDD-HHMMSS--category--project--client--original.ext",
                "naming_scheme_v2": "<ClientLast>-<AddressNumberStreet>-<Location>--<Category>--<Discipline>--<Sheet>--r-<Revision>--<IssueDate>--<Original>.ext",
                "extraction_queued": extraction_queued,
                "findings_path": str(findings_path) if findings_path else "",
                "findings_preview": findings_preview,
                "dtools_quote_id": dtools_qv.get("quote_id", ""),
                "dtools_version": dtools_qv.get("version", ""),
                "dtools_quote_version": dtools_qv.get("quote_version", ""),
            }
        except Exception as e:
            return {"success": False, "error": str(e)}

    @app.post("/tasks/upload_project_bundle")
    async def upload_project_bundle(
        project_name: str = Form(""),
        client_name: str = Form(""),
        address_line: str = Form(""),
        location_name: str = Form(""),
        source_folder_path: str = Form(""),
        enable_watch: str = Form("true"),
        priority: str = Form("high"),
        bundle: UploadFile = File(...),
    ):
        """
        Upload a full project ZIP, explode it into intake files, and create one
        synthesis task so the team can connect RFIs + client wants/needs across files.
        """
        try:
            source_filename = (bundle.filename or "project_bundle.zip").strip()
            if Path(source_filename).suffix.lower() != ".zip":
                return {"success": False, "error": "Project bundle must be a .zip file"}

            bundle_id = uuid4().hex[:12]
            day_dir = datetime.now().strftime("%Y-%m-%d")
            project_hint = _safe_filename_segment(project_name or client_name, fallback="Project", max_len=48)
            bundle_root = TASK_UPLOAD_DIR / "project_bundles" / day_dir / f"{project_hint}--{bundle_id}"
            bundle_root.mkdir(parents=True, exist_ok=True)
            zip_path = bundle_root / source_filename

            payload = await bundle.read()
            if not payload:
                return {"success": False, "error": "Uploaded ZIP was empty"}
            zip_path.write_bytes(payload)

            extracted_files = _safe_extract_zip(zip_path, bundle_root / "expanded", max_files=250)
            if not extracted_files:
                return {"success": False, "error": "ZIP had no supported files to process"}

            counts: Dict[str, int] = {"drawing": 0, "proposal": 0, "image": 0, "document": 0}
            for path in extracted_files:
                counts[_categorize_bundle_file(path.name)] += 1

            sys.path.insert(0, str(BASE_DIR))
            from orchestrator.task_board import add_task as tb_add_task

            task_title = f"Synthesize project intake: {project_name or client_name or project_hint}"
            task_description = "\n".join(
                [
                    "Project bundle intake for cross-file synthesis.",
                    f"Bundle ID: {bundle_id}",
                    f"Project: {project_name or 'n/a'}",
                    f"Client: {client_name or 'n/a'}",
                    f"Address: {address_line or 'n/a'}",
                    f"Location: {location_name or 'n/a'}",
                    f"Source ZIP: {zip_path}",
                    f"File count: {len(extracted_files)}",
                    f"Category counts: {json.dumps(counts)}",
                    "Goal: connect RFIs + client wants/needs across all uploaded files.",
                ]
            )
            task_id = tb_add_task(
                title=task_title,
                description=task_description,
                task_type="documentation",
                priority=(priority or "high"),
            )

            drawing_findings_count = 0
            for source_path in extracted_files:
                category_norm = _categorize_bundle_file(source_path.name)
                stored_filename = _build_intake_filename(
                    category=category_norm,
                    project_name=project_name,
                    client_name=client_name,
                    original_filename=source_path.name,
                    address_line=address_line,
                    location_name=location_name,
                )
                target_dir = TASK_UPLOAD_DIR / category_norm / day_dir
                target_dir.mkdir(parents=True, exist_ok=True)
                stored_path = target_dir / stored_filename
                shutil.copy2(source_path, stored_path)

                queue_record = {
                    "upload_id": uuid4().hex[:12],
                    "bundle_id": bundle_id,
                    "created_at": datetime.now().isoformat(),
                    "category": category_norm,
                    "project_name": project_name,
                    "client_name": client_name,
                    "address_line": address_line,
                    "location_name": location_name,
                    "source_filename": source_path.name,
                    "stored_filename": stored_filename,
                    "stored_path": str(stored_path),
                    "task_id": task_id,
                }

                if category_norm == "drawing":
                    qv = _extract_dtools_qv(text=stored_filename, fallback_text=source_path.name)
                    queue_payload = {
                        "upload_id": queue_record["upload_id"],
                        "bundle_id": bundle_id,
                        "task_id": task_id,
                        "file_path": str(stored_path),
                        "file_name": stored_filename,
                        "project_name": project_name,
                        "client_name": client_name,
                        "address_line": address_line,
                        "location_name": location_name,
                        "dtools_quote_id": qv.get("quote_id", ""),
                        "dtools_version": qv.get("version", ""),
                        "dtools_quote_version": qv.get("quote_version", ""),
                        "status": "queued",
                        "queued_at": datetime.now().isoformat(),
                    }
                    queue_file = TASK_UPLOAD_EXTRACTION_QUEUE_DIR / f"{queue_record['upload_id']}.json"
                    queue_file.write_text(json.dumps(queue_payload, indent=2), encoding="utf-8")
                    try:
                        findings = _first_pass_drawing_findings(
                            stored_path=stored_path,
                            project_name=project_name,
                            client_name=client_name,
                        )
                        findings["upload_id"] = queue_record["upload_id"]
                        findings["bundle_id"] = bundle_id
                        findings["task_id"] = task_id
                        findings_path = TASK_UPLOAD_FINDINGS_DIR / f"{queue_record['upload_id']}__findings.json"
                        findings_path.write_text(json.dumps(findings, indent=2), encoding="utf-8")
                        queue_record["findings_path"] = str(findings_path)
                        queue_payload["status"] = "first_pass_complete"
                        queue_payload["findings_path"] = str(findings_path)
                        queue_file.write_text(json.dumps(queue_payload, indent=2), encoding="utf-8")
                        drawing_findings_count += 1
                    except Exception as extract_error:
                        queue_payload["status"] = "first_pass_failed"
                        queue_payload["error"] = str(extract_error)
                        queue_file.write_text(json.dumps(queue_payload, indent=2), encoding="utf-8")

                _append_upload_queue_record(queue_record)

            watch_registered = False
            watch_info: Dict[str, Any] = {}
            if source_folder_path.strip() and enable_watch.strip().lower() not in {"0", "false", "no"}:
                try:
                    watch_info = _register_project_watch(
                        project_name=project_name,
                        client_name=client_name,
                        address_line=address_line,
                        location_name=location_name,
                        folder_path=source_folder_path.strip(),
                    )
                    watch_registered = True
                except Exception:
                    watch_registered = False

            return {
                "success": True,
                "bundle_id": bundle_id,
                "task_id": task_id,
                "zip_path": str(zip_path),
                "extracted_count": len(extracted_files),
                "category_counts": counts,
                "drawing_findings_count": drawing_findings_count,
                "watch_registered": watch_registered,
                "watch": watch_info,
                "message": "Project bundle uploaded and queued for cross-file synthesis.",
            }
        except zipfile.BadZipFile:
            return {"success": False, "error": "Invalid ZIP file"}
        except Exception as e:
            return {"success": False, "error": str(e)}

    @app.get("/tasks/uploads_queue")
    async def get_uploads_queue(limit: int = 10):
        """
        Return most-recent uploaded intake files with linked task status.
        `open` = pending/in_progress/blocked, `complete` = completed.
        """
        try:
            if not TASK_UPLOAD_QUEUE_FILE.exists():
                return {"success": True, "uploads": []}
            raw_lines = TASK_UPLOAD_QUEUE_FILE.read_text(encoding="utf-8", errors="ignore").splitlines()
            rows: List[Dict[str, Any]] = []
            for line in raw_lines:
                line = line.strip()
                if not line:
                    continue
                try:
                    rows.append(json.loads(line))
                except Exception:
                    continue
            rows = rows[-max(1, min(limit, 50)) :]
            rows.reverse()

            uploads: List[Dict[str, Any]] = []
            for row in rows:
                task_id = row.get("task_id")
                task_status = _task_status(task_id)
                open_complete = "open"
                if task_status == "completed":
                    open_complete = "complete"
                elif task_status in {"cancelled", "failed"}:
                    open_complete = "closed"
                uploads.append(
                    {
                        "upload_id": row.get("upload_id"),
                        "created_at": row.get("created_at"),
                        "category": row.get("category", "document"),
                        "project_name": row.get("project_name", ""),
                        "client_name": row.get("client_name", ""),
                        "address_line": row.get("address_line", ""),
                        "location_name": row.get("location_name", ""),
                        "stored_filename": row.get("stored_filename", ""),
                        "task_id": task_id,
                        "task_status": task_status,
                        "open_complete_status": open_complete,
                        "findings_path": row.get("findings_path", ""),
                    }
                )
            return {"success": True, "uploads": uploads}
        except Exception as e:
            return {"success": False, "uploads": [], "error": str(e)}

    @app.get("/tasks/project_watches")
    async def get_project_watches():
        try:
            watches = _load_project_watches()
            state = _load_project_watch_state()
            enriched = []
            for w in watches:
                st = state.get(w.get("watch_id", ""), {})
                summary_path = ""
                project_slug = w.get("project_slug", "")
                if project_slug:
                    p = KNOWLEDGE_PROJECTS_DIR / project_slug / "project_intelligence_summary.json"
                    if p.exists():
                        summary_path = str(p)
                enriched.append(
                    {
                        **w,
                        "last_scan_at": st.get("last_scan_at"),
                        "last_processed_count": st.get("last_processed_count", 0),
                        "last_candidate_count": st.get("last_candidate_count", 0),
                        "last_skipped_too_large": st.get("last_skipped_too_large", 0),
                        "last_skipped_ignored": st.get("last_skipped_ignored", 0),
                        "scan_duration_sec": st.get("scan_duration_sec", 0.0),
                        "scan_in_progress": bool(st.get("scan_in_progress", False)),
                        "last_error": st.get("last_error", ""),
                        "summary_path": summary_path,
                    }
                )
            return {
                "success": True,
                "watches": enriched,
                "auto_discover_enabled": PROJECT_WATCH_AUTO_DISCOVER,
                "projects_root": PROJECT_WATCH_PROJECTS_ROOT,
            }
        except Exception as e:
            return {"success": False, "watches": [], "error": str(e)}

    @app.post("/tasks/project_watches/register")
    async def register_project_watch(request: ProjectWatchRequest):
        try:
            watch = _register_project_watch(
                project_name=request.project_name,
                client_name=request.client_name,
                address_line=request.address_line,
                location_name=request.location_name,
                folder_path=request.folder_path,
            )
            scan_summary = _scan_project_watch(watch, max_new_files=60)
            return {"success": True, "watch": watch, "initial_scan": scan_summary}
        except Exception as e:
            return {"success": False, "error": str(e)}

    @app.post("/tasks/project_watches/run")
    async def run_project_watches(watch_id: str = Form("")):
        """
        Run one scan pass:
        - all enabled watches when watch_id is empty
        - one specific watch when watch_id is provided
        """
        try:
            watches = _load_project_watches()
            selected = [w for w in watches if w.get("enabled", True)]
            if watch_id.strip():
                selected = [w for w in selected if w.get("watch_id") == watch_id.strip()]
            results = [_scan_project_watch(w, max_new_files=60) for w in selected]
            total = sum(int(r.get("processed", 0) or 0) for r in results)
            return {"success": True, "ran": len(results), "processed_total": total, "results": results}
        except Exception as e:
            return {"success": False, "error": str(e)}

    @app.post("/tasks/project_watches/discover")
    async def discover_project_watches():
        """
        Auto-register any new project folders found under the configured Projects root.
        """
        try:
            summary = _auto_discover_project_watches(max_new=50)
            return {"success": True, **summary}
        except Exception as e:
            return {"success": False, "error": str(e)}

    @app.post("/tasks/project_watches/backfill_recent")
    async def backfill_recent_project_watch_files(hours: int = Form(24), max_files: int = Form(250)):
        """
        One-time catch-up ingestion for recently added files that may predate watch state.
        Intended for manual bursts where many manuals were added at once.
        """
        try:
            summary = _backfill_recent_watch_files(hours=hours, max_files=max_files)
            return {"success": True, **summary}
        except Exception as e:
            return {"success": False, "error": str(e)}

    @app.get("/tasks/project_summary")
    async def get_project_summary(project_slug: str = ""):
        try:
            slug = _slugify(project_slug, fallback="project", max_len=80)
            path = KNOWLEDGE_PROJECTS_DIR / slug / "project_intelligence_summary.json"
            if not path.exists():
                _update_project_intelligence_summary(slug)
            if not path.exists():
                return {"success": False, "error": "Project summary not found", "project_slug": slug}
            data = json.loads(path.read_text(encoding="utf-8"))
            return {"success": True, "project_slug": slug, "summary": data}
        except Exception as e:
            return {"success": False, "error": str(e)}

    @app.post("/manuals/auto_digest/run")
    async def manuals_auto_digest_run(max_new_files: int = Form(25)):
        try:
            return _manuals_auto_digest_once(max_new_files=max_new_files)
        except Exception as e:
            return {"success": False, "error": str(e)}

    @app.get("/manuals/auto_digest/status")
    async def manuals_auto_digest_status():
        try:
            state = _load_manuals_auto_digest_state()
            return {
                "success": True,
                "enabled": MANUALS_AUTO_DIGEST_ENABLED,
                "root": str(MANUALS_LIBRARY_ROOT),
                "knowledge_file": str(MANUALS_AUTO_DIGEST_KNOWLEDGE_FILE),
                "state": state,
            }
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    @app.get("/tasks/claude_pending")
    async def get_claude_pending():
        """Get Claude tasks awaiting approval (type=claude, status=pending)."""
        try:
            sys.path.insert(0, str(BASE_DIR))
            from orchestrator.task_board import get_claude_pending_tasks
            tasks = get_claude_pending_tasks()
            return {
                "tasks": [
                    {
                        "id": t.id,
                        "title": t.title,
                        "description": t.description or "",
                        "priority": t.priority,
                        "created_at": t.created_at,
                    }
                    for t in tasks
                ]
            }
        except Exception as e:
            return {"tasks": [], "error": str(e)}

    @app.get("/tasks/incidents")
    async def get_incident_queue(limit: int = 20):
        """
        Get high-priority troubleshooting tasks for quick mobile triage.
        Includes pending, in_progress, and blocked incidents.
        """
        try:
            sys.path.insert(0, str(BASE_DIR))
            from orchestrator.task_board import list_tasks

            all_tasks = list_tasks(status=None, limit=max(50, min(limit * 6, 500)))
            incidents = []
            for t in all_tasks:
                if t.task_type != "troubleshooting":
                    continue
                if t.priority not in {"critical", "high"}:
                    continue
                if t.status not in {"pending", "in_progress", "blocked"}:
                    continue
                incidents.append(
                    {
                        "id": t.id,
                        "title": t.title,
                        "description": t.description or "",
                        "priority": t.priority,
                        "status": t.status,
                        "assigned_to": t.assigned_to,
                        "created_at": t.created_at,
                        "updated_at": t.updated_at,
                    }
                )

            priority_rank = {"critical": 0, "high": 1}
            status_rank = {"in_progress": 0, "blocked": 1, "pending": 2}
            incidents.sort(
                key=lambda x: (
                    priority_rank.get(x.get("priority", "high"), 9),
                    status_rank.get(x.get("status", "pending"), 9),
                    x.get("created_at", ""),
                )
            )
            incidents = incidents[: max(1, min(limit, 100))]

            return {
                "success": True,
                "count": len(incidents),
                "incidents": incidents,
                "timestamp": datetime.now().isoformat(),
            }
        except Exception as e:
            return {"success": False, "count": 0, "incidents": [], "error": str(e)}
    
    @app.post("/tasks/{task_id}/approve_claude")
    async def approve_claude(task_id: int):
        """Approve a Claude task: Bob sends command to Claude Code via terminal."""
        try:
            sys.path.insert(0, str(BASE_DIR))
            from orchestrator.task_board import approve_claude_task
            success, message = approve_claude_task(task_id)
            return {"success": success, "message": message}
        except Exception as e:
            return {"success": False, "message": str(e)}

    @app.get("/claude/workflows")
    async def get_claude_workflows():
        """Get Claude Code workflow prompts for copy/paste."""
        try:
            path = BASE_DIR / "setup" / "claude_code" / "workflow_prompts.json"
            if not path.exists():
                return {"workflows": []}
            data = json.loads(path.read_text())
            return {"workflows": data}
        except Exception as e:
            return {"workflows": [], "error": str(e)}

    # ─────────────────────────────────────────────────────────────
    # Social / X (Twitter) — same as Telegram SEO menu
    # ─────────────────────────────────────────────────────────────
    
    @app.get("/social/story")
    async def social_story():
        """Generate project story tweet and queue it."""
        return run_command(
            ["python3", str(BASE_DIR / "tools" / "social_content.py"), "--story", "--queue"],
            timeout=45
        )
    
    @app.get("/social/tip")
    async def social_tip():
        """Generate daily tip tweet and queue it."""
        return run_command(
            ["python3", str(BASE_DIR / "tools" / "social_content.py"), "--tip", "--queue"],
            timeout=45
        )
    
    @app.get("/social/video")
    async def social_video():
        """Generate video prompt + tweet and queue it."""
        return run_command(
            ["python3", str(BASE_DIR / "tools" / "social_content.py"), "--video-prompt", "--queue"],
            timeout=45
        )
    
    @app.get("/social/week")
    async def social_week():
        """Generate full week of content (takes ~60 sec)."""
        return run_command(
            ["python3", str(BASE_DIR / "tools" / "social_content.py"), "--series", "--queue"],
            timeout=120
        )
    
    @app.get("/social/x-queue")
    async def social_x_queue():
        """Show X post queue."""
        return run_command(
            ["python3", str(BASE_DIR / "tools" / "x_poster.py"), "--queue"],
            timeout=15
        )
    
    @app.get("/social/x-post")
    async def social_x_post():
        """Post next tweet from queue to @symphonysmart."""
        return run_command(
            ["python3", str(BASE_DIR / "tools" / "x_poster.py"), "--auto"],
            timeout=30
        )
    
    @app.get("/social/x-usage")
    async def social_x_usage():
        """Show X API usage this month."""
        return run_command(
            ["python3", str(BASE_DIR / "tools" / "x_poster.py"), "--usage"],
            timeout=15
        )
    
    # ─────────────────────────────────────────────────────────────
    # SEO Endpoints
    # ─────────────────────────────────────────────────────────────
    
    @app.get("/seo/keywords")
    async def seo_keywords():
        """Research SEO keywords for Vail Valley."""
        return run_tool_endpoint("seo_manager.py", ["--keywords"])
    
    @app.get("/seo/content")
    async def seo_content_ideas():
        """Get content/blog ideas."""
        return run_tool_endpoint("seo_manager.py", ["--content"])
    
    @app.get("/seo/local")
    async def seo_local_audit():
        """Run local SEO audit."""
        return run_tool_endpoint("seo_manager.py", ["--local"])
    
    @app.get("/seo/backlinks")
    async def seo_backlinks():
        """Find backlink opportunities."""
        return run_tool_endpoint("seo_manager.py", ["--backlinks"])
    
    @app.get("/seo/meta")
    async def seo_meta_tags():
        """Generate optimized meta tags."""
        return run_tool_endpoint("seo_manager.py", ["--meta"])
    
    @app.post("/seo/generate")
    async def seo_generate_post():
        """Generate a new blog post draft."""
        result = subprocess.run(
            ["python3", str(BASE_DIR / "orchestrator" / "seo_content_generator.py"), "--generate"],
            capture_output=True, text=True, timeout=120
        )
        return {"success": result.returncode == 0, "output": result.stdout or result.stderr}
    
    @app.get("/seo/drafts")
    async def seo_list_drafts():
        """List all SEO content drafts."""
        drafts_dir = BASE_DIR / "knowledge" / "seo" / "drafts"
        if not drafts_dir.exists():
            return {"success": True, "drafts": []}
        
        drafts = []
        for f in sorted(drafts_dir.glob("*.json"), reverse=True)[:10]:
            try:
                import json
                data = json.loads(f.read_text())
                drafts.append({
                    "file": f.name,
                    "title": data.get("title"),
                    "keyword": data.get("keyword"),
                    "status": data.get("status", "draft"),
                    "generated": data.get("generated")
                })
            except:
                pass
        
        return {"success": True, "drafts": drafts}


def main():
    if not HAS_FASTAPI:
        print("Install FastAPI: pip install fastapi uvicorn")
        return
    
    print(f"""
╔══════════════════════════════════════════════════╗
║     Symphony AI Mobile API                       ║
║     http://localhost:{API_PORT}                         ║
╠══════════════════════════════════════════════════╣
║  Endpoints:                                      ║
║    GET  /              - Health check            ║
║    GET  /dashboard     - Main dashboard          ║
║    GET  /services      - Service status          ║
║    GET  /bids          - Check bids              ║
║    GET  /proposals     - List proposals          ║
║    POST /research      - Search knowledge        ║
║    POST /cortex/curator/run - Curate cortex      ║
║    GET  /cortex/curator/status - Curator stats   ║
║    GET  /cortex/curator/review - Review queue    ║
║    GET  /website/status - Website health         ║
║    GET  /subscriptions - List subscriptions      ║
║    GET  /morning       - Morning checklist       ║
╚══════════════════════════════════════════════════╝
""")
    
    uvicorn.run(app, host=API_BIND_HOST, port=API_PORT)


if __name__ == "__main__":
    main()
